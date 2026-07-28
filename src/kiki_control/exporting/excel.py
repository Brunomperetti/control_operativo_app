"""Generación pura en memoria de reportes Excel auditables."""

from __future__ import annotations

from collections.abc import Iterable
from datetime import datetime, time
from decimal import Decimal
from io import BytesIO
from typing import Any, Literal
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from kiki_control.domain.control_consolidado import ReporteControlConsolidado, ResultadoControlConsolidado
from kiki_control.domain.reconciliation import ReporteConciliacion, ResultadoConciliacion
from kiki_control.domain.ml_eccomapp_diagnostic import DiagnosticoMlEccomapp, EstadoCruceMlEccomapp
from kiki_control.presentation.bloque_b_diagnostics import (
    CategoriaPrincipalMpSinVenta,
    DiagnosticoBloqueB,
    ESTADOS_EXPLICACION_VISIBLES,
    GrupoConDiferencia,
    MovimientoMpSinVentaML,
)
from kiki_control.presentation.control_consolidado_diagnostics import DiagnosticoControlConsolidado, diagnosticar_control_consolidado
from kiki_control.presentation.control_consolidado_view import texto_tratamiento_neto_comparable
from kiki_control.presentation.ml_eccomapp_view import (
    accion_visible_ml_eccomapp,
    conclusion_ejecutiva_ml_eccomapp,
    etiqueta_aptitud_ml_eccomapp,
    etiqueta_estado_ml_eccomapp,
    motivo_visible_ml_eccomapp,
)
from kiki_control.presentation.review_cases import caso_a_fila, clasificar_revisiones
from kiki_control.presentation.reconciliation_view import (
    CoberturaArchivosPresentacion,
    clave_resultado,
    conclusion_ejecutiva,
    es_excepcion_o_caso_especial,
    etiqueta_estado,
    resumen_kpis_tipado,
)

TIPO_COMPLETO = "Reporte completo"
TIPO_EXCEPCIONES = "Solo excepciones"
TIPO_REVISIONES = "Revisiones pendientes"
ACLARACION_UTILIDAD = "La utilidad es informada por Mercado Libre y no representa resultado contable definitivo."
ACLARACION_FONDOS = "Los movimientos de fondos se informan separados y no se consideran pérdidas comerciales."
ACLARACION_PRIVACIDAD = "El archivo se generó en memoria y excluye datos personales, metadatos sensibles, contenido crudo y nombres de archivos originales."
COLUMNAS_OPERACIONES = (
    "ID de orden",
    "Estado",
    "Neto informado ML",
    "Neto aprobado MP",
    "Diferencia",
    "Neto financiero total",
    "Utilidad informada ML",
    "Pago dividido",
    "Devolución",
    "Reclamo o disputa",
    "Pendiente de acreditación",
    "Requiere revisión",
    "Explicación",
    "Motivos técnicos",
    "Filas ML de origen",
    "Filas MP de origen",
    "Cantidad de pagos aprobados",
    "Cantidad de movimientos financieros",
    "Versión de regla",
    "Tolerancia aplicada",
)
_COLUMNAS_MONETARIAS = {"Neto informado ML", "Neto aprobado MP", "Diferencia", "Neto financiero total", "Utilidad informada ML", "Tolerancia aplicada"}
_KPIS_MONETARIOS = {"Utilidad informada ML", "Neto ML comparable", "Neto MP comparable", "Diferencia comparable", "Neto MP fuera del archivo ML"}
_CAMPOS_ENTEROS_RESUMEN = {"Movimientos sin fecha de liquidación", "Cantidad de filas incluidas en Todas las operaciones", "Cantidad de filas incluidas en Excepciones"}
_COLUMNAS_WRAP = {"Explicación", "Motivos técnicos"}
_FORMATO_MONEDA_ARS = '[$$-es-AR] #,##0.00;[Red]-[$$-es-AR] #,##0.00;[$$-es-AR] 0.00'
_FORMATO_FECHA = "dd/mm/yyyy hh:mm:ss"
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def generar_reporte_completo_excel(reporte: ReporteConciliacion, cobertura: CoberturaArchivosPresentacion | None, zona_horaria: str) -> bytes:
    """Devuelve bytes XLSX del reporte completo, sin escribir archivos."""

    resultados = _resultados_ordenados(reporte.resultados)
    excepciones = [r for r in resultados if es_excepcion_o_caso_especial(r)]
    return _generar_excel(reporte, cobertura, zona_horaria, TIPO_COMPLETO, resultados, excepciones, incluir_todas=True)


def generar_reporte_excepciones_excel(reporte: ReporteConciliacion, cobertura: CoberturaArchivosPresentacion | None, zona_horaria: str) -> bytes:
    """Devuelve bytes XLSX con solo excepciones, sin escribir archivos."""

    resultados = _resultados_ordenados(reporte.resultados)
    excepciones = [r for r in resultados if es_excepcion_o_caso_especial(r)]
    return _generar_excel(reporte, cobertura, zona_horaria, TIPO_EXCEPCIONES, resultados, excepciones, incluir_todas=False)


def generar_revisiones_pendientes_excel(reporte: ReporteConciliacion, cobertura: CoberturaArchivosPresentacion | None, zona_horaria: str) -> bytes:
    """Devuelve bytes XLSX con únicamente resultados que requieren revisión."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    casos = clasificar_revisiones(reporte.resultados)
    _escribir_resumen(ws, reporte, cobertura, zona_horaria, TIPO_REVISIONES, len(reporte.resultados), len(casos))
    _escribir_revisiones(wb.create_sheet("Revisiones pendientes"), casos)
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _generar_excel(reporte: ReporteConciliacion, cobertura: CoberturaArchivosPresentacion | None, zona_horaria: str, tipo: str, resultados: list[ResultadoConciliacion], excepciones: list[ResultadoConciliacion], incluir_todas: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    _escribir_resumen(ws, reporte, cobertura, zona_horaria, tipo, len(resultados), len(excepciones))
    if incluir_todas:
        _escribir_operaciones(wb.create_sheet("Todas las operaciones"), resultados)
    _escribir_operaciones(wb.create_sheet("Excepciones"), excepciones)
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _escribir_resumen(ws: Worksheet, reporte: ReporteConciliacion, cobertura: CoberturaArchivosPresentacion | None, zona_horaria: str, tipo: str, filas_todas: int, filas_excepciones: int) -> None:
    ws.append(["Campo", "Valor"])
    conclusion, _ = conclusion_ejecutiva(reporte)
    filas: list[tuple[str, Any]] = [
        ("Nombre", "Kiki Control Financiero"),
        ("Tipo de reporte", tipo),
        ("Fecha y hora del procesamiento (zona operativa)", _fecha_operativa_sin_tz(reporte.fecha_procesamiento_utc, zona_horaria)),
        ("Zona horaria operativa", zona_horaria),
        ("Versión de la regla de conciliación", reporte.version_regla),
        ("Tolerancia aplicada", _decimal_o_vacio(reporte.tolerancia)),
        ("Cobertura de ventas ML", cobertura.periodo_ventas_ml.texto if cobertura else ""),
        ("Cobertura de origen MP", cobertura.periodo_origen_mp.texto if cobertura else ""),
        ("Cobertura de liquidaciones MP", cobertura.periodo_liquidacion_mp.texto if cobertura else ""),
        ("Movimientos sin fecha de liquidación", cobertura.movimientos_sin_fecha_liquidacion if cobertura else ""),
        ("Conclusión ejecutiva", conclusion),
    ]
    filas.extend((nombre, _decimal_o_vacio(valor) if nombre in _KPIS_MONETARIOS else valor) for nombre, valor in resumen_kpis_tipado(reporte).items())
    if tipo == TIPO_COMPLETO:
        filas.append(("Cantidad de filas incluidas en Todas las operaciones", filas_todas))
    filas.append(("Cantidad de filas incluidas en Excepciones", filas_excepciones))
    filas.extend([("Aclaración", ACLARACION_UTILIDAD), ("Aclaración", ACLARACION_FONDOS), ("Aclaración de privacidad", ACLARACION_PRIVACIDAD)])
    for fila in filas:
        ws.append(list(fila))
    _formatear_tabla(ws, moneda_columnas=set(), wrap_columnas={2}, freeze=False)
    for celda in ws[1]:
        _estilo_header(celda)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _asegurar_celda(cell)
        etiqueta = row[0].value
        if etiqueta == "Fecha y hora del procesamiento (zona operativa)":
            row[1].number_format = _FORMATO_FECHA
        elif etiqueta == "Tolerancia aplicada" or etiqueta in _KPIS_MONETARIOS:
            if row[1].value != "":
                row[1].number_format = _FORMATO_MONEDA_ARS
        elif etiqueta in _CAMPOS_ENTEROS_RESUMEN or etiqueta in resumen_kpis_tipado(reporte):
            if isinstance(row[1].value, int):
                row[1].number_format = "0"
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 90


def _escribir_operaciones(ws: Worksheet, resultados: Iterable[ResultadoConciliacion]) -> None:
    ws.append(list(COLUMNAS_OPERACIONES))
    for resultado in resultados:
        ws.append(_fila_operacion(resultado))
    _formatear_tabla(ws, moneda_columnas={idx for idx, c in enumerate(COLUMNAS_OPERACIONES, start=1) if c in _COLUMNAS_MONETARIAS}, wrap_columnas={idx for idx, c in enumerate(COLUMNAS_OPERACIONES, start=1) if c in _COLUMNAS_WRAP}, freeze=True)


def _fila_operacion(r: ResultadoConciliacion) -> list[Any]:
    return [
        _texto_seguro(r.id_orden or clave_resultado(r)),
        _texto_seguro(etiqueta_estado(r.estado)),
        _decimal_o_vacio(r.neto_comercial_informado),
        _decimal_o_vacio(r.neto_pagos_aprobados),
        _decimal_o_vacio(r.diferencia_control),
        _decimal_o_vacio(r.neto_financiero_total),
        _decimal_o_vacio(r.utilidad_neta_informada),
        _si_no(r.es_pago_dividido),
        _si_no(r.tiene_devolucion),
        _si_no(r.tiene_reclamo or r.tiene_disputa),
        _si_no(r.tiene_liquidacion_pendiente),
        _si_no(r.requiere_revision),
        _texto_seguro(" | ".join(r.explicaciones)),
        _texto_seguro(", ".join(m.value for m in r.motivos)),
        _texto_seguro(", ".join(str(n) for n in r.numeros_fila_comercial)),
        _texto_seguro(", ".join(str(n) for n in r.numeros_fila_financiera)),
        r.cantidad_pagos_aprobados,
        r.cantidad_movimientos_financieros,
        _texto_seguro(r.version_regla),
        _decimal_o_vacio(r.tolerancia_aplicada),
    ]


def _resultados_ordenados(resultados: Iterable[ResultadoConciliacion]) -> list[ResultadoConciliacion]:
    return sorted(resultados, key=lambda r: (r.id_orden is None, r.id_orden or "", r.numeros_fila_financiera, r.estado.value))


def _fecha_operativa_sin_tz(fecha_utc: datetime, zona_horaria: str) -> datetime:
    return fecha_utc.astimezone(ZoneInfo(zona_horaria)).replace(tzinfo=None)


def _decimal_o_vacio(valor: Decimal | None) -> Decimal | Literal[""]:
    return "" if valor is None else valor


def _si_no(valor: bool) -> str:
    return "Sí" if valor else "No"


def _texto_seguro(valor: Any) -> str:
    texto = "" if valor is None else str(valor)
    if texto.startswith(_FORMULA_PREFIXES):
        return f"'{texto}"
    return texto


def _formatear_tabla(ws: Worksheet, moneda_columnas: set[int], wrap_columnas: set[int], freeze: bool) -> None:
    for celda in ws[1]:
        _estilo_header(celda)
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    if freeze:
        ws.freeze_panes = "A2"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            _asegurar_celda(cell)
            if cell.column in moneda_columnas and cell.value != "":
                cell.number_format = _FORMATO_MONEDA_ARS
            if cell.column in wrap_columnas:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
    for idx, encabezado in enumerate((cell.value for cell in ws[1]), start=1):
        ancho = 18
        if encabezado in {"Explicación", "Motivos técnicos"}:
            ancho = 48
        elif encabezado in {"ID de orden", "Versión de regla"}:
            ancho = 26
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = ancho


def _estilo_header(celda: Cell) -> None:
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="1F4E78")
    celda.alignment = Alignment(wrap_text=True)


def _asegurar_celda(celda: Cell) -> None:
    celda.value = _valor_seguro_excel(celda.value)


def _valor_seguro_excel(valor: Any) -> Any:
    """Adapta valores a Excel sin alterar la fecha/hora comercial local.

    OpenPyXL no serializa ``datetime`` ni ``time`` con zona horaria. Las fechas
    llegan aquí después de su normalización operativa, por lo que se elimina
    solamente ``tzinfo``: no se convierten a UTC ni a texto.
    """
    if isinstance(valor, (datetime, time)) and valor.tzinfo is not None:
        return valor.replace(tzinfo=None)
    if isinstance(valor, str):
        return _texto_seguro(valor)
    return valor


def _escribir_revisiones(ws: Worksheet, casos: Iterable[Any]) -> None:
    columnas = ("ID de orden o referencia", "Tipo de revisión", "Estado", "Motivo explicado", "Acción recomendada", "Neto ML", "Neto aprobado MP", "Neto financiero total", "Filas ML", "Filas MP", "Columnas de origen")
    ws.append(list(columnas))
    for caso in casos:
        r = caso.resultado
        fila_presentacion = caso_a_fila(caso)
        ws.append([
            _texto_seguro(fila_presentacion.id_orden_o_referencia),
            _texto_seguro(caso.nombre_visible),
            _texto_seguro(etiqueta_estado(r.estado)),
            _texto_seguro(caso.descripcion),
            _texto_seguro(caso.accion_recomendada),
            _decimal_o_vacio(r.neto_comercial_informado),
            _decimal_o_vacio(r.neto_pagos_aprobados),
            _decimal_o_vacio(r.neto_financiero_total),
            _texto_seguro(", ".join(str(n) for n in r.numeros_fila_comercial)),
            _texto_seguro(", ".join(str(n) for n in r.numeros_fila_financiera)),
            _texto_seguro(", ".join(caso.columnas_utilizadas)),
        ])
    _formatear_tabla(ws, moneda_columnas={6, 7, 8}, wrap_columnas={4, 5, 11}, freeze=True)

TIPO_CONSOLIDADO_TRES_FUENTES = "Reporte consolidado de tres fuentes"
TIPO_EXCEPCIONES_CONSOLIDADAS = "Excepciones del control consolidado"
TIPO_REVISIONES_CONSOLIDADAS = "Revisiones consolidadas"
COLUMNAS_CONTROL_CONSOLIDADO = (
    "Grupo u orden", "Estado", "Neto ML", "Costo de productos Eccomapp", "Neto Eccomapp", "Neto aprobado bruto MP", "Reclamos/disputas MP", "Devoluciones MP", "Envíos MP", "Otros impactos MP", "Neto financiero total MP",
    "Eccomapp − ML", "MP − Eccomapp", "MP − ML", "Utilidad preliminar", "Motivo principal", "Filas ML", "Filas Eccomapp", "Filas MP",
)


def generar_reporte_consolidado_excel(reporte: ReporteControlConsolidado, diagnostico: DiagnosticoControlConsolidado | None = None, diag_bloque_b: DiagnosticoBloqueB | None = None, diagnostico_ml_eccomapp: DiagnosticoMlEccomapp | None = None) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Resumen"
    diag = diagnostico or diagnosticar_control_consolidado(reporte)
    _escribir_resumen_consolidado(ws, reporte, TIPO_CONSOLIDADO_TRES_FUENTES)
    _escribir_cobertura_consolidada(wb.create_sheet("Cobertura y universos"), diag)
    _escribir_puente_consolidado(wb.create_sheet("Puente de fuentes"), diag)
    _escribir_control_consolidado(wb.create_sheet("Control por operación"), reporte.resultados)
    _escribir_temporal_consolidado(wb.create_sheet("Distribución temporal MP"), diag)
    _escribir_revisiones_consolidadas(wb.create_sheet("Revisiones"), diag)
    if diag_bloque_b is not None:
        _escribir_resumen_bloque_b(wb.create_sheet("Bloque B — Resumen"), diag_bloque_b)
        _escribir_diferencias_bloque_b(wb.create_sheet("Bloque B — Diferencias"), diag_bloque_b)
        _escribir_movimientos_bloque_b(wb.create_sheet("Bloque B — Movimientos"), diag_bloque_b)
        _escribir_resumen_mp_sin_ml(wb.create_sheet("Bloque B — Resumen MP sin ML"), diag_bloque_b)
        _escribir_mp_sin_venta_bloque_b(wb.create_sheet("Bloque B — MP sin venta ML"), diag_bloque_b)
        _escribir_fondos_bloque_b(wb.create_sheet("Bloque B — Fondos y payouts"), diag_bloque_b)
        _escribir_candidatos_venta_faltante(wb.create_sheet("Candidatos venta faltante"), diag_bloque_b)
        _escribir_pagos_inconsistentes(wb.create_sheet("Pagos MP inconsistentes"), diag_bloque_b)
    if diagnostico_ml_eccomapp is not None:
        _agregar_hojas_ml_eccomapp(wb, diagnostico_ml_eccomapp)
    _escribir_diccionario_consolidado(wb.create_sheet("Diccionario de cálculos"))
    salida = BytesIO(); wb.save(salida); return salida.getvalue()


def generar_diagnostico_ml_eccomapp_excel(diag: DiagnosticoMlEccomapp) -> bytes:
    """Genera las cinco hojas auditables del cruce comercial, sin PII."""
    wb = Workbook(); wb.remove(wb.active); _agregar_hojas_ml_eccomapp(wb, diag)
    salida = BytesIO(); wb.save(salida); return salida.getvalue()


def _agregar_hojas_ml_eccomapp(wb: Workbook, diag: DiagnosticoMlEccomapp) -> None:
    ws = wb.create_sheet("ML-Eccomapp — Resumen")
    ws.append(["Métrica", "Cantidad"])
    for row in (("Filas ML", diag.cantidad_filas_ml), ("Ventas únicas ML", diag.cantidad_ventas_unicas_ml),
                ("Grupos ML", diag.cantidad_grupos_ml), ("Filas Eccomapp", diag.cantidad_filas_eccomapp),
                ("Operaciones únicas Eccomapp", diag.cantidad_operaciones_unicas_eccomapp), ("Grupos Eccomapp", diag.cantidad_grupos_eccomapp),
                ("Coincidencias exactas", diag.cantidad_coincidencias_exactas), ("Coincidencias por grupo", diag.cantidad_coincidencias_por_grupo),
                ("Grupos comerciales con coincidencia", diag.cantidad_coincidencias),
                ("Grupos ML sin Eccomapp", diag.cantidad_solo_ml), ("Grupos Eccomapp sin ML", diag.cantidad_solo_eccomapp),
                ("Incompletas", diag.cantidad_identificador_incompleto), ("Ambiguas", diag.cantidad_ambiguas),
                ("Duplicadas", diag.cantidad_duplicadas),
                ("Grupos ambiguos o incompletos", diag.cantidad_ambiguas + diag.cantidad_identificador_incompleto + diag.cantidad_duplicadas),
                ("Grupos aptos para utilidad", diag.cantidad_apta_utilidad), ("No aptas", diag.cantidad_no_apta_utilidad)):
        ws.append(row)
    ws.append(("Conclusión ejecutiva", conclusion_ejecutiva_ml_eccomapp(diag)))
    _formatear_tabla(ws, moneda_columnas=set(), wrap_columnas=set(), freeze=True)
    exactas = {EstadoCruceMlEccomapp.COINCIDENCIA_EXACTA, EstadoCruceMlEccomapp.COINCIDENCIA_POR_GRUPO}
    ambiguas = {EstadoCruceMlEccomapp.IDENTIFICADOR_INCOMPLETO, EstadoCruceMlEccomapp.IDENTIFICADOR_AMBIGUO, EstadoCruceMlEccomapp.DUPLICADO_ML, EstadoCruceMlEccomapp.DUPLICADO_ECCOMAPP}
    for name, cases in (("ML sin Eccomapp", (c for c in diag.casos if c.estado == EstadoCruceMlEccomapp.SOLO_ML)),
                        ("Eccomapp sin ML", (c for c in diag.casos if c.estado == EstadoCruceMlEccomapp.SOLO_ECCOMAPP)),
                        ("ML-Eccomapp — Coincidencias", (c for c in diag.casos if c.estado in exactas)),
                        ("ML-Eccomapp — Ambiguos", (c for c in diag.casos if c.estado in ambiguas))):
        _escribir_casos_ml_eccomapp(wb.create_sheet(name), tuple(cases))


def _escribir_casos_ml_eccomapp(ws: Worksheet, cases) -> None:
    ws.append(["Grupo u orden", "IDs venta ML", "IDs orden Eccomapp", "Estado", "Aptitud utilidad", "Fecha", "Filas originales ML", "Filas originales Eccomapp", "Ingresos productos ML", "Ingresos envío ML", "Cargos e impuestos ML", "Costos envío ML", "Anulaciones/reembolsos ML", "Total informado ML", "SKU/publicación ML", "Importe venta Eccomapp", "Costo Eccomapp", "Utilidad informada Eccomapp", "Motivo", "Acción recomendada", "Estado técnico", "Aptitud técnica", "Motivo técnico original"])
    for c in cases:
        sum_ml = lambda attr: _decimal_o_vacio(sum((getattr(v, attr) for v in c.ventas_ml if getattr(v, attr) is not None), Decimal("0"))) if any(getattr(v, attr) is not None for v in c.ventas_ml) else ""
        sum_ec = lambda attr: _decimal_o_vacio(sum((getattr(o, attr) for o in c.operaciones_eccomapp if getattr(o, attr) is not None), Decimal("0"))) if any(getattr(o, attr) is not None for o in c.operaciones_eccomapp) else ""
        ws.append([_texto_seguro(c.id_grupo), _texto_seguro(", ".join(c.ids_venta_ml)), _texto_seguro(", ".join(c.ids_orden_eccomapp)), etiqueta_estado_ml_eccomapp(c.estado), etiqueta_aptitud_ml_eccomapp(c.aptitud_utilidad), c.fecha, _texto_seguro(", ".join(str(v.fila_origen) for v in c.ventas_ml)), _texto_seguro(", ".join(str(o.numero_fila_origen) for o in c.operaciones_eccomapp)), sum_ml("ingresos_productos"), sum_ml("ingresos_envio"), sum_ml("cargo_venta_impuestos"), sum_ml("costos_envio"), sum_ml("anulaciones_reembolsos"), _decimal_o_vacio(c.total_ml), _texto_seguro(", ".join(filter(None, (f"{v.sku or ''}/{v.id_publicacion or ''}" for v in c.ventas_ml)))), sum_ec("monto_venta"), _decimal_o_vacio(c.costo_eccomapp), sum_ec("utilidad_neta_informada"), _texto_seguro(motivo_visible_ml_eccomapp(c)), _texto_seguro(accion_visible_ml_eccomapp(c)), c.estado.value, c.aptitud_utilidad.value, _texto_seguro(c.motivo)])
    _formatear_tabla(ws, moneda_columnas=set(range(9, 19)), wrap_columnas={19, 20, 23}, freeze=True)
    for cell in ws["F"][1:]:
        if cell.value is not None:
            cell.number_format = _FORMATO_FECHA


def generar_excepciones_consolidadas_excel(reporte: ReporteControlConsolidado) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Resumen"
    excepciones = tuple(r for r in reporte.resultados if r.requiere_revision or r.estado.value != "COMPLETA" or (r.diferencia_ml_mp is not None and abs(r.diferencia_ml_mp) > r.tolerancia))
    _escribir_resumen_consolidado(ws, reporte, TIPO_EXCEPCIONES_CONSOLIDADAS)
    _escribir_control_consolidado(wb.create_sheet("Excepciones"), excepciones)
    salida = BytesIO(); wb.save(salida); return salida.getvalue()


def generar_revisiones_consolidadas_excel(reporte: ReporteControlConsolidado) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Resumen"
    diag = diagnosticar_control_consolidado(reporte)
    _escribir_resumen_consolidado(ws, reporte, TIPO_REVISIONES_CONSOLIDADAS)
    _escribir_revisiones_consolidadas(wb.create_sheet("Revisiones"), diag)
    salida = BytesIO(); wb.save(salida); return salida.getvalue()


def _escribir_resumen_consolidado(ws: Worksheet, reporte: ReporteControlConsolidado, tipo: str) -> None:
    ws.append(["Campo", "Valor"])
    diag = diagnosticar_control_consolidado(reporte)
    for fila in (("Nombre", "Kiki Control Financiero"), ("Tipo de reporte", tipo), ("Versión de regla", reporte.version_regla), ("Tolerancia", _decimal_o_vacio(reporte.tolerancia)), ("Total grupos", reporte.total_resultados), ("Venta oficial sin Total (ARS)", reporte.total_total_ml_ausente), ("Estado conciliación Bloque A", diag.residual_ml.estado_conciliacion), ("Método cupón Bloque A", diag.residual_ml.metodo_cupones), ("Diferencia final Bloque A", _decimal_o_vacio(diag.residual_ml.diferencia_final)), ("Costo total Eccomapp", _decimal_o_vacio(diag.utilidad.costo_productos_universo_utilidad + diag.utilidad.costo_eccomapp_fuera_universo_calculable)), ("Costo utilizado en utilidad", _decimal_o_vacio(diag.utilidad.costo_productos_universo_utilidad)), ("Costo excluido", _decimal_o_vacio(diag.utilidad.costo_eccomapp_fuera_universo_calculable)), ("Grupos calculables", diag.utilidad.grupos_calculables), ("Grupos excluidos", diag.utilidad.grupos_excluidos), ("Fórmula utilidad preliminar", "utilidad_preliminar = neto_ml_universo_calculable - costo_eccomapp_universo_calculable"), ("Motivos de exclusión", "; ".join(f"{k}: {v}" for k, v in diag.utilidad.motivos_exclusion.items() if v)), ("Aclaración", "Control operativo preliminar; no es resultado contable ni fiscal definitivo.")):
        ws.append(list(fila))
    _formatear_tabla(ws, moneda_columnas=set(), wrap_columnas={2}, freeze=False)
    for row in ws.iter_rows(min_row=2):
        if row[0].value in {"Tolerancia", "Diferencia final Bloque A", "Costo total Eccomapp", "Costo utilizado en utilidad", "Costo excluido"} and row[1].value != "": row[1].number_format = _FORMATO_MONEDA_ARS


def _escribir_cobertura_consolidada(ws: Worksheet, diag: Any) -> None:
    ws.append(["Fuente", "Universo", "Cantidad total", "Importe total", "Cantidad usada", "Importe usado", "Cantidad excluida", "Importe excluido", "Motivo de exclusión"])
    for c in diag.cobertura_monetaria:
        ws.append([_texto_seguro(c.fuente), _texto_seguro(c.universo), c.cantidad_total, c.importe_total, c.cantidad_usada, c.importe_usado, c.cantidad_excluida, c.importe_excluido, _texto_seguro(c.motivo_exclusion)])
    _formatear_tabla(ws, moneda_columnas={4,6,8}, wrap_columnas={9}, freeze=True)


def _escribir_puente_consolidado(ws: Worksheet, diag: Any) -> None:
    ws.append(["Concepto", "Valor", "Origen", "Método"])
    p = diag.puente
    residual = diag.residual_ml
    filas_monetarias = set()
    for componente in residual.componentes:
        ws.append([_texto_seguro(componente.concepto), componente.importe, _texto_seguro(f"{componente.origen} · {componente.columna_origen}"), _texto_seguro(componente.metodo)])
        filas_monetarias.add(ws.max_row)
    ws.append(["Estado conciliación Bloque A", _texto_seguro(residual.estado_conciliacion), "", ""])
    ws.append(["Método cupón Bloque A", _texto_seguro(residual.metodo_cupones), "", ""])
    ws.append(["Diferencia final Bloque A", residual.diferencia_final, "", ""])
    filas_monetarias.add(ws.max_row)
    ws.append(["", "", "", ""])
    for fila in (("Neto ML", p.neto_oficial_ml), ("Neto Eccomapp", p.neto_informado_eccomapp), ("Neto aprobado MP", p.neto_aprobado_mp), ("Eccomapp − ML", p.eccomapp_menos_ml), ("MP − Eccomapp", p.mp_menos_eccomapp), ("MP − ML", p.mp_menos_ml), ("Aporte excluidos a diferencia ML–MP", p.aporte_excluidos_a_diferencia_ml_mp)):
        ws.append([fila[0], fila[1], "", ""])
        filas_monetarias.add(ws.max_row)
    for fila in (
        ("Residual ML", residual.importe),
        ("Universo ML oficial", residual.grupos_universo_ml_oficial),
        ("Grupos calculables residual ML", residual.grupos_calculables),
        ("Grupos excluidos residual ML", residual.grupos_excluidos),
        ("Suma Total (ARS)", residual.suma_total_ars),
        ("Suma Ingresos por productos (ARS)", residual.suma_ingresos_productos),
        ("Suma Ingresos por envío (ARS)", residual.suma_ingresos_envio),
        ("Suma Cargo por venta e impuestos (ARS)", residual.suma_cargo_venta_impuestos),
        ("Suma Costos de envío (ARS)", residual.suma_costos_envio),
        ("Suma Anulaciones y reembolsos (ARS)", residual.suma_anulaciones_reembolsos),
        ("Suma Cupones de descuento", residual.suma_cupones_descuento),
        ("Identidad residual ML cierra", "Sí" if residual.identidad_cierra_exactamente else "No"),
        ("Motivos exclusión residual ML", "; ".join(f"{k}: {v}" for k, v in residual.motivos_exclusion.items() if v)),
    ):
        ws.append([fila[0], fila[1], "", ""])
        if fila[0] in {"Residual ML", "Suma Total (ARS)", "Suma Ingresos por productos (ARS)", "Suma Ingresos por envío (ARS)", "Suma Cargo por venta e impuestos (ARS)", "Suma Costos de envío (ARS)", "Suma Anulaciones y reembolsos (ARS)", "Suma Cupones de descuento"}:
            filas_monetarias.add(ws.max_row)
    ws.append(["Advertencia", "No comparar importes de universos distintos sin revisar Cobertura y universos.", "", ""])
    ws.append(["Aclaración temporal", "Si se genera sin diagnóstico de sesión, la distribución temporal no puede clasificar contra período ML ni fechas MP y queda sin fecha.", "", ""])
    _formatear_tabla(ws, moneda_columnas=set(), wrap_columnas={3, 4}, freeze=False)
    for row_idx in filas_monetarias:
        ws.cell(row=row_idx, column=2).number_format = _FORMATO_MONEDA_ARS
    for row in range(2, ws.max_row + 1):
        if row not in filas_monetarias and isinstance(ws.cell(row=row, column=2).value, int):
            ws.cell(row=row, column=2).number_format = "0"


def _escribir_control_consolidado(ws: Worksheet, resultados: Iterable[ResultadoControlConsolidado]) -> None:
    ws.append(list(COLUMNAS_CONTROL_CONSOLIDADO))
    for r in resultados:
        ws.append([_texto_seguro(r.id_grupo_canonico or ", ".join(r.ids_orden) or f"fila MP {', '.join(map(str, r.filas_origen_mp))}"), _texto_seguro(r.estado.value), _decimal_o_vacio(r.total_informado_ml), _decimal_o_vacio(r.costo_productos_eccomapp), _decimal_o_vacio(r.neto_mp_eccomapp_informado), _decimal_o_vacio(r.neto_aprobado_mp), _decimal_o_vacio(r.impacto_reclamos_disputas_mp), _decimal_o_vacio(r.impacto_devoluciones_mp), _decimal_o_vacio(r.impacto_pagos_envio_mp), _decimal_o_vacio(r.impacto_otros_mp), _decimal_o_vacio(r.neto_financiero_total_mp), _decimal_o_vacio(r.diferencia_neto_ml_eccomapp), _decimal_o_vacio((r.neto_aprobado_mp - r.neto_mp_eccomapp_informado) if r.neto_aprobado_mp is not None and r.neto_mp_eccomapp_informado is not None else None), _decimal_o_vacio(r.diferencia_ml_mp), _decimal_o_vacio(r.utilidad_preliminar_control), _texto_seguro("; ".join(r.motivos)), _texto_seguro(", ".join(map(str, r.filas_origen_ml))), _texto_seguro(", ".join(map(str, r.filas_origen_eccomapp))), _texto_seguro(", ".join(map(str, r.filas_origen_mp)))])
    _formatear_tabla(ws, moneda_columnas=set(range(3, 16)), wrap_columnas={16}, freeze=True)


def _escribir_temporal_consolidado(ws: Worksheet, diag: Any) -> None:
    ws.append(["Categoría", "Cantidad", "Neto aprobado MP", "Neto financiero total MP", "Aclaración"])
    for nombre, item in (("Anteriores", diag.temporal_mp_sin_venta.anteriores), ("Dentro", diag.temporal_mp_sin_venta.dentro), ("Posteriores", diag.temporal_mp_sin_venta.posteriores), ("Sin fecha", diag.temporal_mp_sin_venta.sin_fecha), ("Fechas mixtas", diag.temporal_mp_sin_venta.fechas_mixtas)):
        ws.append([nombre, item.cantidad, item.neto_aprobado_mp, item.neto_financiero_total_mp, _texto_seguro(diag.temporal_mp_sin_venta.aclaracion)])
    _formatear_tabla(ws, moneda_columnas={3, 4}, wrap_columnas={5}, freeze=True)


def _escribir_revisiones_consolidadas(ws: Worksheet, diag: Any) -> None:
    ws.append(["Motivo", "Cantidad", "Importe afectado", "Acción", "Grupos"])
    for r in diag.revisiones.revisiones_multietiqueta:
        ws.append([_texto_seguro(r.motivo_visible), r.cantidad, _decimal_o_vacio(r.importe_afectado), _texto_seguro(r.accion_recomendada), _texto_seguro(", ".join(r.grupos_involucrados))])
    _formatear_tabla(ws, moneda_columnas={3}, wrap_columnas={4, 5}, freeze=True)


def _escribir_diccionario_consolidado(ws: Worksheet) -> None:
    ws.append(["Cálculo", "Fórmula", "Universo", "Columnas utilizadas"])
    filas = [
        ("Costo de productos Eccomapp", "Suma de Costo Total (Con IVA) ($)", "grupos con Eccomapp informado", "Eccomapp: Costo Total (Con IVA) ($)"),
        ("Costo Eccomapp utilizado en utilidad", "Suma de costo_productos_eccomapp dentro del universo calculable", "grupos con Total (ARS) ML y costo Eccomapp presentes", "ML: Total (ARS); Eccomapp: Costo Total (Con IVA) ($)"),
        ("Costo Eccomapp excluido", "costo_total_eccomapp - costo_eccomapp_universo_calculable", "grupos Eccomapp fuera del universo calculable", "Eccomapp: Costo Total (Con IVA) ($); motivos de exclusión"),
        ("Utilidad preliminar", "Total (ARS) ML - Costo Total (Con IVA) Eccomapp", "universo calculable de utilidad", "ML: Total (ARS); Eccomapp: Costo Total (Con IVA) ($)"),
        ("Formación del neto informado por Mercado Libre", "Ingresos por productos + Ingresos por envío + Cargo por venta e impuestos + Costos de envío + Anulaciones y reembolsos + Cupones de descuento + Otros conceptos pendientes de clasificación = Total (ARS)", "universo ML oficial con Bloque A auditable", "Total (ARS); Ingresos por productos (ARS); Ingresos por envío (ARS); Cargo por venta e impuestos (ARS); Costos de envío (ARS); Anulaciones y reembolsos (ARS); Descuentos y bonificaciones"),
        ("MP − ML", "Neto aprobado MP - Neto ML", "universo ML–Eccomapp–MP para puente triple", "Total (ARS); neto Eccomapp; movimientos aprobados MP"),
        ("Diferencia ML–MP (Bloque B)", "neto_financiero_total_mp − total_informado_ml", "universo comparable ML + MP", "Total (ARS) ML; todos los movimientos MP asociados"),
    ]
    for fila in filas: ws.append([_texto_seguro(x) for x in fila])
    _formatear_tabla(ws, moneda_columnas=set(), wrap_columnas={2,3,4}, freeze=True)


# ---------------------------------------------------------------------------
# Bloque B — Exportaciones
# ---------------------------------------------------------------------------

_COLUMNAS_DIFERENCIAS_BLOQUE_B = (
    "ID de grupo u orden",
    "Fecha de venta ML",
    "Neto informado ML",
    "Neto aprobado bruto MP",
    "Reclamos/disputas MP", "Devoluciones MP", "Envíos MP", "Otros impactos MP",
    "Neto financiero total MP", "Diferencia financiera MP − ML",
    "Movimientos MP",
    "Origen MP desde",
    "Origen MP hasta",
    "Liquidación desde",
    "Liquidación hasta",
    "Tipos de movimientos",
    "Estado de explicación",
    "Motivo",
    "Motivos secundarios",
    "Acción recomendada",
)
_COLS_MONETARIAS_DIF_B = {"Neto informado ML", "Neto aprobado bruto MP", "Reclamos/disputas MP", "Devoluciones MP", "Envíos MP", "Otros impactos MP", "Neto financiero total MP", "Diferencia financiera MP − ML"}

_COLUMNAS_MP_SIN_VENTA = (
    "ID de grupo u orden",
    "IDs de movimiento MP",
    "Categoría temporal principal", "Prioridad operativa", "Subclasificación financiera",
    "Combinación resumida", "Tipos de movimiento", "Interpretación", "posible_venta_faltante",
    "Tiene ID de orden", "Fecha de origen desde", "Fecha de origen hasta",
    "Fecha de liquidación desde", "Fecha de liquidación hasta",
    "Neto aprobado bruto MP",
    "Neto financiero total MP",
    "Suma reconstruida desde movimientos", "Agregado financiero original",
    "Diferencia agregado − detalle", "Coherencia del grupo", "Estado de coherencia",
    "Motivo de coherencia", "Advertencia de inconsistencia",
    "Cantidad de movimientos", "Motivo visible", "Acción recomendada", "Filas de origen MP",
)
_COLS_MONETARIAS_MP_SIN = {"Neto aprobado bruto MP", "Neto financiero total MP", "Suma reconstruida desde movimientos", "Agregado financiero original", "Diferencia agregado − detalle"}


def generar_bloque_b_mp_sin_venta_excel(diag: DiagnosticoBloqueB) -> bytes:
    """Devuelve bytes XLSX con los movimientos MP sin venta ML."""
    wb = Workbook()
    ws = wb.active
    ws.title = "MP sin venta ML"
    _escribir_mp_sin_venta_bloque_b(ws, diag)
    _escribir_resumen_mp_sin_ml(wb.create_sheet("Resumen MP sin ML"), diag)
    _escribir_candidatos_venta_faltante(wb.create_sheet("Candidatos venta faltante"), diag)
    _escribir_pagos_inconsistentes(wb.create_sheet("Pagos MP inconsistentes"), diag)
    salida = BytesIO()
    wb.save(salida)
    return salida.getvalue()


def _escribir_resumen_bloque_b(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    ws.append(["Indicador", "Valor"])
    r = diag.resumen
    filas: list[tuple[str, Any]] = [
        ("Grupos comparables", r.comparables_totales),
        ("Coinciden dentro de tolerancia", r.coincidencias),
        ("Grupos con diferencia", r.con_diferencia),
        ("Neto ML comparable", r.neto_ml_comparable),
        ("Neto financiero total MP comparable", r.neto_mp_comparable),
        ("Diferencia universo comparable completo (MP − ML)", r.diferencia_universo_comparable),
        ("Diferencia operaciones fuera de tolerancia", r.diferencia_operaciones_fuera_tolerancia),
        ("Diferencia subuniverso conciliado", r.diferencia_subuniverso_conciliado),
        ("Suma individual de diferencias", diag.suma_diferencias_individuales),
        ("Coherencia suma diferencias", "Sí" if diag.coherencia_suma_diferencias else "No"),
        ("Cantidad MP sin venta ML", diag.cantidad_mp_sin_venta),
        ("Neto aprobado MP sin venta", diag.neto_aprobado_mp_sin_venta),
        ("Neto financiero total MP sin venta", diag.neto_financiero_total_mp_sin_venta),
        ("Coherencia detalle/importes MP sin venta", "Sí" if diag.coherencia_detalle_importes_mp_sin_venta else "No"),
        ("Validez KPI monetario MP sin venta", "Válido" if diag.coherencia_detalle_importes_mp_sin_venta else "NO VÁLIDO — revisar inconsistencias por grupo"),
        ("Convención", "diferencia_ml_mp = neto_financiero_total_mp − total_informado_ml"),
        ("Tratamiento PAGO_ENVIO", "Componente ya incluido en el neto aprobado bruto MP: se muestra, pero no se suma nuevamente."),
        ("Positiva", "MP informa más neto que ML"),
        ("Negativa", "MP informa menos neto que ML"),
        ("Aclaración", "Control operativo preliminar; no es resultado contable ni fiscal definitivo."),
    ]
    for fila in filas:
        ws.append(list(fila))
    _formatear_tabla(ws, moneda_columnas=set(), wrap_columnas={2}, freeze=False)
    monetarios = {"Neto ML comparable", "Neto MP comparable", "Diferencia universo comparable completo (MP − ML)", "Diferencia operaciones fuera de tolerancia", "Diferencia subuniverso conciliado", "Suma individual de diferencias", "Neto aprobado MP sin venta", "Neto financiero total MP sin venta"}
    for row in ws.iter_rows(min_row=2):
        if row[0].value in monetarios and isinstance(row[1].value, Decimal):
            row[1].number_format = _FORMATO_MONEDA_ARS


def _escribir_diferencias_bloque_b(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    ws.append(list(_COLUMNAS_DIFERENCIAS_BLOQUE_B))
    for g in diag.grupos_con_diferencia:
        ws.append([
            _texto_seguro(g.id_grupo),
            _texto_seguro(g.fecha_venta_ml),
            _decimal_o_vacio(g.total_informado_ml),
            _decimal_o_vacio(g.neto_aprobado_mp),
            _decimal_o_vacio(g.impacto_reclamos_disputas_mp),
            _decimal_o_vacio(g.impacto_devoluciones_mp),
            _decimal_o_vacio(g.impacto_pagos_envio_mp),
            _decimal_o_vacio(g.impacto_otros_mp),
            _decimal_o_vacio(g.neto_financiero_total_mp),
            _decimal_o_vacio(g.diferencia_ml_mp),
            g.cantidad_movimientos_mp,
            _texto_seguro(g.fecha_min_origen_mp),
            _texto_seguro(g.fecha_max_origen_mp),
            _texto_seguro(g.fecha_min_liquidacion),
            _texto_seguro(g.fecha_max_liquidacion),
            _texto_seguro(", ".join(g.tipos_movimientos)),
            _texto_seguro(ESTADOS_EXPLICACION_VISIBLES.get(g.estado_explicacion, g.estado_explicacion)),
            _texto_seguro(g.motivo_visible),
            _texto_seguro("; ".join(g.motivos_secundarios)),
            _texto_seguro(g.accion_recomendada),
        ])
    mon_cols = {idx for idx, c in enumerate(_COLUMNAS_DIFERENCIAS_BLOQUE_B, start=1) if c in _COLS_MONETARIAS_DIF_B}
    _formatear_tabla(ws, moneda_columnas=mon_cols, wrap_columnas={18, 19, 20}, freeze=True)


def _escribir_mp_sin_venta_bloque_b(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    ws.append(list(_COLUMNAS_MP_SIN_VENTA))
    for m in diag.movimientos_mp_sin_venta:
        ws.append([
            _texto_seguro(m.id_grupo),
            _texto_seguro(", ".join(m.ids_movimiento_mp)),
            _texto_seguro(m.categoria_principal.value),
            _texto_seguro(m.prioridad_operativa.value),
            _texto_seguro(m.subclasificacion_financiera.value),
            _texto_seguro(m.combinacion_resumida.value),
            _texto_seguro(", ".join(m.tipos_movimiento)),
            _texto_seguro(m.interpretacion),
            m.posible_venta_faltante,
            _si_no(m.tiene_id_orden_utilizable),
            _texto_seguro(m.fecha_min_origen),
            _texto_seguro(m.fecha_origen_maxima),
            _texto_seguro(m.fecha_liquidacion_minima),
            _texto_seguro(m.fecha_max_liquidacion),
            _decimal_o_vacio(m.neto_aprobado_mp),
            _decimal_o_vacio(m.neto_financiero_total_mp),
            _decimal_o_vacio(m.suma_reconstruida_movimientos_mp),
            _decimal_o_vacio(m.neto_financiero_agregado_original_mp),
            _decimal_o_vacio(m.diferencia_agregado_detalle_mp),
            _si_no(m.coherencia_grupo),
            _texto_seguro(m.estado_coherencia.value),
            _texto_seguro(m.motivo_coherencia),
            _texto_seguro(m.advertencia_inconsistencia),
            m.cantidad_movimientos,
            _texto_seguro(m.motivo_sin_venta),
            _texto_seguro(m.accion_recomendada),
            _texto_seguro(", ".join(map(str, m.filas_origen_mp))),
        ])
    mon_cols = {idx for idx, c in enumerate(_COLUMNAS_MP_SIN_VENTA, start=1) if c in _COLS_MONETARIAS_MP_SIN}
    _formatear_tabla(ws, moneda_columnas=mon_cols, wrap_columnas={14, 15}, freeze=True)


def _escribir_resumen_mp_sin_ml(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    columnas = ("Categoría", "Cantidad de grupos", "Cantidad de movimientos", "Neto aprobado bruto",
                "Neto financiero total", "Con ID de orden", "Sin ID de orden", "Acción recomendada")
    ws.append(list(columnas))
    for r in diag.resumen_mp_sin_venta:
        ws.append([r.categoria.value, r.cantidad_grupos, r.cantidad_movimientos, r.neto_aprobado_bruto,
                   r.neto_financiero_total, r.con_id_orden, r.sin_id_orden, _texto_seguro(r.accion_recomendada)])
    ws.append(["VALIDACIÓN", sum(r.cantidad_grupos for r in diag.resumen_mp_sin_venta), "", "", "", "", "",
               "Coherente" if diag.coherencia_mp_sin_venta else "INCONSISTENTE"])
    _formatear_tabla(ws, moneda_columnas={4, 5}, wrap_columnas={8}, freeze=True)
    pagos = diag.diagnostico_pagos_aprobados
    if pagos is not None:
        ws.append([])
        ws.append(["Diagnóstico de pagos aprobados puros", "Valor"])
        for etiqueta, valor in (
            ("Pagos puros detectados", len(pagos.detectados)),
            ("Candidatos válidos", len(pagos.candidatos_validos)),
            ("Pagos inconsistentes", len(pagos.inconsistentes)),
            ("No candidatos por importe no positivo", len(pagos.no_candidatos_importe_no_positivo)),
            ("Importe válido", pagos.importe_valido_candidatos),
            ("Detectados con ID", pagos.detectados_con_id),
            ("Detectados sin ID", pagos.detectados_sin_id),
            ("Candidatos con ID", pagos.candidatos_con_id),
            ("Candidatos sin ID", pagos.candidatos_sin_id),
            ("Conclusión ejecutiva", pagos.conclusion_ejecutiva),
        ):
            ws.append([etiqueta, valor])
            if isinstance(valor, Decimal):
                ws.cell(ws.max_row, 2).number_format = _FORMATO_MONEDA_ARS
    ws.append([])
    ws.append(["Composición de movimientos dentro del período ML sin venta encontrada"])
    ws.append(["Clasificación operativa", "Subclasificación financiera", "Cantidad de grupos",
               "Cantidad de movimientos", "Neto aprobado bruto", "Neto financiero total",
               "Con ID de orden", "Sin ID de orden", "Interpretación", "Acción recomendada",
               "posible_venta_faltante"])
    for r in diag.resumen_operativo_dentro_periodo:
        ws.append([r.prioridad_operativa.value, r.subclasificacion_financiera.value,
                   r.cantidad_grupos, r.cantidad_movimientos, r.neto_aprobado_bruto,
                   r.neto_financiero_total, r.con_id_orden, r.sin_id_orden,
                   _texto_seguro(r.interpretacion), _texto_seguro(r.accion_recomendada),
                   (r.subclasificacion_financiera.value == "PAGO_APROBADO" and all(
                       m.coherencia_grupo for m in diag.movimientos_mp_sin_venta
                       if m.categoria_principal == CategoriaPrincipalMpSinVenta.DENTRO_DEL_PERIODO_ML_SIN_VENTA
                       and m.subclasificacion_financiera == r.subclasificacion_financiera
                   ))])
        ws.cell(ws.max_row, 5).number_format = _FORMATO_MONEDA_ARS
        ws.cell(ws.max_row, 6).number_format = _FORMATO_MONEDA_ARS
    calidad = diag.calidad_monetaria_mp_sin_venta
    if calidad is not None:
        ws.append([])
        ws.append(["Control de calidad monetaria por fila", "Valor"])
        filas_calidad = (
            ("Grupos coherentes", calidad.grupos_coherentes),
            ("Grupos incoherentes", calidad.grupos_incoherentes),
            ("Grupos no verificables", calidad.grupos_no_verificables),
            ("Movimientos con correspondencia inconsistente", calidad.movimientos_correspondencia_inconsistente),
            ("Pagos aprobados negativos", calidad.pagos_aprobados_negativos),
            ("Importe reconstruido confiable", calidad.importe_reconstruido_confiable),
            ("Importe reconstruido excluido de KPI", calidad.importe_reconstruido_excluido_kpi),
            ("Agregado original de referencia", calidad.agregado_original_referencia),
            ("Diferencia agregado − detalle", calidad.diferencia_agregado_detalle),
            ("Importe no verificable", calidad.importe_no_verificable),
            ("Cantidad de grupos excluidos", calidad.cantidad_grupos_excluidos),
            ("Cantidad de grupos sin reconstrucción", calidad.cantidad_grupos_sin_reconstruccion),
        )
        for etiqueta, valor in filas_calidad:
            ws.append([etiqueta, _decimal_o_vacio(valor) if isinstance(valor, Decimal) else valor])
            if isinstance(valor, Decimal):
                ws.cell(ws.max_row, 2).number_format = _FORMATO_MONEDA_ARS


def _escribir_candidatos_venta_faltante(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    columnas = ("ID de grupo", "ID movimiento MP", "ID de orden", "Fila original MP",
                "Fecha de origen", "Fecha de aprobación", "Fecha de liquidación",
                "Importe crudo", "Importe normalizado", "Neto reconstruido",
                "Estado de correspondencia", "Estado monetario",
                "Motivo de posible venta faltante", "Acción recomendada")
    ws.append(list(columnas))
    pagos = diag.diagnostico_pagos_aprobados
    for grupo in (() if pagos is None else pagos.candidatos_validos):
        for mov in grupo.movimientos_asociados:
            ws.append([_texto_seguro(grupo.id_grupo), _texto_seguro(mov.id_movimiento_mp),
                       _texto_seguro(mov.id_orden), mov.fila_origen, _texto_seguro(mov.fecha_origen),
                       _texto_seguro(mov.fecha_aprobacion), _texto_seguro(mov.fecha_liquidacion),
                       _texto_seguro(mov.importe_crudo), _decimal_o_vacio(mov.monto_neto_impactado),
                       _decimal_o_vacio(grupo.suma_reconstruida_movimientos_mp),
                       _texto_seguro(mov.estado_correspondencia_fila),
                       _texto_seguro(grupo.estado_coherencia.value),
                       _texto_seguro(grupo.interpretacion), _texto_seguro(grupo.accion_recomendada)])
    _formatear_tabla(ws, moneda_columnas={9, 10}, wrap_columnas={13, 14}, freeze=True)


def _escribir_pagos_inconsistentes(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    columnas = ("Fila original", "ID movimiento", "ID orden", "Importe",
                "Motivo de exclusión", "Estado", "Acción recomendada")
    ws.append(list(columnas))
    pagos = diag.diagnostico_pagos_aprobados
    for grupo in (() if pagos is None else pagos.inconsistentes):
        for mov in grupo.movimientos_asociados:
            ws.append([mov.fila_origen, _texto_seguro(mov.id_movimiento_mp),
                       _texto_seguro(mov.id_orden), _decimal_o_vacio(mov.monto_neto_impactado),
                       _texto_seguro(grupo.motivo_coherencia), _texto_seguro(grupo.estado_coherencia.value),
                       _texto_seguro("Revisar la semántica del movimiento en Mercado Pago antes de considerarlo candidato.")])
    _formatear_tabla(ws, moneda_columnas={4}, wrap_columnas={5, 7}, freeze=True)


def _escribir_fondos_bloque_b(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    """Escribe fondos/payouts en un universo separado de las ventas faltantes."""
    ws.append(list(_COLUMNAS_MP_SIN_VENTA))
    for m in diag.movimientos_fondos:
        ws.append([_texto_seguro(m.id_grupo), _texto_seguro(", ".join(m.ids_movimiento_mp)),
                   "MOVIMIENTO_DE_FONDOS", _texto_seguro(m.prioridad_operativa.value),
                   _texto_seguro(m.subclasificacion_financiera.value), _texto_seguro(m.combinacion_resumida.value),
                   _texto_seguro(", ".join(m.tipos_movimiento)), _texto_seguro(m.interpretacion), False,
                   _si_no(m.tiene_id_orden_utilizable),
                   _texto_seguro(m.fecha_min_origen), _texto_seguro(m.fecha_origen_maxima),
                   _texto_seguro(m.fecha_liquidacion_minima), _texto_seguro(m.fecha_max_liquidacion),
                   _decimal_o_vacio(m.neto_aprobado_mp), _decimal_o_vacio(m.neto_financiero_total_mp),
                   _decimal_o_vacio(m.suma_reconstruida_movimientos_mp),
                   _decimal_o_vacio(m.neto_financiero_agregado_original_mp),
                   _decimal_o_vacio(m.diferencia_agregado_detalle_mp), _si_no(m.coherencia_grupo),
                   _texto_seguro(m.estado_coherencia.value), _texto_seguro(m.motivo_coherencia),
                   _texto_seguro(m.advertencia_inconsistencia),
                   m.cantidad_movimientos, _texto_seguro(m.motivo_sin_venta),
                   _texto_seguro(m.accion_recomendada), _texto_seguro(", ".join(map(str, m.filas_origen_mp)))])
    mon_cols = {idx for idx, c in enumerate(_COLUMNAS_MP_SIN_VENTA, start=1) if c in _COLS_MONETARIAS_MP_SIN}
    _formatear_tabla(ws, moneda_columnas=mon_cols, wrap_columnas={9, 10}, freeze=True)


def _escribir_movimientos_bloque_b(ws: Worksheet, diag: DiagnosticoBloqueB) -> None:
    columnas = ("ID de grupo", "ID de movimiento MP", "ID de orden", "Tipo", "Clasificación normalizada", "Tratamiento en neto comparable",
                "Fecha de origen", "Fecha de aprobación", "Fecha de liquidación",
                "Importe crudo", "Neto impactado", "Columna fuente del importe",
                "Fila de origen", "Estado correspondencia de fila")
    ws.append(list(columnas))
    for id_grupo, movimientos in diag.grupos_movimientos_asociados:
        for mov in movimientos:
            ws.append([_texto_seguro(id_grupo), _texto_seguro(mov.id_movimiento_mp),
                       _texto_seguro(mov.id_orden), _texto_seguro(mov.tipo_movimiento),
                       _texto_seguro(mov.clasificacion_normalizada),
                       _texto_seguro(texto_tratamiento_neto_comparable(mov.tratamiento_neto_comparable)),
                       _texto_seguro(mov.fecha_origen),
                       _texto_seguro(mov.fecha_aprobacion), _texto_seguro(mov.fecha_liquidacion),
                       _texto_seguro(mov.importe_crudo), _decimal_o_vacio(mov.monto_neto_impactado),
                       _texto_seguro(mov.columna_fuente_importe), mov.fila_origen,
                       _texto_seguro(mov.estado_correspondencia_fila)])
    _formatear_tabla(ws, moneda_columnas={11}, wrap_columnas={6, 14}, freeze=True)
