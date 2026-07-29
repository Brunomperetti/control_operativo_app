"""Parser estricto del Account Statement de Mercado Pago."""

from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
from kiki_control.normalization.values import normalizar_identificador

CABECERA_RESUMEN = ("INITIAL_BALANCE", "CREDITS", "DEBITS", "FINAL_BALANCE")
CABECERA_MOVIMIENTOS = ("RELEASE_DATE", "TRANSACTION_TYPE", "REFERENCE_ID", "TRANSACTION_NET_AMOUNT", "PARTIAL_BALANCE")


class ErrorEstadoCuentaMp(ValueError):
    """El archivo no cumple el contrato documentado por Mercado Pago."""


def parsear_decimal_estado_cuenta(valor: object, campo: str) -> Decimal:
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        raise ErrorEstadoCuentaMp(f"{campo}: valor obligatorio vacío")
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        return Decimal(str(valor))
    texto = str(valor).strip().replace("$", "").replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto)
    except InvalidOperation as exc:
        raise ErrorEstadoCuentaMp(f"{campo}: importe inválido {valor!r}") from exc


def normalizar_estado_cuenta_mp(nombre: str, contenido: bytes, tolerancia: Decimal = Decimal("0.01")) -> ResumenEstadoCuentaMp:
    if Path(nombre).suffix.lower() != ".xlsx":
        raise ErrorEstadoCuentaMp("El Account Statement debe ser XLSX")
    if tolerancia < Decimal("0"):
        raise ErrorEstadoCuentaMp("La tolerancia de validación no puede ser negativa")
    # El modo normal evita el acceso cuadrático de ``ReadOnlyWorksheet.cell`` y
    # sigue procesando el libro completamente en memoria, como el resto del flujo.
    wb = load_workbook(BytesIO(contenido), read_only=False, data_only=True)
    ws = wb.active
    cabecera_resumen = tuple(ws.cell(1, c).value for c in range(1, 5))
    cabecera_movimientos = tuple(ws.cell(4, c).value for c in range(1, 6))
    if cabecera_resumen != CABECERA_RESUMEN:
        raise ErrorEstadoCuentaMp(f"Firma de resumen inválida: {cabecera_resumen!r}")
    if cabecera_movimientos != CABECERA_MOVIMIENTOS:
        raise ErrorEstadoCuentaMp(f"Firma de movimientos inválida: {cabecera_movimientos!r}")
    saldos = [parsear_decimal_estado_cuenta(ws.cell(2, c).value, CABECERA_RESUMEN[c - 1]) for c in range(1, 5)]
    digest = sha256(contenido).hexdigest()
    movimientos = []
    for numero in range(5, ws.max_row + 1):
        valores = [ws.cell(numero, c).value for c in range(1, 6)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in valores):
            continue
        fecha = _fecha(valores[0], numero)
        tipo = str(valores[1]).strip() if valores[1] is not None else ""
        if not tipo:
            raise ErrorEstadoCuentaMp(f"Fila {numero}: TRANSACTION_TYPE vacío")
        try:
            reference_id = normalizar_identificador(valores[2], campo="REFERENCE_ID", opcional=True)
        except ValueError as exc:
            raise ErrorEstadoCuentaMp(f"Fila {numero}: {exc}") from exc
        importe = parsear_decimal_estado_cuenta(valores[3], f"Fila {numero} TRANSACTION_NET_AMOUNT")
        parcial = None if valores[4] is None or str(valores[4]).strip() == "" else parsear_decimal_estado_cuenta(valores[4], f"Fila {numero} PARTIAL_BALANCE")
        movimientos.append(MovimientoEstadoCuentaMp(numero, fecha, tipo, reference_id, importe, parcial, digest, ws.title))
    if not movimientos:
        raise ErrorEstadoCuentaMp("El estado de cuenta no contiene movimientos")
    resumen = ResumenEstadoCuentaMp(*saldos, tuple(movimientos), min(m.fecha_liberacion for m in movimientos), max(m.fecha_liberacion for m in movimientos))
    positivos = sum((m.importe_neto for m in movimientos if m.importe_neto > Decimal("0")), Decimal("0"))
    negativos = sum((m.importe_neto for m in movimientos if m.importe_neto < Decimal("0")), Decimal("0"))
    _validar_cercania("créditos", positivos, resumen.creditos_informados, tolerancia)
    _validar_cercania("débitos", negativos, resumen.debitos_informados, tolerancia)
    _validar_cercania("saldo final", resumen.saldo_final_calculado, resumen.saldo_final_informado, tolerancia)
    return resumen


def _fecha(valor: object, fila: int) -> datetime:
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, date):
        return datetime.combine(valor, datetime.min.time())
    if valor is not None:
        texto = str(valor).strip()
        for formato in (None, "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.fromisoformat(texto.replace("Z", "+00:00")) if formato is None else datetime.strptime(texto, formato)
            except ValueError:
                pass
    raise ErrorEstadoCuentaMp(f"Fila {fila}: RELEASE_DATE inválida")


def _validar_cercania(nombre: str, calculado: Decimal, informado: Decimal, tolerancia: Decimal) -> None:
    diferencia = calculado - informado
    if diferencia > tolerancia or diferencia < -tolerancia:
        raise ErrorEstadoCuentaMp(f"No coincide {nombre}: calculado {calculado}, informado {informado}")
