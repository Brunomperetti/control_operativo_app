"""Normalización del reporte XLSX oficial de ventas de Mercado Libre."""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from zipfile import BadZipFile
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from openpyxl import load_workbook

from kiki_control.domain.enums import SeveridadValidacion, TipoFuente
from kiki_control.domain.official_sale import VentaOficialMercadoLibre
from kiki_control.ingestion.file_inspector import inspeccionar_archivo
from kiki_control.normalization.values import es_vacio, normalizar_decimal, normalizar_identificador
from kiki_control.validation.results import ProblemaValidacion

ZONA_HORARIA_PREDETERMINADA = "America/Argentina/Cordoba"

MESES_ES = {
    "enero": 1,
    "febrero": 2,
    "marzo": 3,
    "abril": 4,
    "mayo": 5,
    "junio": 6,
    "julio": 7,
    "agosto": 8,
    "septiembre": 9,
    "setiembre": 9,
    "octubre": 10,
    "noviembre": 11,
    "diciembre": 12,
}

_FECHA_TEXTUAL_ES_RE = re.compile(
    r"^\s*(?P<dia>\d{1,2})\s+de\s+(?P<mes>[a-záéíóúñ]+)\s+de\s+(?P<anio>\d{4})"
    r"\s+(?P<hora>\d{1,2}):(?P<minuto>\d{2})(?::(?P<segundo>\d{2}))?\s+hs\.?\s*$",
    re.IGNORECASE,
)


@dataclass(frozen=True)
class FilaRechazadaVentaMercadoLibre:
    numero_fila_origen: int
    errores: tuple[ProblemaValidacion, ...]


@dataclass(frozen=True)
class ResultadoNormalizacionVentasMercadoLibre:
    ventas: tuple[VentaOficialMercadoLibre, ...]
    filas_rechazadas: tuple[FilaRechazadaVentaMercadoLibre, ...]
    errores: tuple[ProblemaValidacion, ...]
    advertencias: tuple[ProblemaValidacion, ...]
    cantidad_total_recibida: int
    cantidad_normalizada: int
    cantidad_rechazada: int
    hash_importacion: str
    hoja_procesada: str | None
    columnas_originales: tuple[str, ...]


def normalizar_ventas_mercado_libre(nombre_archivo: str, contenido: bytes, zona_horaria: str = ZONA_HORARIA_PREDETERMINADA) -> ResultadoNormalizacionVentasMercadoLibre:
    """Normaliza exclusivamente XLSX oficiales de ventas de Mercado Libre en memoria."""

    inspeccion = inspeccionar_archivo(nombre_archivo, contenido)
    if inspeccion.fuente_detectada != TipoFuente.MERCADO_LIBRE_VENTAS:
        error = ProblemaValidacion("FUENTE_NO_COMPATIBLE", "El adaptador de ventas oficiales de Mercado Libre solo procesa XLSX con encabezado '# de venta'.", SeveridadValidacion.ERROR)
        return _vacio(error, inspeccion)
    if not inspeccion.es_valido:
        return ResultadoNormalizacionVentasMercadoLibre(tuple(), tuple(), inspeccion.errores, inspeccion.advertencias, inspeccion.metadatos.cantidad_filas, 0, inspeccion.metadatos.cantidad_filas, inspeccion.metadatos.sha256, inspeccion.metadatos.nombre_hoja, inspeccion.metadatos.columnas_encontradas)
    if inspeccion.metadatos.extension != ".xlsx":
        error = ProblemaValidacion("FORMATO_NO_COMPATIBLE", "Las ventas oficiales de Mercado Libre se procesan exclusivamente desde XLSX.", SeveridadValidacion.ERROR)
        return _vacio(error, inspeccion)
    try:
        zona = ZoneInfo(zona_horaria)
    except ZoneInfoNotFoundError:
        error = ProblemaValidacion("ZONA_HORARIA_INVALIDA", "La zona horaria configurada no existe en zoneinfo.", SeveridadValidacion.ERROR)
        return _vacio(error, inspeccion)

    filas = _leer_filas(contenido, inspeccion.metadatos.nombre_hoja)
    ventas: list[VentaOficialMercadoLibre] = []
    rechazadas: list[FilaRechazadaVentaMercadoLibre] = []
    errores_globales: list[ProblemaValidacion] = []
    for numero, fila in filas:
        venta, errores = _normalizar_fila(fila, numero, inspeccion.metadatos.sha256, zona)
        if errores:
            rechazadas.append(FilaRechazadaVentaMercadoLibre(numero, tuple(errores)))
            errores_globales.extend(errores)
        else:
            ventas.append(venta)
    return ResultadoNormalizacionVentasMercadoLibre(tuple(ventas), tuple(rechazadas), tuple(errores_globales), inspeccion.advertencias, len(filas), len(ventas), len(rechazadas), inspeccion.metadatos.sha256, inspeccion.metadatos.nombre_hoja, inspeccion.metadatos.columnas_encontradas)


def _vacio(error: ProblemaValidacion, inspeccion):
    return ResultadoNormalizacionVentasMercadoLibre(tuple(), tuple(), (error,), inspeccion.advertencias, 0, 0, 0, inspeccion.metadatos.sha256, inspeccion.metadatos.nombre_hoja, inspeccion.metadatos.columnas_encontradas)


def _leer_filas(contenido: bytes, hoja: str | None) -> list[tuple[int, dict[str, object]]]:
    try:
        wb = load_workbook(BytesIO(contenido), read_only=True, data_only=True)
    except (BadZipFile, OSError, ValueError) as exc:
        raise ValueError("XLSX inválido") from exc
    ws = wb[hoja] if hoja and hoja in wb.sheetnames else wb.active
    encabezado = None
    header_idx = 0
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        valores = tuple("" if v is None else str(v).strip() for v in row)
        if any(v == "# de venta" for v in valores):
            encabezado = _desambiguar_encabezados(valores)
            header_idx = idx
            break
    if encabezado is None:
        return []
    filas = []
    for idx, row in enumerate(ws.iter_rows(min_row=header_idx + 1, values_only=True), start=header_idx + 1):
        if any(not es_vacio(v) for v in row):
            filas.append((idx, dict(zip(encabezado, row, strict=False))))
    return filas


def _desambiguar_encabezados(encabezado: tuple[str, ...]) -> tuple[str, ...]:
    vistos: dict[str, int] = {}
    resultado: list[str] = []
    for nombre in encabezado:
        cantidad = vistos.get(nombre, 0)
        resultado.append(nombre if cantidad == 0 else f"{nombre}.{cantidad}")
        vistos[nombre] = cantidad + 1
    return tuple(resultado)


def _normalizar_fila(fila: dict[str, object], numero: int, hash_importacion: str, zona: ZoneInfo):
    errores: list[ProblemaValidacion] = []
    id_venta = _id(fila.get("# de venta"), "# de venta", numero, errores)
    fecha_venta = _fecha(fila.get("Fecha de venta"), "Fecha de venta", numero, errores, zona, opcional=True)
    sku = _id(fila.get("SKU"), "SKU", numero, errores, opcional=True)
    id_publicacion = _id(fila.get("# de publicación"), "# de publicación", numero, errores, opcional=True)
    if not id_venta:
        return None, errores
    venta = VentaOficialMercadoLibre(
        fila_origen=numero,
        hash_importacion=hash_importacion,
        id_venta=id_venta,
        fecha_venta=fecha_venta,
        estado=_txt(fila.get("Estado")),
        descripcion_estado=_txt(fila.get("Descripción del estado")),
        paquete_varios_productos=_bool(fila.get("Paquete de varios productos"), "Paquete de varios productos", numero, errores),
        pertenece_kit=_bool(fila.get("Pertenece a un kit"), "Pertenece a un kit", numero, errores),
        unidades=_int(fila.get("Unidades"), "Unidades", numero, errores),
        ingresos_productos=_dec(fila.get("Ingresos por productos (ARS)"), "Ingresos por productos (ARS)", numero, errores),
        cargo_venta_impuestos=_dec(fila.get("Cargo por venta e impuestos (ARS)"), "Cargo por venta e impuestos (ARS)", numero, errores),
        ingresos_envio=_dec(fila.get("Ingresos por envío (ARS)"), "Ingresos por envío (ARS)", numero, errores),
        costos_envio=_dec(fila.get("Costos de envío (ARS)"), "Costos de envío (ARS)", numero, errores),
        costo_envio_declarado=_dec(fila.get("Costo de envío basado en medidas y peso declarados"), "Costo de envío basado en medidas y peso declarados", numero, errores),
        cargo_diferencias_envio=_dec(fila.get("Cargo por diferencias en medidas y peso del paquete"), "Cargo por diferencias en medidas y peso del paquete", numero, errores),
        descuentos_bonificaciones=_dec(fila.get("Descuentos y bonificaciones"), "Descuentos y bonificaciones", numero, errores),
        anulaciones_reembolsos=_dec(fila.get("Anulaciones y reembolsos (ARS)"), "Anulaciones y reembolsos (ARS)", numero, errores),
        total_informado_ml=_dec(fila.get("Total (ARS)"), "Total (ARS)", numero, errores),
        sku=sku,
        id_publicacion=id_publicacion,
        canal_venta=_txt(fila.get("Canal de venta")),
        titulo_publicacion=_txt(fila.get("Título de la publicación")),
        variante=_txt(fila.get("Variante")),
        precio_unitario=_dec(fila.get("Precio unitario de venta de la publicación (ARS)"), "Precio unitario de venta de la publicación (ARS)", numero, errores),
        forma_entrega=_txt(fila.get("Forma de entrega")),
        reclamo_abierto=_bool(fila.get("Reclamo abierto"), "Reclamo abierto", numero, errores),
        reclamo_cerrado=_bool(fila.get("Reclamo cerrado"), "Reclamo cerrado", numero, errores),
        con_mediacion=_bool(fila.get("Con mediación"), "Con mediación", numero, errores),
    )
    return venta, errores


def _id(v: object, campo: str, n: int, errores: list[ProblemaValidacion], opcional: bool=False) -> str | None:
    try:
        return normalizar_identificador(v, campo=campo, opcional=opcional)
    except Exception:
        errores.append(_problema("IDENTIFICADOR_INVALIDO", "Identificador inválido.", n, campo))
        return None


def _dec(v: object, campo: str, n: int, errores: list[ProblemaValidacion]) -> Decimal | None:
    if es_vacio(v):
        return None
    try:
        return normalizar_decimal(v, campo=campo, opcional=True)
    except Exception:
        errores.append(_problema("IMPORTE_INVALIDO", "Importe inválido.", n, campo))
        return None


def _int(v: object, campo: str, n: int, errores: list[ProblemaValidacion]) -> int | None:
    if es_vacio(v):
        return None
    try:
        valor = Decimal(str(v))
    except InvalidOperation:
        errores.append(_problema("ENTERO_INVALIDO", "Entero inválido.", n, campo))
        return None
    if valor != valor.to_integral_value():
        errores.append(_problema("ENTERO_INVALIDO", "Entero inválido.", n, campo))
        return None
    return int(valor)


def _txt(v: object) -> str | None:
    return None if es_vacio(v) else str(v).strip()


def _bool(v: object, campo: str, n: int, errores: list[ProblemaValidacion]) -> bool | None:
    if es_vacio(v):
        return None
    t = str(v).strip().lower()
    if t in {"sí", "si", "s", "yes", "true", "1"}:
        return True
    if t in {"no", "false", "0"}:
        return False
    errores.append(_problema("BOOLEANO_INVALIDO", "Indicador booleano inválido.", n, campo))
    return None


def _fecha(v: object, campo: str, n: int, errores: list[ProblemaValidacion], zona: ZoneInfo, opcional: bool = False) -> datetime | None:
    if es_vacio(v):
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=zona) if v.tzinfo is None else v.astimezone(zona)
    texto = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(texto, fmt).replace(tzinfo=zona)
        except ValueError:
            pass
    fecha_textual = _fecha_textual_es(texto, zona)
    if fecha_textual is not None:
        return fecha_textual
    try:
        dt = datetime.fromisoformat(texto)
        return dt.replace(tzinfo=zona) if dt.tzinfo is None else dt.astimezone(zona)
    except ValueError:
        errores.append(_problema("FECHA_INVALIDA", "Fecha inválida.", n, campo))
        return None


def _fecha_textual_es(texto: str, zona: ZoneInfo) -> datetime | None:
    coincidencia = _FECHA_TEXTUAL_ES_RE.fullmatch(texto)
    if coincidencia is None:
        return None
    mes = MESES_ES.get(coincidencia.group("mes").lower())
    if mes is None:
        return None
    try:
        return datetime(
            int(coincidencia.group("anio")),
            mes,
            int(coincidencia.group("dia")),
            int(coincidencia.group("hora")),
            int(coincidencia.group("minuto")),
            int(coincidencia.group("segundo") or 0),
            tzinfo=zona,
        )
    except ValueError:
        return None


def _problema(codigo: str, mensaje: str, numero: int, columna: str) -> ProblemaValidacion:
    return ProblemaValidacion(codigo, mensaje, SeveridadValidacion.ERROR, columna=columna, fila=numero)
