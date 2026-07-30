from datetime import datetime
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook
import pytest

from kiki_control.domain.account_statement import CategoriaEstadoCuentaMp, EstadoVinculacionEstadoCuentaMp
from kiki_control.exporting import generar_control_estado_cuenta_mp_excel
from kiki_control.linking.account_statement import controlar_estado_cuenta_mp
from kiki_control.normalization.account_statement import normalizar_estado_cuenta_mp, parsear_decimal_estado_cuenta


def archivo_estado():
    wb = Workbook(); ws = wb.active; ws.title = "ACCOUNT_STATEMENT"
    ws.append(["INITIAL_BALANCE", "CREDITS", "DEBITS", "FINAL_BALANCE"])
    ws.append(["1.000,00", "160,00", "-40,00", "1.120,00"]); ws.append([])
    ws.append(["RELEASE_DATE", "TRANSACTION_TYPE", "REFERENCE_ID", "TRANSACTION_NET_AMOUNT", "PARTIAL_BALANCE"])
    ws.append([datetime(2026, 7, 28), "Liquidación de dinero", "169679883346", "100,00", "1.100,00"])
    ws.append([datetime(2026, 7, 28), "Dinero retenido", "169679883346", "-40,00", None])
    ws.append([datetime(2026, 7, 28), "Rendimientos", None, "10,00", None])
    ws.append([datetime(2026, 7, 28), "Liquidación de dinero", "SIN-VINCULO", "50,00", "1.120,00"])
    out = BytesIO(); wb.save(out); return out.getvalue()


def settlement():
    return SimpleNamespace(id_operacion_mercado_pago="169679883346", numero_fila_origen=781, id_orden=None, canal_venta="Mercado Pago", plataforma_cobro="Código QR")


def test_leyenda_b3_distingue_categoria_tecnica_y_atribucion_comercial():
    from kiki_control.presentation.account_statement_view import detalle_cobertura_tecnica, leyenda_cobertura_tecnica

    control = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    texto = leyenda_cobertura_tecnica(control)
    assert texto == ("Cobertura calculada: 4 líneas procesadas · 4 categorizadas exactamente una vez · "
                     "2 sin atribución comercial suficiente · 0 con conflicto · diferencia monetaria $ 0,00.")
    assert detalle_cobertura_tecnica(control) == "Sin categoría técnica: 0 · procesadas exactamente una vez: 4."
    assert "sin clasificación" not in texto


def test_leyenda_b3_es_dinamica_para_cero_pendientes_faltante_y_conflicto():
    from dataclasses import replace
    from kiki_control.domain.account_statement import ControlEstadoCuentaMp
    from kiki_control.presentation.account_statement_view import detalle_cobertura_tecnica, leyenda_cobertura_tecnica

    base = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    cero = ControlEstadoCuentaMp(base.resumen, tuple(
        replace(m, categoria=CategoriaEstadoCuentaMp.OTRO_INGRESO_NO_ML_IDENTIFICADO)
        if m.categoria == CategoriaEstadoCuentaMp.SIN_ASOCIACION_SUFICIENTE else m
        for m in base.movimientos
    ))
    # Una entrada omitida y otra repetida representan, respectivamente, falta técnica y conflicto.
    inconsistente = ControlEstadoCuentaMp(base.resumen, base.movimientos[:-1] + (base.movimientos[0],))
    assert "0 sin atribución comercial suficiente" in leyenda_cobertura_tecnica(cero)
    assert "Sin categoría técnica: 1" in detalle_cobertura_tecnica(inconsistente)
    assert "1 con conflicto" in leyenda_cobertura_tecnica(inconsistente)
    assert "$ 50,00" in leyenda_cobertura_tecnica(inconsistente)


def test_sintesis_ejecutiva_singular_plural_y_sin_pendientes():
    from dataclasses import replace
    from kiki_control.presentation.account_statement_view import sintesis_ejecutiva

    control = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    b1 = SimpleNamespace(resumen=SimpleNamespace(con_diferencia=0))
    assert sintesis_ejecutiva(b1, control) == ("Conciliación ML–MP sin diferencias. El saldo de Mercado Pago cierra. "
                                               "Quedan 2 movimientos pendientes de atribución comercial.")
    pendiente = next(m for m in control.movimientos if m.categoria == CategoriaEstadoCuentaMp.SIN_ASOCIACION_SUFICIENTE)
    uno = replace(control, movimientos=tuple(m for m in control.movimientos if m is not pendiente))
    assert "Queda 1 movimiento pendiente" in sintesis_ejecutiva(b1, uno)
    sin_pendientes = replace(control, movimientos=tuple(m for m in control.movimientos if m.categoria != CategoriaEstadoCuentaMp.SIN_ASOCIACION_SUFICIENTE))
    assert "No quedan movimientos pendientes" in sintesis_ejecutiva(b1, sin_pendientes)
    assert "con 1 diferencia" in sintesis_ejecutiva(SimpleNamespace(resumen=SimpleNamespace(con_diferencia=1)), uno)


def test_parser_decimal_argentino_y_control_de_saldo():
    assert parsear_decimal_estado_cuenta("$ 26.316.618,06", "saldo") == Decimal("26316618.06")
    resumen = normalizar_estado_cuenta_mp("account_statement.xlsx", archivo_estado())
    assert len(resumen.movimientos) == 4
    assert resumen.variacion_neta == Decimal("120.00")
    assert resumen.saldo_final_calculado == resumen.saldo_final_informado
    assert resumen.movimientos[1].saldo_parcial is None
    assert resumen.movimientos[0].reference_id == "169679883346"


def test_vinculo_repetido_no_es_ambiguo_y_clasificacion_excluyente():
    control = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    assert control.movimientos[0].estado_vinculacion == EstadoVinculacionEstadoCuentaMp.VINCULADO_SETTLEMENT
    assert control.movimientos[1].estado_vinculacion == EstadoVinculacionEstadoCuentaMp.VINCULADO_SIN_ORIGEN_COMERCIAL
    assert control.lineas_vinculadas == 2
    assert control.operaciones_settlement_vinculadas == 1
    assert control.movimientos[0].categoria == CategoriaEstadoCuentaMp.OTRO_INGRESO_NO_ML_IDENTIFICADO
    assert control.movimientos[0].subtipo == "Venta por mostrador con Código QR"
    assert control.movimientos[1].categoria == CategoriaEstadoCuentaMp.SIN_CLASIFICACION_COMERCIAL
    assert control.movimientos[2].categoria == CategoriaEstadoCuentaMp.OTRO_INGRESO_NO_ML_IDENTIFICADO
    assert control.movimientos[3].categoria == CategoriaEstadoCuentaMp.SIN_ASOCIACION_SUFICIENTE
    assert len(control.movimientos) == 4 and control.diferencia_cobertura == Decimal("0.00")


def test_excel_real_preserva_ids_fechas_y_no_contiene_formulas_rotas():
    control = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    wb = load_workbook(BytesIO(generar_control_estado_cuenta_mp_excel(control)), data_only=False)
    assert wb.sheetnames == ["MP — Control de saldo", "MP — Composición diaria", "MP — Otros ingresos", "MP — Salidas y ajustes", "MP — Asociados a ML", "MP — Sin asociación"]
    ws = wb["MP — Otros ingresos"]
    assert ws["A2"].value == "169679883346" and ws["A2"].number_format == "@"
    assert isinstance(ws["E2"].value, datetime)
    assert not any(isinstance(c.value, str) and c.value.startswith("#") for sheet in wb for row in sheet for c in row)


def test_exportacion_b2_independiente_incluye_metadatos_de_periodos_y_estados():
    from dataclasses import replace
    control = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    control = replace(control, metadatos_procesamiento=(("Período solicitado", "18/07/2026 — 20/07/2026"),
                                                       ("Estado B1", "SIN_ACTIVIDAD_COMERCIAL"),
                                                       ("Estado B2/B3", "COMPLETO")))
    wb = load_workbook(BytesIO(generar_control_estado_cuenta_mp_excel(control)), data_only=True)
    valores = {(r[0].value, r[1].value) for r in wb["MP — Control de saldo"].iter_rows(min_row=2) if r[0].value}
    assert ("Estado B1", "SIN_ACTIVIDAD_COMERCIAL") in valores
    assert ("Estado B2/B3", "COMPLETO") in valores


def _settlement(fila, operacion="OP-1", orden="ORD-1", canal="Mercado Libre", plataforma="Checkout"):
    return SimpleNamespace(id_operacion_mercado_pago=operacion, numero_fila_origen=fila, id_orden=orden, canal_venta=canal, plataforma_cobro=plataforma)


def test_varias_filas_settlement_coherentes_no_son_ambiguas():
    from kiki_control.linking.account_statement import agrupar_settlement_por_operacion
    grupo = agrupar_settlement_por_operacion([_settlement(10), _settlement(11)])["OP-1"]
    assert not grupo.es_ambiguo
    assert grupo.filas_origen == (10, 11)
    assert grupo.ids_orden == ("ORD-1",)
    assert len(grupo.movimientos) == 2


def test_filas_settlement_contradictorias_son_ambiguas():
    from kiki_control.linking.account_statement import agrupar_settlement_por_operacion
    grupo = agrupar_settlement_por_operacion([_settlement(10), _settlement(11, orden="ORD-2")])["OP-1"]
    assert grupo.es_ambiguo and "IDs de orden" in grupo.motivo_ambiguedad


def test_vinculo_canonico_se_construye_por_fila_no_comparando_ids():
    from kiki_control.linking.account_statement import construir_indice_operacion_mp_a_grupo_ml
    resultado = SimpleNamespace(id_grupo_canonico="PACK-CANONICO", tiene_mercado_libre_oficial=True, filas_origen_mp=(10, 11))
    reporte = SimpleNamespace(resultados=(resultado,))
    indice = construir_indice_operacion_mp_a_grupo_ml(reporte, [_settlement(10, orden="ORD-A"), _settlement(11, orden="ORD-B")])
    assert indice == {"OP-1": ("PACK-CANONICO",)}


def test_cobertura_se_calcula_y_detecta_faltantes_y_duplicados():
    from dataclasses import replace
    from kiki_control.domain.account_statement import ControlEstadoCuentaMp
    base = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    inconsistente = ControlEstadoCuentaMp(base.resumen, base.movimientos[:-1] + (base.movimientos[0],))
    assert inconsistente.cantidad_lineas_entrada == 4
    assert inconsistente.cantidad_lineas_clasificadas == 2
    assert inconsistente.cantidad_no_clasificadas == 1
    assert inconsistente.cantidad_clasificadas_mas_de_una_vez == 1
    assert not inconsistente.cobertura_completa
    assert inconsistente.diferencia_cobertura_monetaria != Decimal("0")


def archivo_productivo_anonimizado():
    wb = Workbook(); ws = wb.active; ws.title = "ACCOUNT_STATEMENT"
    ws.append(["INITIAL_BALANCE", "CREDITS", "DEBITS", "FINAL_BALANCE"])
    ws.append([Decimal("26316618.06"), Decimal("9222725.75"), Decimal("-16575362.17"), Decimal("18963981.64")]); ws.append([])
    ws.append(["RELEASE_DATE", "TRANSACTION_TYPE", "REFERENCE_ID", "TRANSACTION_NET_AMOUNT", "PARTIAL_BALANCE"])
    linked_ids = ["169679883346"] + [f"L-{i:03}" for i in range(1, 79)]
    for i in range(646):
        if i == 0:
            ref, tipo, importe = linked_ids[0], "Liquidación de dinero", Decimal("54875.24")
        elif i == 1:
            ref, tipo, importe = linked_ids[1], "Rendimiento", Decimal("9167850.36")
        elif i < 17:
            ref, tipo, importe = linked_ids[i], "Rendimiento", Decimal("0.01")
        elif i == 17:
            ref, tipo, importe = linked_ids[i], "Transferencia enviada", Decimal("-16575362.17")
        elif i < 137:
            ref, tipo, importe = linked_ids[i % 79], "Transferencia enviada", Decimal("0")
        elif i < 139:
            ref, tipo, importe = linked_ids[i % 79], "Liquidación de dinero", Decimal("0")
        else:
            ref, tipo, importe = f"U-{i:03}", "Liquidación de dinero", Decimal("0")
        ws.append([datetime(2026, 7, 28), tipo, ref, importe, None])
    out = BytesIO(); wb.save(out); return out.getvalue()


def test_regresion_anonimizada_cuatro_archivos_productivos():
    resumen = normalizar_estado_cuenta_mp("productivo_anonimizado.xlsx", archivo_productivo_anonimizado())
    linked_ids = ["169679883346"] + [f"L-{i:03}" for i in range(1, 79)]
    settlements = []
    for i in range(139):
        ref = linked_ids[0] if i == 0 else linked_ids[1 + (i - 1) % 78]
        if i == 0:
            settlements.append(_settlement(781, ref, None, "Mercado Pago", "Código QR"))
        else:
                # Esta regresión histórica prueba cobertura sin atribución de
                # canal; la evidencia ML histórica se cubre en pruebas propias.
                settlements.append(_settlement(800 + i, ref, f"ORD-{linked_ids.index(ref):03}", None, None))
    control = controlar_estado_cuenta_mp(resumen, settlements)
    assert (len(resumen.movimientos), resumen.saldo_inicial, resumen.creditos_informados, resumen.debitos_informados, resumen.saldo_final_informado) == (646, Decimal("26316618.06"), Decimal("9222725.75"), Decimal("-16575362.17"), Decimal("18963981.64"))
    assert resumen.diferencia_control == Decimal("0.00")
    assert control.lineas_vinculadas == 139
    assert control.operaciones_settlement_vinculadas == 79
    assert control.lineas_sin_vinculo_settlement == 507
    categorias = {categoria: control.estadisticas_categoria(categoria) for categoria in CategoriaEstadoCuentaMp}
    assert [categorias[c].cantidad_movimientos for c in CategoriaEstadoCuentaMp] == [0, 17, 120, 509]
    vinculadas = {categoria: control.estadisticas_vinculadas_categoria(categoria) for categoria in CategoriaEstadoCuentaMp}
    assert [vinculadas[c].cantidad_movimientos for c in CategoriaEstadoCuentaMp] == [0, 17, 120, 2]
    assert sum(v.cantidad_movimientos for v in vinculadas.values()) == control.lineas_vinculadas
    assert sum((v.impacto_neto for v in vinculadas.values()), Decimal("0")) == control.importe_neto_lineas_vinculadas
    assert control.estadisticas_estado(EstadoVinculacionEstadoCuentaMp.VINCULADO_SIN_ORIGEN_COMERCIAL).cantidad_movimientos == 2
    assert categorias[CategoriaEstadoCuentaMp.SIN_ASOCIACION_SUFICIENTE].cantidad_movimientos == 507 + 2
    assert [categorias[c].impacto_neto for c in CategoriaEstadoCuentaMp] == [Decimal("0"), Decimal("9222725.75"), Decimal("-16575362.17"), Decimal("0")]
    assert sum((categorias[c].impacto_neto for c in CategoriaEstadoCuentaMp), Decimal("0")) == resumen.variacion_neta
    qr = next(m for m in control.movimientos if m.movimiento.reference_id == "169679883346")
    assert qr.fila_settlement == 781 and qr.movimiento.importe_neto == Decimal("54875.24")
    assert qr.categoria == CategoriaEstadoCuentaMp.OTRO_INGRESO_NO_ML_IDENTIFICADO
    assert qr.subtipo == "Venta por mostrador con Código QR"
    assert control.cobertura_completa


def test_exportacion_incluye_desgloses_de_vinculacion_e_importes():
    resumen = normalizar_estado_cuenta_mp("a.xlsx", archivo_estado())
    control = controlar_estado_cuenta_mp(resumen, [settlement()])
    wb = load_workbook(BytesIO(generar_control_estado_cuenta_mp_excel(control)), data_only=True)
    composicion = tuple(cell.value for row in wb["MP — Composición diaria"] for cell in row)
    sin_asociacion = tuple(cell.value for row in wb["MP — Sin asociación"] for cell in row)
    assert "Importes por categoría" in composicion
    assert "Composición de las líneas vinculadas al settlement" in composicion
    assert "Estados de vinculación" in composicion
    assert "VINCULADO_SIN_ORIGEN_COMERCIAL" in sin_asociacion


def test_textos_b1_b2_distinguen_ventas_de_movimientos_de_saldo():
    from kiki_control.presentation.account_statement_view import aclaracion_b1_b2, aclaracion_sin_movimientos_ml
    assert "37 grupos conciliados" in aclaracion_b1_b2(37)
    assert "37 grupos conciliados" in aclaracion_sin_movimientos_ml(37)
    assert "611" not in aclaracion_b1_b2(37)
    assert "611" not in aclaracion_sin_movimientos_ml(37)


def test_acciones_recomendadas_se_diferencian_por_estado_de_vinculacion():
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp

    def resumen(reference_id="REF", tipo="Liquidación de dinero", importe=Decimal("1")):
        movimiento = MovimientoEstadoCuentaMp(5, datetime(2026, 7, 28), tipo, reference_id, importe, None, "hash", "ACCOUNT_STATEMENT")
        return ResumenEstadoCuentaMp(Decimal("0"), Decimal("0"), Decimal("0"), importe, (movimiento,), datetime(2026, 7, 28), datetime(2026, 7, 28))

    sin_vinculo = controlar_estado_cuenta_mp(resumen(), []).movimientos[0]
    vinculado_sin_origen = controlar_estado_cuenta_mp(resumen(), [_settlement(10, "REF", canal=None, plataforma=None)]).movimientos[0]
    ambiguo = controlar_estado_cuenta_mp(resumen(), [_settlement(10, "REF", "A"), _settlement(11, "REF", "B")]).movimientos[0]
    id_vacio = controlar_estado_cuenta_mp(resumen(reference_id=None), []).movimientos[0]
    identificado = controlar_estado_cuenta_mp(resumen("QR"), [_settlement(12, "QR", None, "Mercado Pago", "Código QR")]).movimientos[0]

    assert sin_vinculo.accion_recomendada == "Revisar con un settlement que cubra la fecha de origen."
    assert vinculado_sin_origen.accion_recomendada == "Revisar el detalle comercial de la operación o ampliar la información del canal de cobro."
    assert ambiguo.accion_recomendada == "Revisar las filas settlement contradictorias y definir la operación comercial correcta."
    assert id_vacio.accion_recomendada == "Completar o verificar el reference ID en el estado de cuenta."
    assert identificado.accion_recomendada == "Sin acción; conservar para trazabilidad."


def test_prioridad_reintegros_comision_positivos_sin_settlement():
    """La semántica explícita prevalece sobre la palabra genérica comisión."""
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp

    tipos = ("Reintegro de comisión", "Reintegro comision", "Reintegro de comisiones",
             "Devolución de comisión cobrada")
    movimientos = tuple(
        MovimientoEstadoCuentaMp(i, datetime(2026, 1, 1), tipo, f"R-{i}", Decimal("1"),
                                 None, "hash", "ACCOUNT_STATEMENT")
        for i, tipo in enumerate(tipos, 1)
    )
    resumen = ResumenEstadoCuentaMp(Decimal("0"), Decimal("4"), Decimal("0"), Decimal("4"),
                                    movimientos, datetime(2026, 1, 1), datetime(2026, 1, 1))
    control = controlar_estado_cuenta_mp(resumen, [])

    assert all(m.categoria == CategoriaEstadoCuentaMp.OTROS_INGRESOS_NO_ML for m in control.movimientos)
    assert all(m.subtipo == "Reintegro de comisiones" for m in control.movimientos)
    assert all("Account Statement" in m.motivo for m in control.movimientos)
    assert control.cobertura_comercial_completa


def test_salida_requiere_evidencia_y_ml_prevalece_sobre_comision():
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp

    movimientos = (
        MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), "Transferencia enviada", "OUT", Decimal("-2"), None, "h", "S"),
        MovimientoEstadoCuentaMp(2, datetime(2026, 1, 1), "Movimiento", "UNKNOWN", Decimal("-3"), None, "h", "S"),
        MovimientoEstadoCuentaMp(3, datetime(2026, 1, 1), "Comisión", "ML", Decimal("-4"), None, "h", "S"),
    )
    resumen = ResumenEstadoCuentaMp(Decimal("9"), Decimal("0"), Decimal("-9"), Decimal("0"), movimientos,
                                    datetime(2026, 1, 1), datetime(2026, 1, 1))
    settlement_ml = _settlement(3, "ML", "ORD", "Mercado Libre", "Checkout")
    control = controlar_estado_cuenta_mp(resumen, [settlement_ml], {"ML": "G-ML"})

    assert control.movimientos[0].categoria == CategoriaEstadoCuentaMp.SALIDAS_NO_ML
    assert control.movimientos[1].categoria == CategoriaEstadoCuentaMp.SIN_CLASIFICACION_COMERCIAL
    assert control.movimientos[2].categoria == CategoriaEstadoCuentaMp.MOVIMIENTOS_ASOCIADOS_A_ML
    assert not control.cobertura_comercial_completa


def test_settlement_grande_retiene_solo_ids_requeridos_antes_de_agrupar():
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp

    movimiento = MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), "Pago a proveedor", "NEEDED",
                                          Decimal("-1"), None, "h", "S")
    resumen = ResumenEstadoCuentaMp(Decimal("1"), Decimal("0"), Decimal("-1"), Decimal("0"),
                                    (movimiento,), datetime(2026, 1, 1), datetime(2026, 1, 1))
    settlement_grande = tuple(_settlement(i, f"OTHER-{i}", canal=None, plataforma=None)
                              for i in range(1, 2001)) + (_settlement(3000, "NEEDED", canal=None, plataforma=None),)
    control = controlar_estado_cuenta_mp(resumen, settlement_grande)
    metricas = dict(control.metadatos_procesamiento)

    assert metricas["IDs requeridos por Statement"] == "1"
    assert metricas["IDs encontrados en Settlement"] == "1"
    assert metricas["Filas Settlement retenidas para B2/B3"] == "1"


def test_mensaje_cobertura_distingue_causas_y_resume_combinaciones():
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
    from kiki_control.presentation.account_statement_view import mensaje_cobertura_comercial_parcial

    def control_para(referencias, settlements=()):
        movimientos = tuple(
            MovimientoEstadoCuentaMp(i, datetime(2026, 7, 28), "Movimiento", referencia,
                                     Decimal("-1"), None, "h", "S")
            for i, referencia in enumerate(referencias, 1)
        )
        resumen = ResumenEstadoCuentaMp(Decimal(len(movimientos)), Decimal("0"),
                                        -Decimal(len(movimientos)), Decimal("0"), movimientos,
                                        datetime(2026, 7, 28), datetime(2026, 7, 28))
        return controlar_estado_cuenta_mp(resumen, settlements)

    falta = mensaje_cobertura_comercial_parcial(control_para(("FALTA",)))
    vinculado = mensaje_cobertura_comercial_parcial(
        control_para(("VINC",), (_settlement(10, "VINC", canal=None, plataforma=None),))
    )
    ambiguo = mensaje_cobertura_comercial_parcial(control_para(
        ("AMB",), (_settlement(11, "AMB", "A"), _settlement(12, "AMB", "B"))))
    vacio = mensaje_cobertura_comercial_parcial(control_para((None,)))
    combinado = mensaje_cobertura_comercial_parcial(control_para(
        ("FALTA", "VINC", "AMB", None),
        (_settlement(20, "VINC", canal=None, plataforma=None),
         _settlement(21, "AMB", "A"), _settlement(22, "AMB", "B"))))

    assert "período de origen más amplio" in falta
    assert vinculado == ("La cobertura comercial es parcial porque existen movimientos vinculados al "
                         "Settlement cuya evidencia no permite determinar responsablemente el origen comercial.")
    assert "más amplio" not in vinculado
    assert "evidencia contradictoria" in ambiguo
    assert "referencia vacía" in vacio
    assert "varias causas" in combinado
    assert all(estado.value in combinado for estado in (
        EstadoVinculacionEstadoCuentaMp.SIN_VINCULO_SETTLEMENT,
        EstadoVinculacionEstadoCuentaMp.VINCULADO_SIN_ORIGEN_COMERCIAL,
        EstadoVinculacionEstadoCuentaMp.ID_AMBIGUO,
        EstadoVinculacionEstadoCuentaMp.ID_VACIO,
    ))


@pytest.mark.parametrize("tipo", [
    "Devolución", "Reclamo", "Comisión", "Impuesto", "Retención",
    "Cancelación", "Contracargo", "Ajuste",
])
def test_ajuste_generico_negativo_sin_canal_permanece_pendiente(tipo):
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
    mov = MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), tipo, "REF", Decimal("-1"), None, "h", "S")
    resumen = ResumenEstadoCuentaMp(Decimal("1"), Decimal("0"), Decimal("-1"), Decimal("0"),
                                    (mov,), datetime(2026, 1, 1), datetime(2026, 1, 1))
    clasificado = controlar_estado_cuenta_mp(resumen, []).movimientos[0]
    assert clasificado.categoria == CategoriaEstadoCuentaMp.SIN_CLASIFICACION_COMERCIAL
    assert "no acredita" in clasificado.motivo


@pytest.mark.parametrize("tipo", ["Devolución", "Reclamo", "Comisión", "Impuesto", "Retención"])
def test_ajuste_generico_con_evidencia_ml_permanece_en_ml(tipo):
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
    mov = MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), tipo, "REF", Decimal("-1"), None, "h", "S")
    resumen = ResumenEstadoCuentaMp(Decimal("1"), Decimal("0"), Decimal("-1"), Decimal("0"),
                                    (mov,), datetime(2026, 1, 1), datetime(2026, 1, 1))
    clasificado = controlar_estado_cuenta_mp(
        resumen, [_settlement(1, "REF", "ORD", "Mercado Libre", "Checkout")], {"REF": "G"}
    ).movimientos[0]
    assert clasificado.categoria == CategoriaEstadoCuentaMp.MOVIMIENTOS_ASOCIADOS_A_ML


@pytest.mark.parametrize("tipo", ["Transferencia enviada", "Pago a proveedor", "Retiro", "Extracción", "Envío de dinero"])
def test_salida_financiera_inequivoca_sigue_siendo_salida_no_ml(tipo):
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
    mov = MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), tipo, "REF", Decimal("-1"), None, "h", "S")
    resumen = ResumenEstadoCuentaMp(Decimal("1"), Decimal("0"), Decimal("-1"), Decimal("0"),
                                    (mov,), datetime(2026, 1, 1), datetime(2026, 1, 1))
    assert controlar_estado_cuenta_mp(resumen, []).movimientos[0].categoria == CategoriaEstadoCuentaMp.SALIDAS_NO_ML


@pytest.mark.parametrize(("plataforma", "subtipo"), [
    ("Código QR", "Venta por mostrador con Código QR"),
    ("Point Smart", "Venta con Point"),
    (None, "Ingreso Mercado Pago — medio no determinado"),
    ("Checkout", "Ingreso Mercado Pago — medio no determinado"),
    ("Link de pago", "Ingreso Mercado Pago — medio no determinado"),
])
def test_canal_mercado_pago_no_presupone_qr(plataforma, subtipo):
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
    mov = MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), "Liquidación", "REF", Decimal("1"), None, "h", "S")
    resumen = ResumenEstadoCuentaMp(Decimal("0"), Decimal("1"), Decimal("0"), Decimal("1"),
                                    (mov,), datetime(2026, 1, 1), datetime(2026, 1, 1))
    clasificado = controlar_estado_cuenta_mp(
        resumen, [_settlement(1, "REF", None, "Mercado Pago", plataforma)]
    ).movimientos[0]
    assert clasificado.categoria == CategoriaEstadoCuentaMp.OTROS_INGRESOS_NO_ML
    assert clasificado.subtipo == subtipo


def test_tipo_original_qr_es_evidencia_explicita_sin_plataforma():
    from kiki_control.domain.account_statement import MovimientoEstadoCuentaMp, ResumenEstadoCuentaMp
    mov = MovimientoEstadoCuentaMp(1, datetime(2026, 1, 1), "Cobro con Codigo QR", "REF", Decimal("1"), None, "h", "S")
    resumen = ResumenEstadoCuentaMp(Decimal("0"), Decimal("1"), Decimal("0"), Decimal("1"),
                                    (mov,), datetime(2026, 1, 1), datetime(2026, 1, 1))
    item = controlar_estado_cuenta_mp(resumen, [_settlement(1, "REF", None, "Mercado Pago", None)]).movimientos[0]
    assert item.subtipo == "Venta por mostrador con Código QR"


def test_consolidado_agrega_hojas_solo_cuando_hay_account_statement():
    from kiki_control.exporting import generar_reporte_consolidado_excel
    from tests.test_control_consolidado_diagnostics import r, rep
    reporte = rep([r("grupo")])
    control = controlar_estado_cuenta_mp(normalizar_estado_cuenta_mp("a.xlsx", archivo_estado()), [settlement()])
    sin = load_workbook(BytesIO(generar_reporte_consolidado_excel(reporte)))
    con = load_workbook(BytesIO(generar_reporte_consolidado_excel(reporte, control_estado_cuenta=control)))
    hojas_mp = {"MP — Control de saldo", "MP — Composición diaria", "MP — Otros ingresos", "MP — Salidas y ajustes", "MP — Asociados a ML", "MP — Sin asociación"}
    assert not hojas_mp.intersection(sin.sheetnames)
    assert hojas_mp.issubset(con.sheetnames)


def test_vinculo_canonico_clasifica_ml_y_multiples_grupos_son_ambiguos():
    resumen = normalizar_estado_cuenta_mp("a.xlsx", archivo_estado())
    mov = _settlement(10, "169679883346", "ORD-DISTINTO", "Mercado Libre", "Checkout")
    control_ml = controlar_estado_cuenta_mp(resumen, [mov], {"169679883346": ("GRUPO-CANONICO",)})
    assert control_ml.movimientos[0].categoria == CategoriaEstadoCuentaMp.ASOCIADO_A_VENTA_ML
    assert control_ml.movimientos[0].id_grupo_ml == "GRUPO-CANONICO"
    control_ambiguo = controlar_estado_cuenta_mp(resumen, [mov], {"169679883346": ("GRUPO-A", "GRUPO-B")})
    assert control_ambiguo.movimientos[0].estado_vinculacion == EstadoVinculacionEstadoCuentaMp.ID_AMBIGUO
    assert control_ambiguo.movimientos[0].categoria == CategoriaEstadoCuentaMp.SIN_ASOCIACION_SUFICIENTE
