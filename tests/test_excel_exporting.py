from copy import deepcopy
from dataclasses import replace
from datetime import UTC, date
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import ast

import pytest
from openpyxl import load_workbook
from openpyxl.cell.cell import TYPE_FORMULA

from kiki_control.domain.financial_movement import TipoOperacionFinanciera
from kiki_control.exporting.excel import generar_reporte_completo_excel, generar_reporte_excepciones_excel
from kiki_control.presentation.reconciliation_view import CoberturaArchivosPresentacion, RangoFechasPresentacion, es_excepcion_o_caso_especial
from kiki_control.reconciliation import reconciliar
from tests.test_reconciliation_engine import FECHA, mov, op


def cobertura_sintetica():
    rango = RangoFechasPresentacion(date(2026, 1, 1), date(2026, 1, 2), "01/01/2026 a 02/01/2026")
    return CoberturaArchivosPresentacion(rango, rango, rango, 1, True, None)


def reporte_sintetico():
    operaciones = [
        op(id_orden="00000000000000012345", neto="100", fila=2, utilidad="22"),
        op(id_orden="200", neto="50", fila=3, utilidad="7"),
        op(id_orden="=CMD()", neto="30", fila=4, utilidad="3"),
    ]
    movimientos = [
        mov(id_orden="00000000000000012345", monto="100", fila=20, id_mp="p1"),
        mov(id_orden="200", monto="49.50", fila=21, id_mp="p2"),
        mov(id_orden="=CMD()", monto="-30", tipo=TipoOperacionFinanciera.DEVOLUCION_DINERO, fila=22, id_mp="dev"),
        mov(id_orden=None, monto="-10", tipo=TipoOperacionFinanciera.PAYOUT, fila=23, id_mp="pay"),
        mov(id_orden="300", monto="5", fila=24, id_mp="p3", liquidado=False),
    ]
    return replace(reconciliar(operaciones, movimientos, tolerancia=Decimal("0.01")), fecha_procesamiento_utc=FECHA)


def cargar(bytes_xlsx):
    return load_workbook(BytesIO(bytes_xlsx), data_only=False)


def mapa_resumen(ws):
    return {row[0].value: row[1].value for row in ws.iter_rows(min_row=2, max_col=2)}


def celda_resumen(ws, etiqueta):
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[0].value == etiqueta:
            return row[1]
    raise AssertionError(f"No existe la fila de resumen {etiqueta!r}")


def filas_operaciones(ws):
    headers = [cell.value for cell in ws[1]]
    return [dict(zip(headers, row, strict=False)) for row in ws.iter_rows(min_row=2, values_only=True)]


def test_reportes_xlsx_validos_hojas_y_contenido_en_memoria(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    antes = set(Path.cwd().iterdir())
    reporte = reporte_sintetico()
    completo = generar_reporte_completo_excel(reporte, cobertura_sintetica(), "America/Argentina/Cordoba")
    excepciones = generar_reporte_excepciones_excel(reporte, cobertura_sintetica(), "America/Argentina/Cordoba")
    assert completo.startswith(b"PK") and excepciones.startswith(b"PK")
    assert set(Path.cwd().iterdir()) == antes
    wb = cargar(completo)
    wb_exc = cargar(excepciones)
    assert wb.sheetnames == ["Resumen", "Todas las operaciones", "Excepciones"]
    assert wb_exc.sheetnames == ["Resumen", "Excepciones"]
    assert wb["Todas las operaciones"].max_row - 1 == len(reporte.resultados)
    esperadas = sum(1 for r in reporte.resultados if es_excepcion_o_caso_especial(r))
    assert wb["Excepciones"].max_row - 1 == esperadas
    assert wb_exc["Excepciones"].max_row - 1 == esperadas


def test_resumen_contiene_contexto_kpis_aclaraciones_y_no_filtra_pii():
    ws = cargar(generar_reporte_completo_excel(reporte_sintetico(), cobertura_sintetica(), "America/Argentina/Cordoba"))["Resumen"]
    resumen = mapa_resumen(ws)
    assert resumen["Nombre"] == "Kiki Control Financiero"
    assert resumen["Tipo de reporte"] == "Reporte completo"
    assert resumen["Fecha y hora del procesamiento (zona operativa)"].hour == 21
    assert resumen["Zona horaria operativa"] == "America/Argentina/Cordoba"
    assert resumen["Versión de la regla de conciliación"] == "ML_MP_ID_ORDER_NETO_V1"
    assert Decimal(str(resumen["Tolerancia aplicada"])) == Decimal("0.01")
    assert resumen["Cobertura de ventas ML"] == "01/01/2026 a 02/01/2026"
    assert "Se compararon" in resumen["Conclusión ejecutiva"]
    assert "Comparables" in resumen and "Diferencia comparable" in resumen
    assert "la utilidad es informada" in " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None).lower()
    assert "movimientos de fondos" in " ".join(str(v).lower() for v in resumen.values())
    contenido = " ".join(str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None).lower()
    for prohibido in ("sha", "hash", "json", "archivo_real", "documento", "tarjeta", "pagador"):
        assert prohibido not in contenido


def test_columnas_tipos_formatos_orden_excepciones_y_formulas_seguras():
    reporte = reporte_sintetico()
    wb = cargar(generar_reporte_completo_excel(reporte, cobertura_sintetica(), "UTC"))
    ws = wb["Todas las operaciones"]
    headers = [c.value for c in ws[1]]
    assert headers == [
        "ID de orden", "Estado", "Neto informado ML", "Neto aprobado MP", "Diferencia", "Neto financiero total", "Utilidad informada ML", "Pago dividido", "Devolución", "Reclamo o disputa", "Pendiente de acreditación", "Requiere revisión", "Explicación", "Motivos técnicos", "Filas ML de origen", "Filas MP de origen", "Cantidad de pagos aprobados", "Cantidad de movimientos financieros", "Versión de regla", "Tolerancia aplicada",
    ]
    filas = filas_operaciones(ws)
    assert [f["ID de orden"] for f in filas] == ["00000000000000012345", "200", "300", "'=CMD()", "movimiento_de_fondos-fila-23"]
    assert ws["A2"].data_type == "s" and ws["A2"].value == "00000000000000012345"
    formula_cell = next(cell for cell in ws["A"] if cell.value == "'=CMD()")
    assert formula_cell.data_type != TYPE_FORMULA
    assert ws["C2"].data_type == "n" and Decimal(str(ws["C2"].value)) == Decimal("100")
    assert ws["F5"].value < 0
    assert ws["C4"].value is None
    assert ws["H2"].value in {"Sí", "No"}
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws["C2"].number_format
    exc_ids = [f["ID de orden"] for f in filas_operaciones(wb["Excepciones"])]
    assert "movimiento_de_fondos-fila-23" in exc_ids
    assert "00000000000000012345" not in exc_ids


def test_resumen_aplica_tipos_y_formatos_por_campo():
    ws = cargar(generar_reporte_completo_excel(reporte_sintetico(), cobertura_sintetica(), "America/Argentina/Cordoba"))["Resumen"]
    for etiqueta in ("Comparables", "Movimientos sin fecha de liquidación", "Cantidad de filas incluidas en Todas las operaciones", "Cantidad de filas incluidas en Excepciones"):
        celda = celda_resumen(ws, etiqueta)
        assert isinstance(celda.value, int)
        assert celda.data_type == "n"
        assert "$" not in celda.number_format

    for etiqueta in ("Utilidad informada ML", "Neto ML comparable", "Neto MP comparable", "Diferencia comparable", "Neto MP fuera del archivo ML", "Tolerancia aplicada"):
        celda = celda_resumen(ws, etiqueta)
        assert celda.data_type == "n"
        assert not isinstance(celda.value, str)
        assert "$" not in str(celda.value)
        assert "$" in celda.number_format

    diferencia = celda_resumen(ws, "Diferencia comparable")
    assert Decimal(str(diferencia.value)) < Decimal("0")
    assert "$" not in celda_resumen(ws, "Conclusión ejecutiva").number_format
    fecha = celda_resumen(ws, "Fecha y hora del procesamiento (zona operativa)")
    assert fecha.is_date
    assert fecha.number_format == "dd/mm/yyyy hh:mm:ss"



@pytest.mark.parametrize("id_malicioso", ["=CMD()", "+SUMA", "-ORDEN", "@INDICE"])
def test_ids_sinteticos_peligrosos_no_son_formula_y_conservan_valor_visible(id_malicioso):
    reporte = replace(reconciliar([op(id_orden=id_malicioso)], [mov(id_orden=id_malicioso)], tolerancia=Decimal("0.01")), fecha_procesamiento_utc=FECHA)
    ws = cargar(generar_reporte_completo_excel(reporte, cobertura_sintetica(), "UTC"))["Todas las operaciones"]
    celda = ws["A2"]
    assert celda.data_type != TYPE_FORMULA
    assert celda.value == f"'{id_malicioso}"
    assert ws["C2"].data_type == "n"

def test_exportacion_no_usa_float_y_no_muta_dominio():
    import kiki_control.exporting.excel as excel

    arbol = ast.parse(Path(excel.__file__).read_text())
    assert not any(isinstance(node, ast.Constant) and isinstance(node.value, float) for node in ast.walk(arbol))
    reporte = reporte_sintetico()
    antes = deepcopy(reporte)
    generar_reporte_excepciones_excel(reporte, cobertura_sintetica(), "UTC")
    assert reporte == antes
