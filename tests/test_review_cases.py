from dataclasses import asdict, replace
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from kiki_control.domain.financial_movement import TipoOperacionFinanciera
from kiki_control.domain.reconciliation import EstadoConciliacion, MotivoConciliacion
from kiki_control.exporting import generar_revisiones_pendientes_excel
from kiki_control.presentation.review_cases import TipoRevision, caso_a_fila, clasificar_revision, clasificar_revisiones, conteo_por_tipo, filas_revisiones, filtrar_casos
from kiki_control.reconciliation import reconciliar
from tests.test_reconciliation_engine import FECHA, mov, op


def _uno(ops, movs):
    return reconciliar(ops, movs).resultados[0]


def _tipo(resultado):
    caso = clasificar_revision(resultado)
    assert caso is not None
    return caso.tipo


def test_clasifica_categorias_sinteticas_y_respaldo():
    assert _tipo(_uno([], [mov(id_orden=None, tipo=TipoOperacionFinanciera.CASHBACK)])) == TipoRevision.MP_SIN_ID_ORDEN
    assert _tipo(_uno([], [mov(id_orden="9")])) == TipoRevision.ORDEN_MP_SIN_VENTA_ML
    assert _tipo(_uno([op()], [])) == TipoRevision.VENTA_ML_SIN_MOVIMIENTO_MP
    assert _tipo(_uno([op()], [mov(), mov(tipo=TipoOperacionFinanciera.RECLAMO, monto="-1", id_mp="r")])) == TipoRevision.RECLAMO_O_DISPUTA
    assert _tipo(_uno([op()], [mov(), mov(tipo=TipoOperacionFinanciera.DISPUTA_ENVIO, monto="-1", id_mp="d")])) == TipoRevision.RECLAMO_O_DISPUTA
    assert _tipo(_uno([op()], [mov(), mov(tipo=TipoOperacionFinanciera.DESCONOCIDA, monto="1", id_mp="x")])) == TipoRevision.MOVIMIENTO_DESCONOCIDO_EN_REVISION
    assert _tipo(_uno([op(fila=1), op(fila=2, hash_="h2")], [mov()])) == TipoRevision.DUPLICACION_COMERCIAL
    assert _tipo(_uno([op()], [mov(id_mp="z"), mov(id_mp="z", fila=11)])) == TipoRevision.DUPLICACION_FINANCIERA
    base = _uno([op()], [])
    respaldo = replace(base, estado=EstadoConciliacion.CONCILIADA_CON_DIFERENCIA, motivos=(MotivoConciliacion.DIFERENCIA_SUPERA_TOLERANCIA,), requiere_revision=True, cantidad_operaciones_comerciales=0, cantidad_movimientos_financieros=0)
    assert _tipo(respaldo) == TipoRevision.OTRA_REVISION


def test_prioridad_con_varias_condiciones_e_invariancia_kpi():
    resultado = _uno([op(fila=1), op(fila=2, hash_="h2")], [mov(), mov(tipo=TipoOperacionFinanciera.DESCONOCIDA, monto="1", id_mp="x", fila=12)])
    caso = clasificar_revision(resultado)
    assert caso.tipo == TipoRevision.DUPLICACION_COMERCIAL
    assert TipoRevision.MOVIMIENTO_DESCONOCIDO_EN_REVISION in caso.condiciones_adicionales
    reporte = reconciliar([op(id_orden="1"), op(id_orden="2")], [mov(id_orden="1"), mov(id_orden="3"), mov(id_orden=None, tipo=TipoOperacionFinanciera.CASHBACK, id_mp="c")])
    casos = clasificar_revisiones(reporte.resultados)
    assert sum(conteo_por_tipo(casos).values()) == sum(1 for r in reporte.resultados if r.requiere_revision)


def test_tabla_segura_filtros_y_sin_mutar_dominio_sin_floats():
    peligro = _uno([], [mov(id_orden="=PELIGRO")])
    antes = asdict(peligro)
    casos = clasificar_revisiones((peligro,))
    filas = filas_revisiones(casos)
    assert "Pagador" not in asdict(filas[0]) and "Documento" not in asdict(filas[0]) and "Tarjeta" not in asdict(filas[0])
    assert filtrar_casos(casos, TipoRevision.ORDEN_MP_SIN_VENTA_ML, "PELIGRO") == list(casos)
    assert asdict(peligro) == antes
    assert not any(isinstance(v, float) for v in asdict(peligro).values())


def test_descarga_revisiones_solo_revisiones_ids_texto_decimales_formula_y_sin_pii():
    reporte = reconciliar([], [mov(id_orden="12345678901234567890"), mov(id_orden="=FORMULA", id_mp="x")])
    data = generar_revisiones_pendientes_excel(reporte, None, "UTC")
    wb = load_workbook(BytesIO(data), data_only=False)
    assert wb.sheetnames == ["Resumen", "Revisiones pendientes"]
    ws = wb["Revisiones pendientes"]
    rows = list(ws.iter_rows(values_only=False))
    assert len(rows) == 3  # header + only resultados requiere_revision
    assert rows[1][0].data_type == "s" and rows[1][0].value == "12345678901234567890"
    id_cell = rows[2][0]
    assert id_cell.data_type == "s" and str(id_cell.value).startswith("'=FORMULA")
    assert isinstance(rows[1][6].value, Decimal | int) or rows[1][6].value == 100
    headers = [c.value for c in rows[0]]
    forbidden = {"Pagador", "Documento", "Tarjeta", "Datos extra", "Hash"}
    assert forbidden.isdisjoint(headers)


def test_modulo_puro_no_usa_dependencias_prohibidas_y_no_archivos_reales():
    import kiki_control.presentation.review_cases as rc
    import kiki_control.reconciliation.engine as engine

    assert "streamlit" not in rc.__dict__ and "pandas" not in rc.__dict__ and "openpyxl" not in rc.__dict__
    assert "review_cases" not in engine.__dict__


def test_excel_revisiones_referencia_amigable_para_movimiento_sin_id_orden():
    reporte = reconciliar([], [mov(id_orden=None, tipo=TipoOperacionFinanciera.CASHBACK, fila=77, id_mp="mp-sintetico")])
    caso = clasificar_revisiones(reporte.resultados)[0]
    referencia_esperada = caso_a_fila(caso).id_orden_o_referencia

    data = generar_revisiones_pendientes_excel(reporte, None, "UTC")
    ws = load_workbook(BytesIO(data), data_only=False)["Revisiones pendientes"]
    celda = ws["A2"]

    assert referencia_esperada == "Movimiento MP sin ID de orden — referencia interna fila 77"
    assert celda.value == referencia_esperada
    assert celda.data_type == "s"
    assert "movimiento_sin_operacion_comercial-fila-77" not in celda.value
    valores = [cell.value for row in ws.iter_rows() for cell in row]
    assert not any("hash" in str(v).lower() for v in valores if v is not None)
    assert not any(pii in str(v).lower() for v in valores if v is not None for pii in ("pagador", "documento", "tarjeta", "datos extra"))
