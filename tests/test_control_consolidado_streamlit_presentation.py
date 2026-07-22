from decimal import Decimal

from kiki_control.presentation.control_consolidado_view import (
    FilaControlConsolidado,
    filtrar_filas_consolidadas,
    formato_importe,
    kpis_consolidados,
)
from kiki_control.ui.session_cycle import (
    construir_firma_procesamiento_tres_fuentes,
    invalidar_resultados_conocidos,
    limpiar_filtros_de_vista,
)


def test_firma_tres_fuentes_estable_y_sensible_a_cambios_o_retiros():
    base = construir_firma_procesamiento_tres_fuentes("ml", "eco", "mp", "UTC", Decimal("0.01"))
    assert base == construir_firma_procesamiento_tres_fuentes("ml", "eco", "mp", "UTC", Decimal("0.010"))
    assert base != construir_firma_procesamiento_tres_fuentes("ml2", "eco", "mp", "UTC", Decimal("0.01"))
    assert base != construir_firma_procesamiento_tres_fuentes("ml", None, "mp", "UTC", Decimal("0.01"))
    assert base != construir_firma_procesamiento_tres_fuentes("ml", "eco", "mp", "America/Argentina/Cordoba", Decimal("0.01"))
    assert base != construir_firma_procesamiento_tres_fuentes("ml", "eco", "mp", "UTC", Decimal("1"))


def test_invalidacion_completa_y_limpieza_de_filtros_sin_borrar_reportes():
    estado = {k: object() for k in ("normalizacion", "cobertura_consolidada", "reporte_comercial", "reporte_financiero", "reporte_consolidado", "firma_procesamiento", "filtro_estados", "detalle_operacion")}
    invalidar_resultados_conocidos(estado)
    assert not estado
    estado = {"reporte_consolidado": object(), "reporte_financiero": object(), "filtro_estados": ["x"], "detalle_operacion": "a"}
    limpiar_filtros_de_vista(estado)
    assert "reporte_consolidado" in estado and "reporte_financiero" in estado
    assert "filtro_estados" not in estado and "detalle_operacion" not in estado


def test_streamlit_no_duplica_formulas_financieras_y_usa_apis_existentes():
    source = open("src/kiki_control/ui/streamlit_app.py", encoding="utf-8").read()
    for api in ("normalizar_ventas_mercado_libre", "normalizar_mercado_libre", "normalizar_mercado_pago", "vincular_ventas_oficiales_con_eccomapp", "reconciliar", "consolidar_control_financiero"):
        assert api in source
    assert "Utilidad preliminar =" not in source
    assert "Diferencia =" not in source
    assert "Auditoría de conciliación Eccomapp–Mercado Pago" in source
    assert "Descargar reporte completo" in source


def test_presentacion_sin_float_y_negativos_con_signo():
    assert formato_importe(Decimal("-10.50")).startswith("$ -10,50")
    source = open("src/kiki_control/presentation/control_consolidado_view.py", encoding="utf-8").read()
    assert "float(" not in source


def test_filtros_estados_diferencias_faltantes_y_revision():
    filas = [
        FilaControlConsolidado("a", "ORD-1", "Completa", "COMPLETA", "ML oficial, Eccomapp, MP", "$ 1,00", "$ 0,00", "$ 0,00", "$ 1,00", "$ 1,00", "$ 1,00", "$ 1,00", "$ 0,00", "$ 0,00", "No", False, False),
        FilaControlConsolidado("b", "ORD-2", "Sin Mp", "SIN_MOVIMIENTO_FINANCIERO", "ML oficial, Eccomapp", "$ 1,00", "$ 0,00", "$ 0,00", "$ 1,00", "$ 1,00", "No calculado", "No calculado", "No calculado", "$ 0,00", "Sí", False, True),
        FilaControlConsolidado("c", "ORD-3", "Con Diferencia", "CON_DIFERENCIA", "ML oficial, Eccomapp, MP", "$ 1,00", "$ 0,00", "$ 0,00", "$ 1,00", "$ 1,00", "$ 2,00", "$ 2,00", "$ 1,00", "$ 0,00", "Sí", True, False),
    ]
    assert [f.clave for f in filtrar_filas_consolidadas(filas, {"SIN_MOVIMIENTO_FINANCIERO"}, "", False, False, False)] == ["b"]
    assert [f.clave for f in filtrar_filas_consolidadas(filas, set(), "ORD-3", False, True, False)] == ["c"]
    assert [f.clave for f in filtrar_filas_consolidadas(filas, set(), "", True, False, True)] == ["b"]


def test_kpis_comparables_utilidad_parcial_y_mp_sin_ml():
    from kiki_control.domain.control_consolidado import EstadoControlConsolidado, IndicadoresFinancieros, ReporteControlConsolidado, ResultadoControlConsolidado
    ind = IndicadoresFinancieros(False, False, False, False, False, False, False, False)
    def res(clave, ml, mp, costo, dif, utilidad, tiene_ml=True):
        return ResultadoControlConsolidado(clave, clave, (clave,), tiene_ml, costo is not None, mp is not None, None, ml, Decimal("-1") if ml is not None else None, None, None, None, None, ml, None, costo, None, None, None, mp, mp, None, None, None, None, None, None, dif, utilidad, Decimal("0.01"), EstadoControlConsolidado.COMPLETA, False, (), (), ind, "v", (), (), (), (), (), (), (), ())
    resultados = (res("a", Decimal("100"), Decimal("90"), Decimal("40"), Decimal("-10"), Decimal("60")), res("b", Decimal("50"), None, None, None, None), res("c", None, Decimal("30"), None, None, None, False))
    reporte = ReporteControlConsolidado(resultados, "v", Decimal("0.01"), 0, 0, 3, 0, 1, 0, 0, 1, 1, 1, 0, 0, Decimal("150"), Decimal("120"), Decimal("40"))
    bloques = kpis_consolidados(reporte)
    b = {k.nombre: k.valor for ks in bloques.values() for k in ks}
    assert b["Neto ML comparable"] == "$ 100,00"
    assert b["Neto MP comparable"] == "$ 90,00"
    assert b["Diferencia comparable ML–MP"] == "$ -10,00"
    assert b["Neto MP sin venta oficial asociada"] == "$ 30,00"
    assert b["Utilidad preliminar calculable"] == "$ 60,00"
    assert b["Cobertura de utilidad"] == "1 de 3"


def test_integral_tres_fuentes_sintetico_en_memoria_presentacion_y_cobertura():
    from csv import DictWriter
    from io import StringIO

    from kiki_control.adapters.mercado_libre import normalizar_mercado_libre
    from kiki_control.adapters.mercado_libre_ventas import normalizar_ventas_mercado_libre
    from kiki_control.adapters.mercado_pago import normalizar_mercado_pago
    from kiki_control.linking.commercial import vincular_ventas_oficiales_con_eccomapp
    from kiki_control.linking.control_financiero import consolidar_control_financiero
    from kiki_control.presentation.control_consolidado_view import (
        advertir_periodos_distintos,
        cobertura_tres_fuentes,
        detalle_control,
        explicacion_resultado,
        filas_tabla_consolidada,
        tabla_consolidada,
    )
    from kiki_control.reconciliation import reconciliar
    from tests.test_mercado_libre_ventas_normalization import PII_SENTINELS, fila as fila_ml_oficial, xlsx_ventas
    from tests.test_mercado_pago_normalization import fila as fila_mp, xlsx as xlsx_mp
    from tests.test_streamlit_integration import ML_COLS

    id_orden = "10000000000000000001"
    venta = fila_ml_oficial(id_venta=id_orden, total="120.00")
    venta["Ingresos por productos (ARS)"] = "150.00"
    venta["Cargo por venta e impuestos (ARS)"] = "-20.00"
    venta["Costos de envío (ARS)"] = "-10.00"
    ventas_ml = normalizar_ventas_mercado_libre("ventas_oficiales.xlsx", xlsx_ventas([venta]))

    salida = StringIO()
    writer = DictWriter(salida, fieldnames=ML_COLS)
    writer.writeheader()
    writer.writerow(
        {
            "Fecha de venta": "2026-07-01",
            "Hora": "10:00:00",
            "Producto": "Producto sintético sin datos reales",
            "Sku": "SKU-SINTETICO-0001",
            "ID Order": id_orden,
            "Cantidad": "1",
            "Monto de venta ($)": "150,00",
            "Costo Total (Con IVA) ($)": "70,00",
            "Comisión MeLi  ($)": "20,00",
            "Costo de envío (Seller) ($)": "10,00",
            "Monto neto (en MP) ($)": "115,00",
            "Utilidades netas ($)": "45,00",
            "Parámetros cálculo": "Costo inc. alíc.: Sí | Precio inc. alíc.: Sí | IIBB: 0 (0%)",
        }
    )
    eccomapp = normalizar_mercado_libre("eccomapp.csv", salida.getvalue().encode("utf-8"))

    mercado_pago = normalizar_mercado_pago(
        "mercado_pago.xlsx",
        xlsx_mp(
            [
                fila_mp(
                    **{
                        "ID DE OPERACIÓN EN MERCADO PAGO": "mp-sintetico-1",
                        "ID DE LA ORDEN": id_orden,
                        "FECHA DE ORIGEN": "2026-07-01T13:00:00.000-03:00",
                        "FECHA DE APROBACIÓN": "2026-07-01T13:01:00.000-03:00",
                        "FECHA DE LIQUIDACIÓN DEL DINERO": "2026-07-10T13:00:00.000-03:00",
                        "VALOR DE LA COMPRA": "150.00",
                        "MONTO NETO DE LA OPERACIÓN QUE IMPACTÓ TU DINERO": "118.00",
                        "MONTO NETO DE LA OPERACIÓN": "118.00",
                        "COMISIONES + IVA": "-20.00",
                        "COSTO DE ENVÍO": "-10.00",
                    }
                )
            ]
        ),
    )

    assert ventas_ml.cantidad_normalizada == eccomapp.cantidad_normalizada == mercado_pago.cantidad_normalizada == 1
    reporte_comercial = vincular_ventas_oficiales_con_eccomapp(ventas_ml.ventas, eccomapp.operaciones)
    reporte_financiero = reconciliar(eccomapp.operaciones, mercado_pago.movimientos, Decimal("0.01"))
    consolidado = consolidar_control_financiero(reporte_comercial, reporte_financiero)
    resultado = consolidado.resultados[0]

    assert resultado.monto_venta_ml == Decimal("150.00")
    assert resultado.total_informado_ml == Decimal("120.00")
    assert resultado.cargo_venta_impuestos_ml == Decimal("-20.00")
    assert resultado.costo_envio_ml == Decimal("-10.00")
    assert resultado.costo_productos_eccomapp == Decimal("70.00")
    assert resultado.neto_aprobado_mp == Decimal("118.00")
    assert resultado.diferencia_ml_mp == Decimal("-2.00")
    assert resultado.utilidad_preliminar_control == Decimal("50.00")

    detalle = detalle_control(resultado)
    assert detalle["Venta ML oficial"] == "$ 150,00"
    assert detalle["Neto esperado ML"] == "$ 120,00"
    assert detalle["Cargos e impuestos ML"] == "$ -20,00"
    assert detalle["Costo de envío ML"] == "$ -10,00"
    assert detalle["Costo productos"] == "$ 70,00"
    assert detalle["Neto aprobado MP"] == "$ 118,00"
    assert detalle["Neto financiero total MP"] == "$ 118,00"
    assert detalle["Diferencia ML–MP"] == "$ -2,00"
    assert detalle["Utilidad preliminar"] == "$ 50,00"

    explicaciones = {fila["Concepto"]: fila for fila in explicacion_resultado(resultado)}
    assert explicaciones["Venta ML oficial"]["Columna utilizada"] == "Ingresos por productos (ARS)"
    assert explicaciones["Cargos e impuestos ML"]["Columna utilizada"] == "Cargo por venta e impuestos (ARS)"
    assert explicaciones["Costo de envío ML"]["Columna utilizada"] == "Costos de envío (ARS)"
    assert explicaciones["Neto esperado ML"]["Columna utilizada"] == "Total (ARS)"
    assert explicaciones["Costo de productos"]["Columna utilizada"] == "Costo Total (Con IVA) ($)"
    assert explicaciones["Neto aprobado MP"]["Columna utilizada"] == "MONTO NETO DE LA OPERACIÓN QUE IMPACTÓ TU DINERO"
    assert explicaciones["Diferencia ML–MP"]["Regla aplicada"] == "neto_aprobado_mp - total_informado_ml"
    assert explicaciones["Utilidad preliminar"]["Regla aplicada"] == "total_informado_ml - costo_productos_eccomapp"

    cobertura = cobertura_tres_fuentes(ventas_ml.ventas, eccomapp.operaciones, mercado_pago.movimientos)
    assert not advertir_periodos_distintos(cobertura)
    assert next(c for c in cobertura if c.nombre == "Liquidaciones MP").minimo == "10/07/2026"

    filas = filas_tabla_consolidada(consolidado.resultados)
    tabla = tabla_consolidada(filas)
    texto_presentacion = repr(detalle) + repr(explicaciones) + repr(tabla)
    for pii in PII_SENTINELS:
        assert pii not in texto_presentacion
    assert "float(" not in open("src/kiki_control/presentation/control_consolidado_view.py", encoding="utf-8").read()


def test_ayudas_y_column_config_declaran_columnas_externas_exactas():
    source = open("src/kiki_control/presentation/control_consolidado_view.py", encoding="utf-8").read()
    assert "Campo interno: costo_envio_ml. Columna utilizada: Costos de envío (ARS)." in source
    assert "Columnas externas: Ingresos por envío (ARS)" not in source
    ui = open("src/kiki_control/ui/streamlit_app.py", encoding="utf-8").read()
    for expected in (
        "Ingresos por productos (ARS)",
        "Cargo por venta e impuestos (ARS)",
        "Costos de envío (ARS)",
        "Total (ARS)",
        "Costo Total (Con IVA) ($)",
        "MONTO NETO DE LA OPERACIÓN QUE IMPACTÓ TU DINERO",
    ):
        assert expected in ui
    assert "column_config=_column_config_control_consolidado()" in ui

def test_formato_monetario_argentino_cliente_y_fechas_ddmmaaaa():
    from datetime import date
    from kiki_control.presentation.control_consolidado_view import formato_fecha
    assert formato_importe(Decimal('1234.56')) == '$ 1.234,56'
    assert formato_fecha(date(2026, 7, 22)) == '22/07/2026'


def test_exportaciones_consolidadas_tres_descargas_y_hojas():
    from io import BytesIO
    from openpyxl import load_workbook
    from tests.test_control_consolidado_diagnostics import r, rep, D
    from kiki_control.exporting import generar_excepciones_consolidadas_excel, generar_reporte_consolidado_excel, generar_revisiones_consolidadas_excel
    reporte = rep([r('a'), r('b')])
    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(reporte)))
    assert wb.sheetnames == ['Resumen', 'Cobertura y universos', 'Puente de fuentes', 'Control por operación', 'Distribución temporal MP', 'Revisiones', 'Diccionario de cálculos']
    assert load_workbook(BytesIO(generar_excepciones_consolidadas_excel(reporte))).sheetnames == ['Resumen', 'Excepciones']
    assert load_workbook(BytesIO(generar_revisiones_consolidadas_excel(reporte))).sheetnames == ['Resumen', 'Revisiones']
    source = open('src/kiki_control/exporting/excel.py', encoding='utf-8').read().lower()
    assert 'float(' not in source and 'comprador' not in source and 'documento' not in source

def test_streamlit_y_excel_temporal_no_duplican_netos_mp_distintos():
    from io import BytesIO
    from openpyxl import load_workbook
    from tests.test_control_consolidado_diagnostics import r, rep, D, E
    from kiki_control.exporting import generar_reporte_consolidado_excel
    from kiki_control.presentation.control_consolidado_diagnostics import diagnosticar_control_consolidado
    from kiki_control.ui.streamlit_app import _fila_temporal

    reporte = rep([
        r('fin:distinto:hash:fila:1', E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D('100'), neto_fin=D('-20'), costo=None, dif=None, tiene_ml=False, tiene_ec=False, filas_mp=(1,)),
    ])
    item = diagnosticar_control_consolidado(reporte, None, None, {1: None}).temporal_mp_sin_venta.sin_fecha
    assert _fila_temporal('Sin fecha', item)['Neto aprobado MP'] == '$ 100,00'
    assert _fila_temporal('Sin fecha', item)['Neto financiero total MP'] == '$ -20,00'

    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(reporte)))
    ws = wb['Distribución temporal MP']
    row = next(row for row in ws.iter_rows(min_row=2, values_only=False) if row[0].value == 'Sin fecha')
    assert row[2].value == D('100')
    assert row[3].value == D('-20')

def test_excel_consolidado_usa_mismo_diagnostico_temporal_que_interfaz():
    from datetime import date
    from io import BytesIO
    from openpyxl import load_workbook
    from tests.test_control_consolidado_diagnostics import r, rep, D, E
    from kiki_control.exporting import generar_reporte_consolidado_excel
    from kiki_control.presentation.control_consolidado_diagnostics import diagnosticar_control_consolidado
    from kiki_control.ui.streamlit_app import _fila_temporal

    reporte = rep([
        r('fin:anterior:hash:fila:1', E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D('10'), neto_fin=D('-1'), costo=None, dif=None, tiene_ml=False, tiene_ec=False, filas_mp=(1,)),
        r('fin:dentro:hash:fila:2', E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D('20'), neto_fin=D('-2'), costo=None, dif=None, tiene_ml=False, tiene_ec=False, filas_mp=(2,)),
        r('fin:posterior:hash:fila:3', E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D('30'), neto_fin=D('-3'), costo=None, dif=None, tiene_ml=False, tiene_ec=False, filas_mp=(3,)),
        r('fin:sin-fecha:hash:fila:4', E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D('40'), neto_fin=D('-4'), costo=None, dif=None, tiene_ml=False, tiene_ec=False, filas_mp=(4,)),
        r('fin:mixto:hash:fila:5', E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D('50'), neto_fin=D('-5'), costo=None, dif=None, tiene_ml=False, tiene_ec=False, filas_mp=(5, 6)),
    ])
    diagnostico = diagnosticar_control_consolidado(
        reporte,
        date(2026, 7, 10),
        date(2026, 7, 20),
        {1: date(2026, 7, 1), 2: date(2026, 7, 15), 3: date(2026, 7, 30), 4: None, 5: date(2026, 7, 1), 6: date(2026, 7, 30)},
    )
    interfaz = {
        'Anteriores': _fila_temporal('Anteriores', diagnostico.temporal_mp_sin_venta.anteriores),
        'Dentro': _fila_temporal('Dentro', diagnostico.temporal_mp_sin_venta.dentro),
        'Posteriores': _fila_temporal('Posteriores', diagnostico.temporal_mp_sin_venta.posteriores),
        'Sin fecha': _fila_temporal('Sin fecha', diagnostico.temporal_mp_sin_venta.sin_fecha),
        'Fechas mixtas': _fila_temporal('Fechas mixtas', diagnostico.temporal_mp_sin_venta.fechas_mixtas),
    }
    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(reporte, diagnostico=diagnostico)))
    excel = {row[0].value: row for row in wb['Distribución temporal MP'].iter_rows(min_row=2, values_only=False)}
    for categoria in interfaz:
        assert excel[categoria][1].value == interfaz[categoria]['Cantidad']
        assert excel[categoria][2].value == D(interfaz[categoria]['Neto aprobado MP'].replace('$ ', '').replace('.', '').replace(',', '.'))
        assert excel[categoria][3].value == D(interfaz[categoria]['Neto financiero total MP'].replace('$ ', '').replace('.', '').replace(',', '.'))
    assert {excel[c][1].value for c in interfaz} == {1}


def test_excel_residual_muestra_universo_sumas_y_no_formatea_cantidades_como_moneda():
    from io import BytesIO
    from openpyxl import load_workbook
    from tests.test_control_consolidado_diagnostics import r, rep, D, con_componentes_ml
    from kiki_control.exporting import generar_reporte_consolidado_excel
    reporte = rep([con_componentes_ml(r('ml'), D('120'), D('150'), D('-20'), D('-10'))])
    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(reporte)))
    ws = wb['Puente de fuentes']
    filas = {row[0].value: row[1] for row in ws.iter_rows(min_row=2, values_only=False)}
    assert filas['Universo ML oficial'].value == 1
    assert '$' not in filas['Universo ML oficial'].number_format
    assert filas['Suma Total (ARS)'].value == D('120')
    assert '$' in filas['Suma Total (ARS)'].number_format
    assert filas['Identidad residual ML cierra'].value == 'Sí'
    assert '$' not in filas['Identidad residual ML cierra'].number_format
    diccionario = wb['Diccionario de cálculos']
    assert any(row[2].value == 'universo ML oficial con los cuatro importes presentes' for row in diccionario.iter_rows(min_row=2))
