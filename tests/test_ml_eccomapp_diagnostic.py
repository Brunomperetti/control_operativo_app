from dataclasses import replace
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook
import pytest

from kiki_control.domain.ml_eccomapp_diagnostic import EstadoAptitudUtilidad as Apt, EstadoCruceMlEccomapp as Estado
from kiki_control.exporting.excel import generar_diagnostico_ml_eccomapp_excel, generar_reporte_consolidado_excel
from kiki_control.linking.ml_eccomapp_diagnostic import diagnosticar_ml_eccomapp
from kiki_control.presentation.ml_eccomapp_view import (
    ETIQUETAS_APTITUD,
    ETIQUETAS_ESTADO,
    conclusion_ejecutiva_ml_eccomapp,
    filas_casos_ml_eccomapp,
)
from tests.test_commercial_linking import op, venta
from tests.test_control_consolidado import reporte


def caso(ventas, ops):
    diag = diagnosticar_ml_eccomapp(ventas, ops)
    assert len(diag.casos) == 1
    return diag.casos[0]


def test_coincidencia_exacta_uno_a_uno_y_utilidad_calculable():
    c = caso([venta("123")], [op("123")])
    assert (c.estado, c.aptitud_utilidad) == (Estado.COINCIDENCIA_EXACTA, Apt.UTILIDAD_CALCULABLE)


def test_carrito_principal_y_ordenes_detalle_forman_un_unico_caso_completo():
    ventas = [venta("C1", None, fila=1), venta("O1", "A", fila=2), venta("O2", "B", fila=3)]
    operaciones = [op("O1", "C1", "A", fila=1), op("O2", "C1", "B", fila=2)]
    d = diagnosticar_ml_eccomapp(ventas, operaciones)
    assert len(d.casos) == 1
    c = d.casos[0]
    assert c.estado == Estado.COINCIDENCIA_POR_GRUPO
    assert [v.fila_origen for v in c.ventas_ml] == [1, 2, 3]
    assert [o.numero_fila_origen for o in c.operaciones_eccomapp] == [1, 2]
    assert c.ids_venta_ml == ("C1", "O1", "O2")
    assert len({(v.hash_importacion, v.fila_origen) for v in c.ventas_ml}) == 3
    assert len({(o.hash_importacion, o.numero_fila_origen) for o in c.operaciones_eccomapp}) == 2
    assert c.aptitud_utilidad == Apt.UTILIDAD_CALCULABLE
    assert (d.cantidad_filas_ml, d.cantidad_filas_eccomapp, d.cantidad_grupos_ml, d.cantidad_grupos_eccomapp) == (3, 2, 1, 1)


def test_dos_carritos_no_mezclan_sus_componentes():
    ventas = [venta("C1", None, 1), venta("O1", "A", 2), venta("C2", None, 3), venta("O2", "B", 4)]
    operaciones = [op("O1", "C1", "A", 1), op("O2", "C2", "B", 2)]
    d = diagnosticar_ml_eccomapp(ventas, operaciones)
    assert len(d.casos) == 2
    por_grupo = {c.id_grupo: c for c in d.casos}
    assert por_grupo["C1"].ids_venta_ml == ("C1", "O1")
    assert por_grupo["C1"].ids_orden_eccomapp == ("O1",)
    assert por_grupo["C2"].ids_venta_ml == ("C2", "O2")
    assert por_grupo["C2"].ids_orden_eccomapp == ("O2",)


def test_id_orden_que_colisiona_con_carrito_de_otro_grupo_es_ambiguo():
    d = diagnosticar_ml_eccomapp([venta("X")], [op("O1", "X", fila=1), op("X", "C2", fila=2)])
    ambiguos = [c for c in d.casos if c.estado == Estado.IDENTIFICADOR_AMBIGUO]
    assert len(ambiguos) == 1
    assert ambiguos[0].operaciones_eccomapp == ()
    assert all("X" not in c.ids_venta_ml for c in d.casos if c is not ambiguos[0])


def test_solo_ml_y_solo_eccomapp_son_universos_independientes():
    d = diagnosticar_ml_eccomapp([venta("ML")], [op("EC")])
    assert (d.cantidad_solo_ml, d.cantidad_solo_eccomapp) == (1, 1)


def test_identificadores_faltantes_no_desaparecen():
    d = diagnosticar_ml_eccomapp([venta("")], [replace(op(""), id_carrito=None)])
    assert d.cantidad_identificador_incompleto == 2
    assert sum(len(c.ventas_ml) for c in d.casos) == 1
    assert sum(len(c.operaciones_eccomapp) for c in d.casos) == 1
    assert d.cantidad_grupos_ml == 0 and d.cantidad_grupos_eccomapp == 0
    assert d.cantidad_ventas_unicas_ml == 0 and d.cantidad_operaciones_unicas_eccomapp == 0


def test_duplicado_ml_inesperado_usa_la_clasificacion_del_vinculador_canonico():
    assert caso([venta("X", "A", 1), venta("X", "A", 2)], []).estado == Estado.DUPLICADO_ML


def test_duplicado_eccomapp_inesperado():
    d = diagnosticar_ml_eccomapp([venta("X")], [op("X", sku="A", fila=1), op("X", sku="A", fila=2)])
    assert all(c.estado == Estado.DUPLICADO_ECCOMAPP for c in d.casos)
    assert sum(len(c.ventas_ml) for c in d.casos) == 1
    assert sum(len(c.operaciones_eccomapp) for c in d.casos) == 2


def test_sin_costo_sin_total_y_none_se_conservan_sin_convertir_a_cero():
    sin_costo = caso([venta("A")], [replace(op("A"), costo_total_con_iva=None)])
    sin_total = caso([venta("B", total=None)], [op("B")])
    assert (sin_costo.costo_eccomapp, sin_costo.aptitud_utilidad) == (None, Apt.SIN_COSTO)
    assert (sin_total.total_ml, sin_total.aptitud_utilidad) == (None, Apt.SIN_TOTAL_ML)


def test_coherencia_total_de_filas_grupos_y_categorias():
    d = diagnosticar_ml_eccomapp([venta("A"), venta("B", fila=2)], [op("A"), op("C", fila=2)])
    assert d.cantidad_filas_ml == 2 and d.cantidad_filas_eccomapp == 2
    assert sum((d.cantidad_coincidencias_exactas, d.cantidad_coincidencias_por_grupo, d.cantidad_solo_ml, d.cantidad_solo_eccomapp, d.cantidad_identificador_incompleto, d.cantidad_ambiguas, d.cantidad_duplicadas)) == len(d.casos)


def test_identidad_de_fila_repetida_en_la_entrada_se_rechaza_explictamente():
    fila = venta("A")
    with pytest.raises(AssertionError, match="más de una vez"):
        diagnosticar_ml_eccomapp([fila, fila], [])


def test_excel_cinco_hojas_ids_extensos_formulas_seguras_y_faltantes_vacios():
    long_id = "00000000000000012345"
    d = diagnosticar_ml_eccomapp([venta(long_id), venta("=CMD()", fila=2, total=None)], [op(long_id)])
    wb = load_workbook(BytesIO(generar_diagnostico_ml_eccomapp_excel(d)), data_only=False)
    assert wb.sheetnames == ["ML-Eccomapp — Resumen", "ML sin Eccomapp", "Eccomapp sin ML", "ML-Eccomapp — Coincidencias", "ML-Eccomapp — Ambiguos"]
    ws = wb["ML-Eccomapp — Coincidencias"]
    assert ws["A2"].value == long_id and ws["A2"].data_type == "s"
    missing = wb["ML sin Eccomapp"]
    assert missing["A2"].value == "'=CMD()" and missing["N2"].value is None


@pytest.mark.parametrize("exportador", ["especifico", "consolidado"])
def test_excel_sanea_fechas_con_timezone_sin_cambiar_hora_local_y_conserva_none(exportador):
    fecha_naive = datetime(2026, 7, 20, 9, 15)
    fecha_operativa = datetime(2026, 7, 20, 11, 40, tzinfo=timezone(timedelta(hours=-3)))
    fecha_utc = datetime(2026, 7, 21, 1, 5, tzinfo=timezone.utc)

    exacto = replace(caso([venta("MATCH")], [op("MATCH")]), fecha=fecha_operativa)
    solo_ml = replace(caso([venta("ONLY-ML")], []), fecha=fecha_naive)
    solo_ec = replace(caso([], [op("ONLY-EC")]), fecha=fecha_utc)
    ambiguo = replace(caso([venta("")], []), fecha=None)
    base = diagnosticar_ml_eccomapp([venta("BASE")], [op("BASE")])
    diag = replace(base, casos=(exacto, solo_ml, solo_ec, ambiguo))

    if exportador == "especifico":
        contenido = generar_diagnostico_ml_eccomapp_excel(diag)
    else:
        contenido = generar_reporte_consolidado_excel(reporte([], [], []), diagnostico_ml_eccomapp=diag)

    wb = load_workbook(BytesIO(contenido))
    esperadas = {
        "ML-Eccomapp — Coincidencias": fecha_operativa.replace(tzinfo=None),
        "ML sin Eccomapp": fecha_naive,
        "Eccomapp sin ML": fecha_utc.replace(tzinfo=None),
        "ML-Eccomapp — Ambiguos": None,
    }
    assert "ML-Eccomapp — Resumen" in wb.sheetnames
    for hoja, esperada in esperadas.items():
        celda = wb[hoja]["F2"]
        assert celda.value == esperada
        if esperada is not None:
            assert celda.is_date


def test_rotulos_excel_explicitan_unidades_sin_alterar_metricas():
    diag = diagnosticar_ml_eccomapp([venta("A"), venta("B", fila=2)], [op("A")])
    wb = load_workbook(BytesIO(generar_diagnostico_ml_eccomapp_excel(diag)))
    valores = {row[0].value: row[1].value for row in wb["ML-Eccomapp — Resumen"].iter_rows(min_row=2)}
    assert valores["Ventas únicas ML"] == diag.cantidad_ventas_unicas_ml
    assert valores["Operaciones únicas Eccomapp"] == diag.cantidad_operaciones_unicas_eccomapp
    assert valores["Grupos comerciales con coincidencia"] == diag.cantidad_coincidencias
    assert valores["Grupos ML sin Eccomapp"] == diag.cantidad_solo_ml
    assert valores["Grupos Eccomapp sin ML"] == diag.cantidad_solo_eccomapp
    assert valores["Grupos ambiguos o incompletos"] == diag.cantidad_ambiguas + diag.cantidad_identificador_incompleto + diag.cantidad_duplicadas
    assert valores["Grupos aptos para utilidad"] == diag.cantidad_apta_utilidad


def test_conclusion_ejecutiva_distingue_ventas_operaciones_y_grupos():
    diag = diagnosticar_ml_eccomapp([venta("A"), venta("B", fila=2)], [op("A")])
    conclusion = conclusion_ejecutiva_ml_eccomapp(diag)
    assert f"{diag.cantidad_ventas_unicas_ml} ventas únicas de Mercado Libre" in conclusion
    assert f"{diag.cantidad_operaciones_unicas_eccomapp} operaciones únicas de Eccomapp" in conclusion
    assert f"{diag.cantidad_coincidencias} grupos coincidentes" in conclusion
    assert f"{diag.cantidad_apta_utilidad} pueden utilizarse para calcular utilidad" in conclusion
    assert f"{diag.cantidad_no_apta_utilidad} requieren revisión o información adicional" in conclusion
    assert "Una coincidencia agrupada puede contener varias ventas o filas de Mercado Libre y una o más operaciones de Eccomapp" in conclusion


def test_presentacion_traduce_todos_los_estados_sin_modificar_enums():
    assert set(ETIQUETAS_ESTADO) == set(Estado)
    assert set(ETIQUETAS_APTITUD) == set(Apt)
    assert Estado.SOLO_ML.value == "SOLO_ML"
    assert Apt.SIN_VINCULO_ECCOMAPP.value == "SIN_VINCULO_ECCOMAPP"
    assert ETIQUETAS_ESTADO[Estado.COINCIDENCIA_POR_GRUPO] == "Coincidencia agrupada por carrito u orden"
    assert ETIQUETAS_APTITUD[Apt.UTILIDAD_CALCULABLE] == "Apta para calcular utilidad"


def test_filas_visibles_formatean_fecha_y_reemplazan_faltantes_y_tecnicismos():
    c = replace(caso([venta("")], []), fecha=datetime(2026, 7, 20, 11, 40))
    fila = filas_casos_ml_eccomapp((c,))[0]
    assert fila["Fecha"] == "20/07/2026 11:40"
    assert fila["Grupo u orden"] == "Sin ID informado"
    assert fila["Costo Eccomapp"] == "Sin costo informado"
    assert not any("None" in str(valor) or "empty" in str(valor).lower() for valor in fila.values())


def test_solo_ml_muestra_explicacion_comercial_y_cero_con_anulacion_es_conservador():
    normal = filas_casos_ml_eccomapp((caso([venta("A")], []),))[0]
    assert normal["Motivo"] == "No se encontró una operación correspondiente en Eccomapp utilizando el ID de carrito o el ID de orden."
    assert normal["Acción recomendada"] == "Verificar si la operación fue cancelada, devuelta o excluida del archivo Eccomapp."
    venta_anulada = replace(venta("B", total=Decimal("0")), anulaciones_reembolsos=Decimal("-100"))
    anulada = filas_casos_ml_eccomapp((caso([venta_anulada], []),))[0]
    assert "No debe considerarse automáticamente una venta faltante" in anulada["Motivo"]


def test_excel_agrega_rotulos_legibles_y_conserva_valores_tecnicos_auditables():
    d = diagnosticar_ml_eccomapp([venta("A")], [])
    wb = load_workbook(BytesIO(generar_diagnostico_ml_eccomapp_excel(d)))
    ws = wb["ML sin Eccomapp"]
    assert ws["D2"].value == "Solo en Mercado Libre"
    assert ws["U2"].value == "SOLO_ML"
    assert ws["V2"].value == "SIN_VINCULO_ECCOMAPP"
