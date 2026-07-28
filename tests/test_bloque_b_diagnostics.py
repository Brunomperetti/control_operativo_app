"""Pruebas para Bloque B — Conciliación ML–MP.

Cubre los 20 casos obligatorios del requisito.
No utiliza archivos reales ni hardcodea valores de los archivos de producción.
"""

from __future__ import annotations

from dataclasses import replace
from datetime import date
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from kiki_control.domain.control_consolidado import (
    EstadoControlConsolidado as E,
    IndicadoresFinancieros,
    ReporteControlConsolidado,
    ResultadoControlConsolidado,
    TipoMovimientoFinanciero,
)
from kiki_control.domain.financial_movement import TipoOperacionFinanciera, TratamientoNetoComparable
from kiki_control.exporting.excel import (
    generar_bloque_b_mp_sin_venta_excel,
    generar_reporte_consolidado_excel,
)
from kiki_control.presentation.bloque_b_diagnostics import (
    DiagnosticoBloqueB,
    CategoriaPrincipalMpSinVenta,
    EstadoCoherenciaGrupo,
    EstadoExplicacionDiferencia,
    SubclasificacionFinanciera,
    categoria_temporal_mp,
    clasificar_diferencia,
    diagnosticar_bloque_b,
    clasificacion_normalizada_movimiento_mp,
    clasificaciones_movimientos_mp_por_fila,
    tratamientos_movimientos_mp_por_fila,
    EnriquecimientoMovimientoMpPorFila,
)
from kiki_control.presentation.control_consolidado_view import (
    TITULO_BLOQUE_B,
    filas_movimientos_bloque_b,
    filas_movimientos_diferencia,
    texto_universo_comparable,
)

D = Decimal


def _enriq(fila, id_mov, tipo, monto):
    return EnriquecimientoMovimientoMpPorFila(
        fila, id_mov, "orden-1", tipo, D(monto),
        (TratamientoNetoComparable.COMPONENTE_YA_INCLUIDO if tipo == "PAGO_ENVIO"
         else TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE),
        date(2026, 7, 2), date(2026, 7, 2), None, monto,
    )
IND = IndicadoresFinancieros(False, False, False, False, False, False, False, False)
IND_DEV = IndicadoresFinancieros(False, True, False, False, False, False, False, False)
IND_REC = IndicadoresFinancieros(False, False, True, False, False, False, False, False)
IND_ENV = IndicadoresFinancieros(False, False, False, False, False, False, False, True)


# ---------------------------------------------------------------------------
# Helpers sintéticos
# ---------------------------------------------------------------------------

def _r(
    clave: str,
    estado: E = E.COMPLETA,
    ml: Decimal | None = D("100"),
    mp: Decimal | None = D("100"),
    dif: Decimal | None = D("0"),
    neto_fin: Decimal | None = None,
    tiene_ml: bool = True,
    tiene_mp: bool = True,
    filas_mp: tuple[int, ...] = (1,),
    filas_ml: tuple[int, ...] = (1,),
    ind: IndicadoresFinancieros = IND,
    tipo: TipoMovimientoFinanciero | None = None,
    imp_dev: Decimal = D("0"),
    imp_rec: Decimal = D("0"),
    imp_env: Decimal = D("0"),
    imp_otros: Decimal = D("0"),
    tolerancia: Decimal = D("0.01"),
) -> ResultadoControlConsolidado:
    return ResultadoControlConsolidado(
        clave_resultado=clave,
        id_grupo_canonico=clave if not clave.startswith("fin:") else None,
        ids_orden=(clave,) if not clave.startswith("fin:") else (),
        tiene_mercado_libre_oficial=tiene_ml,
        tiene_eccomapp=True,
        tiene_mercado_pago=tiene_mp,
        tipo_movimiento_financiero=tipo,
        monto_venta_ml=D("120") if tiene_ml else None,
        cargo_venta_impuestos_ml=None,
        ingresos_envio_ml=None,
        costo_envio_ml=None,
        descuentos_bonificaciones_ml=None,
        anulaciones_reembolsos_ml=None,
        total_informado_ml=ml if tiene_ml else None,
        monto_venta_eccomapp_informado=None,
        costo_productos_eccomapp=D("40"),
        costo_envio_seller_eccomapp_informado=None,
        neto_mp_eccomapp_informado=None,
        utilidad_eccomapp_informada=None,
        neto_aprobado_mp=mp if tiene_mp else None,
        neto_financiero_total_mp=neto_fin if neto_fin is not None else mp,
        impacto_pagos_envio_mp=imp_env,
        impacto_devoluciones_mp=imp_dev,
        impacto_reclamos_disputas_mp=imp_rec,
        impacto_otros_mp=imp_otros,
        diferencia_venta_ml_eccomapp=None,
        diferencia_neto_ml_eccomapp=None,
        diferencia_ml_mp=dif,
        utilidad_preliminar_control=None,
        tolerancia=tolerancia,
        estado=estado,
        requiere_revision=estado != E.COMPLETA,
        motivos=(),
        explicaciones=(),
        indicadores_financieros=ind,
        version_regla="v",
        hashes_importacion_ml=(),
        hashes_importacion_eccomapp=(),
        hashes_importacion_mp=(),
        filas_origen_ml=filas_ml,
        filas_origen_eccomapp=(),
        filas_origen_mp=filas_mp,
        claves_resultados_comerciales=(),
        claves_resultados_financieros=(),
    )


def _rep(resultados: list[ResultadoControlConsolidado], tolerancia: Decimal = D("0.01")) -> ReporteControlConsolidado:
    return ReporteControlConsolidado(
        resultados=tuple(resultados),
        version_regla="v",
        tolerancia=tolerancia,
        total_resultados_comerciales_recibidos=0,
        total_resultados_financieros_recibidos=0,
        total_resultados=len(resultados),
        total_requieren_revision=sum(r.requiere_revision for r in resultados),
        total_completa=sum(r.estado == E.COMPLETA for r in resultados),
        total_con_diferencia=sum(r.estado == E.CON_DIFERENCIA for r in resultados),
        total_sin_movimiento_financiero=sum(r.estado == E.SIN_MOVIMIENTO_FINANCIERO for r in resultados),
        total_solo_movimiento_financiero=sum(r.estado == E.SOLO_MOVIMIENTO_FINANCIERO for r in resultados),
        total_sin_venta_oficial=sum(r.estado == E.SIN_VENTA_OFICIAL for r in resultados),
        total_sin_costo_producto=sum(r.estado == E.SIN_COSTO_PRODUCTO for r in resultados),
        total_en_revision_financiera=sum(r.estado == E.EN_REVISION_FINANCIERA for r in resultados),
        total_duplicada_o_ambigua=sum(r.estado == E.DUPLICADA_O_AMBIGUA for r in resultados),
        suma_total_informado_ml=D("0"),
        suma_neto_aprobado_mp=D("0"),
        suma_costo_productos_eccomapp=D("0"),
        total_total_ml_ausente=0,
    )


# ---------------------------------------------------------------------------
# Test 1 — Grupo ML–MP que coincide dentro de tolerancia
# ---------------------------------------------------------------------------

def test_grupo_ml_mp_coincide_dentro_tolerancia():
    r = _r("ok", ml=D("100"), mp=D("100.005"), dif=D("0.005"))
    reporte = _rep([r])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.resumen.comparables_totales == 1
    assert diag.resumen.coincidencias == 1
    assert diag.resumen.con_diferencia == 0
    assert len(diag.grupos_con_diferencia) == 0


def test_pago_envio_conciliado_sigue_visible_en_ui_y_excel_como_incluido():
    r = _r("envio-incluido", ml=D("8777.90"), mp=D("8777.90"), dif=D("0"),
           filas_mp=(7, 8), ind=IND_ENV, imp_env=D("2449.46"))
    diag = diagnosticar_bloque_b(
        _rep([r]),
        clasificaciones_mp_por_fila={7: "PAGO_APROBADO", 8: "PAGO_ENVIO"},
        tipos_movimiento_mp_por_fila={7: "PAGO_APROBADO", 8: "PAGO_ENVIO"},
        montos_neto_mp_por_fila={7: D("8777.90"), 8: D("2449.46")},
        tratamientos_mp_por_fila={
            7: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            8: TratamientoNetoComparable.COMPONENTE_YA_INCLUIDO,
        },
    )

    assert diag.grupos_con_diferencia == ()
    filas_ui = filas_movimientos_bloque_b(diag)
    assert filas_ui[1]["Tratamiento en neto comparable"] == "Componente ya incluido; no se suma nuevamente"

    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(_rep([r]), diag_bloque_b=diag)))
    movimientos = wb["Bloque B — Movimientos"]
    assert movimientos.max_row == 3
    assert movimientos["F3"].value == "Componente ya incluido; no se suma nuevamente"
    resumen = wb["Bloque B — Resumen"]
    assert any(row[0].value == "Tratamiento PAGO_ENVIO" for row in resumen.iter_rows(min_row=2))


# ---------------------------------------------------------------------------
# Test 2 — Grupo con diferencia positiva (MP > ML)
# ---------------------------------------------------------------------------

def test_grupo_con_diferencia_positiva():
    r = _r("pos", ml=D("100"), mp=D("110"), dif=D("10"))
    reporte = _rep([r])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.resumen.con_diferencia == 1
    assert len(diag.grupos_con_diferencia) == 1
    assert diag.grupos_con_diferencia[0].diferencia_ml_mp == D("10")
    assert diag.resumen.diferencia_operaciones_fuera_tolerancia == D("10")


# ---------------------------------------------------------------------------
# Test 3 — Grupo con diferencia negativa (MP < ML)
# ---------------------------------------------------------------------------

def test_grupo_con_diferencia_negativa():
    r = _r("neg", ml=D("100"), mp=D("85"), dif=D("-15"))
    reporte = _rep([r])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.resumen.con_diferencia == 1
    assert diag.grupos_con_diferencia[0].diferencia_ml_mp == D("-15")
    assert diag.resumen.diferencia_operaciones_fuera_tolerancia == D("-15")


# ---------------------------------------------------------------------------
# Test 4 — Dos grupos cuya suma coincide con la diferencia total
# ---------------------------------------------------------------------------

def test_suma_diferencias_individuales_igual_total():
    r1 = _r("g1", ml=D("100"), mp=D("115"), dif=D("15"))
    r2 = _r("g2", ml=D("200"), mp=D("190"), dif=D("-10"))
    reporte = _rep([r1, r2])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.resumen.con_diferencia == 2
    suma_ind = diag.grupos_con_diferencia[0].diferencia_ml_mp + diag.grupos_con_diferencia[1].diferencia_ml_mp
    assert suma_ind == diag.resumen.diferencia_operaciones_fuera_tolerancia
    assert diag.coherencia_suma_diferencias


# ---------------------------------------------------------------------------
# Test 5 — Diferencia explicada por movimientos financieros identificados
# ---------------------------------------------------------------------------

def test_diferencia_explicada_por_movimientos_financieros():
    # diferencia = 5, impacto_dev = 5 → EXPLICADA dentro de tolerancia
    r = _r("exp", ml=D("100"), mp=D("105"), dif=D("5"), imp_dev=D("5"), ind=IND_DEV)
    estado, motivo, _ = clasificar_diferencia(r)
    assert estado == EstadoExplicacionDiferencia.EXPLICADA
    assert "explicada" in motivo.lower()


# ---------------------------------------------------------------------------
# Test 6 — Temporalidad genera indicio pero no explica monetariamente
# ---------------------------------------------------------------------------

def test_indicio_temporal_no_explica_importe():
    # liquidación posterior a venta pero sin impactos que igualen la diferencia
    r = _r("temp", ml=D("100"), mp=D("108"), dif=D("8"))
    fecha_venta = date(2026, 7, 10)
    fecha_liq = date(2026, 7, 25)
    estado, motivo, secundarios = clasificar_diferencia(r, fecha_venta_ml=fecha_venta, fecha_min_liquidacion=fecha_liq)
    assert estado == EstadoExplicacionDiferencia.INDICIO_TEMPORAL
    assert "Liquidación posterior" in " ".join(secundarios)


# ---------------------------------------------------------------------------
# Test 7 — Diferencia pendiente de clasificación
# ---------------------------------------------------------------------------

def test_diferencia_pendiente_de_clasificacion():
    r = _r("pend", ml=D("100"), mp=D("112"), dif=D("12"))
    estado, motivo, _ = clasificar_diferencia(r)
    assert estado == EstadoExplicacionDiferencia.PENDIENTE_DE_CLASIFICACION


# ---------------------------------------------------------------------------
# Test 8 — Movimiento MP anterior al período ML
# ---------------------------------------------------------------------------

def test_movimiento_mp_anterior_al_periodo_ml():
    r = _r("fin:ant:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("50"), tiene_ml=False, filas_mp=(1,))
    reporte = _rep([r])
    fechas = {1: date(2026, 6, 15)}
    diag = diagnosticar_bloque_b(reporte, inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31), fechas_origen_mp_por_fila=fechas)
    assert diag.movimientos_mp_sin_venta[0].categoria_temporal == "Anterior al período ML"


# ---------------------------------------------------------------------------
# Test 9 — Movimiento MP dentro del período ML
# ---------------------------------------------------------------------------

def test_movimiento_mp_dentro_del_periodo_ml():
    r = _r("fin:den:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("50"), tiene_ml=False, filas_mp=(1,))
    reporte = _rep([r])
    fechas = {1: date(2026, 7, 15)}
    diag = diagnosticar_bloque_b(reporte, inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31), fechas_origen_mp_por_fila=fechas)
    assert diag.movimientos_mp_sin_venta[0].categoria_temporal == "Dentro del período ML"


# ---------------------------------------------------------------------------
# Test 10 — Movimiento MP con liquidación posterior
# ---------------------------------------------------------------------------

def test_movimiento_mp_con_liquidacion_posterior():
    # Un grupo comparable (no SOLO_MOVIMIENTO_FINANCIERO) con liquidación futura
    r = _r("liq-post", ml=D("100"), mp=D("102"), dif=D("2"))
    fecha_venta = date(2026, 7, 10)
    fecha_liq = date(2026, 8, 5)
    estado, _, secundarios = clasificar_diferencia(r, fecha_venta_ml=fecha_venta, fecha_min_liquidacion=fecha_liq)
    assert "Liquidación posterior" in " ".join(secundarios)


# ---------------------------------------------------------------------------
# Test 11 — Movimiento sin fecha
# ---------------------------------------------------------------------------

def test_movimiento_mp_sin_fecha():
    r = _r("fin:sf:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("30"), tiene_ml=False, filas_mp=(5,))
    reporte = _rep([r])
    diag = diagnosticar_bloque_b(reporte, inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31), fechas_origen_mp_por_fila={})
    assert diag.movimientos_mp_sin_venta[0].categoria_temporal == "Sin fecha"
    assert diag.movimientos_mp_sin_venta[0].fecha_min_origen == "Sin fecha"


# ---------------------------------------------------------------------------
# Test 12 — Grupo con fechas mixtas
# ---------------------------------------------------------------------------

def test_grupo_con_fechas_mixtas():
    filas_mp = (1, 2)
    r = _r("fin:mix:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("30"), tiene_ml=False, filas_mp=filas_mp)
    reporte = _rep([r])
    fechas = {1: date(2026, 6, 20), 2: date(2026, 7, 15)}
    diag = diagnosticar_bloque_b(
        reporte,
        inicio_ml=date(2026, 7, 1),
        fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila=fechas,
    )
    assert diag.movimientos_mp_sin_venta[0].categoria_temporal == "Fechas mixtas"


# ---------------------------------------------------------------------------
# Test 13 — Payout separado de ventas faltantes
# ---------------------------------------------------------------------------

def test_payout_separado_de_ventas_faltantes():
    r = _r(
        "fin:pay:h:1",
        E.SOLO_MOVIMIENTO_FINANCIERO,
        ml=None,
        mp=D("200"),
        tiene_ml=False,
        tipo=TipoMovimientoFinanciero.MOVIMIENTO_DE_FONDOS,
    )
    reporte = _rep([r])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.cantidad_mp_sin_venta == 0
    assert diag.neto_aprobado_mp_sin_venta == D("0")
    assert diag.cantidad_movimientos_fondos == 1
    mov = diag.movimientos_fondos[0]
    assert "Payout" in mov.motivo_sin_venta or "movimiento de fondos" in mov.motivo_sin_venta.lower()


# ---------------------------------------------------------------------------
# Test 14 — Listado MP sin venta ML
# ---------------------------------------------------------------------------

def test_listado_mp_sin_venta_ml():
    r1 = _r("fin:a:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("50"), tiene_ml=False, filas_mp=(1,))
    r2 = _r("fin:b:h:2", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("70"), tiene_ml=False, filas_mp=(2,))
    r3 = _r("con-ml", ml=D("100"), mp=D("100"), dif=D("0"))
    reporte = _rep([r1, r2, r3])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.cantidad_mp_sin_venta == 2
    # Sin importes por fila el agregado queda solo como referencia: el KPI no
    # reutiliza silenciosamente 50 + 70.
    assert diag.neto_aprobado_mp_sin_venta == D("0")
    assert diag.neto_financiero_total_mp_sin_venta == D("0")
    assert not diag.coherencia_detalle_importes_mp_sin_venta
    assert len(diag.movimientos_mp_sin_venta) == 2


# ---------------------------------------------------------------------------
# Test 15 — Exportación Excel de diferencias
# ---------------------------------------------------------------------------

def test_exportacion_excel_diferencias():
    r = _r("dif-excel", ml=D("100"), mp=D("125"), dif=D("25"))
    reporte = _rep([r])
    diag_b = diagnosticar_bloque_b(reporte)
    bytes_xlsx = generar_reporte_consolidado_excel(reporte, diag_bloque_b=diag_b)
    assert bytes_xlsx.startswith(b"PK")
    wb = load_workbook(BytesIO(bytes_xlsx))
    assert "Bloque B — Diferencias" in wb.sheetnames
    assert "Bloque B — Resumen" in wb.sheetnames
    ws_dif = wb["Bloque B — Diferencias"]
    # Header row + 1 data row
    assert ws_dif.max_row == 2


# ---------------------------------------------------------------------------
# Test 16 — Exportación Excel de MP sin venta ML
# ---------------------------------------------------------------------------

def test_exportacion_excel_mp_sin_venta():
    r1 = _r("fin:z:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("80"), tiene_ml=False, filas_mp=(1,))
    r2 = _r("fin:y:h:2", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("30"), tiene_ml=False, filas_mp=(2,))
    reporte = _rep([r1, r2])
    diag_b = diagnosticar_bloque_b(reporte)
    bytes_xlsx = generar_bloque_b_mp_sin_venta_excel(diag_b)
    assert bytes_xlsx.startswith(b"PK")
    wb = load_workbook(BytesIO(bytes_xlsx))
    assert "MP sin venta ML" in wb.sheetnames
    ws = wb["MP sin venta ML"]
    assert ws.max_row == 3  # header + 2 data rows


# ---------------------------------------------------------------------------
# Test 17 — Coherencia de conteos: comparables = coincidentes + con diferencia
# ---------------------------------------------------------------------------

def test_coherencia_conteos_comparables():
    r_ok = _r("ok", ml=D("100"), mp=D("100"), dif=D("0"))
    r_dif = _r("dif", ml=D("100"), mp=D("120"), dif=D("20"))
    reporte = _rep([r_ok, r_dif])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.resumen.comparables_totales == diag.resumen.coincidencias + diag.resumen.con_diferencia
    assert diag.resumen.comparables_totales == 2
    assert diag.resumen.coincidencias == 1
    assert diag.resumen.con_diferencia == 1


# ---------------------------------------------------------------------------
# Test 18 — Suma individual de diferencias = diferencia total
# ---------------------------------------------------------------------------

def test_suma_individual_igual_diferencia_total():
    r1 = _r("x1", ml=D("100"), mp=D("118"), dif=D("18"))
    r2 = _r("x2", ml=D("200"), mp=D("192"), dif=D("-8"))
    reporte = _rep([r1, r2])
    diag = diagnosticar_bloque_b(reporte)
    assert abs(diag.suma_diferencias_individuales - diag.resumen.diferencia_operaciones_fuera_tolerancia) <= D("0.01")
    assert diag.coherencia_suma_diferencias


# ---------------------------------------------------------------------------
# Test 19 — La UI no muestra "Identidad cierra: Sí" sin identificar el universo
# ---------------------------------------------------------------------------

def test_texto_universo_comparable_no_dice_identidad_cierra_sin_contexto():
    r_ok = _r("ok", ml=D("100"), mp=D("100"), dif=D("0"))
    r_dif = _r("dif", ml=D("100"), mp=D("124"), dif=D("24"))
    reporte = _rep([r_ok, r_dif])
    diag = diagnosticar_bloque_b(reporte)
    texto = texto_universo_comparable(diag)
    # El texto debe mencionar la cantidad de operaciones pendientes y el importe
    assert "1 operaciones por analizar" in texto or "1 operación" in texto or "quedan" in texto.lower()
    # No debe solo decir "Identidad cierra: Sí"
    assert "identidad cierra: sí" not in texto.lower()
    # Debe identificar el universo
    assert "1" in texto  # al menos menciona que hay operaciones


# ---------------------------------------------------------------------------
# Test 20 — Bloque A no se modifica
# ---------------------------------------------------------------------------

def test_bloque_a_no_modificado():
    """El diagnóstico de Bloque B no afecta los resultados del Bloque A."""
    from kiki_control.presentation.control_consolidado_diagnostics import diagnosticar_control_consolidado

    r1 = _r("a", ml=D("100"), mp=D("100"), dif=D("0"))
    r2 = _r("b", ml=D("200"), mp=D("220"), dif=D("20"))
    reporte = _rep([r1, r2])

    diag_consolidado = diagnosticar_control_consolidado(reporte)
    diag_b = diagnosticar_bloque_b(reporte)

    # Bloque A residual no cambia
    assert diag_consolidado.residual_ml.suma_total_ars == D("300")  # 100 + 200

    # Bloque B no interfiere con partición de Bloque A
    assert diag_consolidado.particion.total_resultados == 2
    assert diag_b.resumen.comparables_totales == 2  # independiente


# ---------------------------------------------------------------------------
# Test adicional — INDICIO_FINANCIERO cuando hay reclamos/devoluciones
# ---------------------------------------------------------------------------

def test_indicio_financiero_con_reclamo():
    r = _r("rec", ml=D("100"), mp=D("107"), dif=D("7"), ind=IND_REC)
    estado, motivo, secundarios = clasificar_diferencia(r)
    assert estado == EstadoExplicacionDiferencia.INDICIO_FINANCIERO
    assert "Reclamo detectado" in secundarios


def test_indicio_financiero_con_devolucion():
    # devolucion presente pero importe no iguala diferencia → INDICIO_FINANCIERO
    r = _r("dev-ind", ml=D("100"), mp=D("107"), dif=D("7"), imp_dev=D("3"), ind=IND_DEV)
    estado, _, _ = clasificar_diferencia(r)
    assert estado == EstadoExplicacionDiferencia.INDICIO_FINANCIERO


def test_categoria_temporal_funcion_directa():
    filas = (1, 2, 3)
    fechas = {
        1: date(2026, 7, 15),
        2: date(2026, 7, 20),
        3: date(2026, 7, 25),
    }
    cat = categoria_temporal_mp(filas, fechas, inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31))
    assert cat == "Dentro del período ML"


def test_categoria_temporal_posterior():
    filas = (1,)
    fechas = {1: date(2026, 8, 5)}
    cat = categoria_temporal_mp(filas, fechas, inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31))
    assert cat == "Posterior al período ML"


def test_bloque_b_sin_enriquecimiento_funciona_sin_error():
    """Con datos mínimos (sin enriquecimiento) el diagnóstico no falla."""
    r1 = _r("ok", ml=D("100"), mp=D("100"), dif=D("0"))
    r2 = _r("fin:x:h:1", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("50"), tiene_ml=False)
    reporte = _rep([r1, r2])
    diag = diagnosticar_bloque_b(reporte)
    assert diag.resumen.comparables_totales == 1
    assert diag.cantidad_mp_sin_venta == 1
    assert diag.movimientos_mp_sin_venta[0].fecha_min_origen == "Sin fecha"


def test_titulo_bloque_b():
    assert "Bloque B" in TITULO_BLOQUE_B
    assert "Conciliación" in TITULO_BLOQUE_B


def test_exportacion_sin_diferencias_no_crea_hoja_con_datos():
    """Si no hay diferencias, la hoja de diferencias tiene solo el header."""
    r = _r("ok", ml=D("100"), mp=D("100"), dif=D("0"))
    reporte = _rep([r])
    diag_b = diagnosticar_bloque_b(reporte)
    bytes_xlsx = generar_reporte_consolidado_excel(reporte, diag_bloque_b=diag_b)
    wb = load_workbook(BytesIO(bytes_xlsx))
    ws = wb["Bloque B — Diferencias"]
    assert ws.max_row == 1  # solo el header


def test_no_float_en_bloque_b_diagnostics():
    """Verificar que bloque_b_diagnostics.py no usa float."""
    source = open("src/kiki_control/presentation/bloque_b_diagnostics.py", encoding="utf-8").read()
    assert "float(" not in source


def test_tres_diferencias_corresponden_a_sus_universos():
    dentro = _r("dentro", ml=D("100"), mp=D("100.005"), dif=D("0.005"))
    fuera = _r("fuera", ml=D("100"), mp=D("110"), dif=D("10"))
    diag = diagnosticar_bloque_b(_rep([dentro, fuera]))
    assert diag.resumen.diferencia_universo_comparable == D("10.005")
    assert diag.resumen.diferencia_operaciones_fuera_tolerancia == D("10")
    assert diag.resumen.diferencia_subuniverso_conciliado == D("0.005")


def test_detalle_movimientos_y_hojas_separadas():
    r = _r("dif-detalle", ml=D("100"), mp=D("110"), dif=D("10"), filas_mp=(7, 8))
    diag = diagnosticar_bloque_b(
        _rep([r]),
        ids_operacion_mp_por_fila={7: "mov-7", 8: "mov-8"},
        ids_orden_mp_por_fila={7: "orden", 8: "orden"},
        tipos_movimiento_mp_por_fila={7: "PAGO", 8: "DEVOLUCION"},
        clasificaciones_mp_por_fila={7: "APROBADO", 8: "APROBADO"},
        fechas_origen_mp_por_fila={7: date(2026, 7, 1)},
        fechas_aprobacion_mp_por_fila={7: date(2026, 7, 2), 8: date(2026, 7, 3)},
        fechas_liquidacion_mp_por_fila={7: date(2026, 7, 4), 8: None},
        montos_neto_mp_por_fila={7: D("120"), 8: D("-10")},
    )
    movimientos = diag.grupos_con_diferencia[0].movimientos_asociados
    assert len(movimientos) == 2
    assert movimientos[0].fecha_origen == "01/07/2026"
    assert movimientos[1].fecha_origen == "Sin fecha"
    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(_rep([r]), diag_bloque_b=diag)))
    assert {"Bloque B — Movimientos", "Bloque B — Fondos y payouts"}.issubset(wb.sheetnames)
    assert wb["Bloque B — Movimientos"].max_row == 3


@pytest.mark.parametrize(
    ("tipo", "esperado"),
    [
        (TipoOperacionFinanciera.PAGO_APROBADO, "PAGO_APROBADO"),
        (TipoOperacionFinanciera.RECLAMO, "RECLAMO"),
        (TipoOperacionFinanciera.PAYOUT, "PAYOUT"),
    ],
)
def test_clasificacion_real_del_movimiento_mp(tipo, esperado):
    movimiento = SimpleNamespace(tipo_operacion=tipo)
    assert clasificacion_normalizada_movimiento_mp(movimiento) == esperado


def test_tratamiento_financiero_se_propaga_tipado_por_fila():
    movimientos = (
        SimpleNamespace(numero_fila_origen=2, tratamiento_neto_comparable=TratamientoNetoComparable.COMPONENTE_YA_INCLUIDO),
        SimpleNamespace(numero_fila_origen=3, tratamiento_neto_comparable=TratamientoNetoComparable.MOVIMIENTO_DE_FONDOS),
    )
    assert tratamientos_movimientos_mp_por_fila(movimientos) == {
        2: TratamientoNetoComparable.COMPONENTE_YA_INCLUIDO,
        3: TratamientoNetoComparable.MOVIMIENTO_DE_FONDOS,
    }


def test_clasificacion_movimiento_mp_solo_usa_fallback_si_esta_ausente():
    assert clasificacion_normalizada_movimiento_mp(SimpleNamespace(tipo_operacion=None)) == "Sin clasificación"


def test_enriquecimiento_indexa_clasificaciones_reales_por_fila():
    movimientos = (
        SimpleNamespace(numero_fila_origen=2, tipo_operacion=TipoOperacionFinanciera.PAGO_APROBADO),
        SimpleNamespace(numero_fila_origen=3, tipo_operacion=TipoOperacionFinanciera.DEVOLUCION_DINERO),
        SimpleNamespace(numero_fila_origen=4, tipo_operacion=TipoOperacionFinanciera.PAYOUT),
        SimpleNamespace(numero_fila_origen=5, tipo_operacion=None),
    )
    assert clasificaciones_movimientos_mp_por_fila(movimientos) == {
        2: "PAGO_APROBADO",
        3: "DEVOLUCION_DINERO",
        4: "PAYOUT",
        5: "Sin clasificación",
    }


def test_ui_y_excel_conservan_clasificaciones_distintas():
    r = _r("dif-estados", ml=D("100"), mp=D("110"), dif=D("10"), filas_mp=(7, 8, 9))
    diag = diagnosticar_bloque_b(
        _rep([r]),
        clasificaciones_mp_por_fila={7: "PAGO_APROBADO", 8: "RECLAMO", 9: "PAYOUT"},
    )
    grupo = diag.grupos_con_diferencia[0]
    assert [fila["Clasificación normalizada"] for fila in filas_movimientos_diferencia(grupo)] == [
        "PAGO_APROBADO", "RECLAMO", "PAYOUT"
    ]
    wb = load_workbook(BytesIO(generar_reporte_consolidado_excel(_rep([r]), diag_bloque_b=diag)))
    hoja_movimientos = wb["Bloque B — Movimientos"]
    assert hoja_movimientos["E1"].value == "Clasificación normalizada"
    clasificaciones_excel = [celda.value for celda in hoja_movimientos["E"][1:]]
    assert clasificaciones_excel == ["PAGO_APROBADO", "RECLAMO", "PAYOUT"]

# Diagnóstico auditable de MP sin venta ML

def test_clasificacion_principal_resumen_filtros_y_vinculacion():
    from kiki_control.presentation.bloque_b_diagnostics import CategoriaPrincipalMpSinVenta, SubclasificacionFinanciera
    from kiki_control.presentation.control_consolidado_view import filtrar_mp_sin_venta
    casos = [
        _r("ant", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("10"), tiene_ml=False, filas_mp=(10,), ind=IND_REC),
        _r("dentro", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("20"), tiene_ml=False, filas_mp=(20, 21)),
        _r("post", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("30"), tiene_ml=False, filas_mp=(30,)),
        _r("sin-fecha", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("40"), tiene_ml=False, filas_mp=(40,)),
        _r("payout", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("999"), tiene_ml=False, filas_mp=(50,), tipo=TipoMovimientoFinanciero.MOVIMIENTO_DE_FONDOS),
    ]
    diag = diagnosticar_bloque_b(
        _rep(casos), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={10: date(2026, 6, 1), 20: date(2026, 7, 2), 21: date(2026, 7, 3), 30: date(2026, 8, 1), 50: date(2026, 7, 1)},
        tipos_movimiento_mp_por_fila={10: "RECLAMO", 20: "PAGO_APROBADO", 21: "DEVOLUCION", 30: "PAGO_APROBADO", 40: "DESCONOCIDO"},
        ids_operacion_mp_por_fila={10: "=riesgo", 20: "m20", 21: "m21", 30: "m30", 40: "m40"},
        ids_orden_mp_por_fila={20: "orden-20"},
        montos_neto_mp_por_fila={10: D("10"), 20: D("30"), 21: D("-10"), 30: D("30"), 40: D("40")},
    )
    assert diag.cantidad_mp_sin_venta == 4
    assert diag.cantidad_movimientos_fondos == 1
    assert diag.coherencia_mp_sin_venta
    assert sum(x.cantidad_grupos for x in diag.resumen_mp_sin_venta) == 4
    assert sum(x.neto_financiero_total for x in diag.resumen_mp_sin_venta) == D("100")
    assert {m.categoria_principal for m in diag.movimientos_mp_sin_venta} == set(CategoriaPrincipalMpSinVenta)
    dentro = next(m for m in diag.movimientos_mp_sin_venta if m.id_grupo == "dentro")
    assert dentro.subclasificacion_financiera == SubclasificacionFinanciera.MULTIPLES_TIPOS
    assert dentro.tiene_id_orden_utilizable and dentro.cantidad_movimientos == 2
    # El checkbox ya no mezcla casos temporales u otros movimientos: solo alta.
    assert len(filtrar_mp_sin_venta(diag.movimientos_mp_sin_venta, solo_prioritarios=True)) == 0


def test_excel_mp_sin_venta_incluye_resumen_detalle_y_seguridad():
    r = _r("grupo", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("12.34"), tiene_ml=False, filas_mp=(7,))
    diag = diagnosticar_bloque_b(
        _rep([r]), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={7: date(2026, 7, 2)}, ids_operacion_mp_por_fila={7: "=1+1"},
        tipos_movimiento_mp_por_fila={7: "PAGO_APROBADO"}, montos_neto_mp_por_fila={7: D("12.34")},
    )
    wb = load_workbook(BytesIO(generar_bloque_b_mp_sin_venta_excel(diag)))
    assert "Resumen MP sin ML" in wb.sheetnames
    ws = wb["MP sin venta ML"]
    assert ws["B2"].data_type == "s" and ws["B2"].value.startswith("'")
    headers = [c.value for c in ws[1]]
    amount = ws.cell(2, headers.index("Neto aprobado bruto MP") + 1)
    assert D(str(amount.value)) == D("12.34") and "$" in amount.number_format

@pytest.mark.parametrize(("fechas", "esperada"), [
    ({1: date(2026, 6, 30), 2: date(2026, 7, 10)}, "DENTRO_DEL_PERIODO_ML_SIN_VENTA"),
    ({1: date(2026, 7, 10), 2: date(2026, 8, 1)}, "DENTRO_DEL_PERIODO_ML_SIN_VENTA"),
    ({}, "SIN_FECHA_DE_ORIGEN"),
    ({1: date(2026, 6, 30), 2: date(2026, 8, 1)}, "DENTRO_DEL_PERIODO_ML_SIN_VENTA"),
])
def test_categoria_principal_grupos_con_fechas_limite(fechas, esperada):
    from kiki_control.presentation.bloque_b_diagnostics import categoria_principal_mp
    assert categoria_principal_mp((1, 2), fechas, date(2026, 7, 1), date(2026, 7, 31)).value == esperada


@pytest.mark.parametrize(("tipos", "sub", "combinacion"), [
    (("PAGO_APROBADO",), "PAGO_APROBADO", "NO_APLICA"),
    (("PAGO_ENVIO",), "ENVIO", "NO_APLICA"),
    (("PAGO_APROBADO", "DEVOLUCION"), "MULTIPLES_TIPOS", "PAGO + DEVOLUCIÓN"),
    (("PAGO_APROBADO", "RECLAMO"), "MULTIPLES_TIPOS", "PAGO + RECLAMO"),
    (("PAGO_ENVIO", "DEVOLUCION"), "MULTIPLES_TIPOS", "ENVÍO + DEVOLUCIÓN"),
    (("PAGO_ENVIO", "DISPUTA"), "MULTIPLES_TIPOS", "ENVÍO + DISPUTA"),
    (("CASHBACK",), "OTRO_MOVIMIENTO", "NO_APLICA"),
])
def test_clasificacion_operativa_y_combinaciones(tipos, sub, combinacion):
    from kiki_control.presentation.bloque_b_diagnostics import combinacion_resumida, subclasificar_financieramente
    assert subclasificar_financieramente(tipos).value == sub
    assert combinacion_resumida(tipos).value == combinacion


def test_resumen_operativo_posible_venta_faltante_filtros_bloque_d_y_excel():
    from kiki_control.presentation.control_consolidado_view import filtrar_mp_sin_venta, kpis_consolidados
    tipos = {
        1: "PAGO_APROBADO", 2: "PAGO_ENVIO", 3: "PAGO_APROBADO", 4: "DEVOLUCION",
        5: "CASHBACK",
    }
    casos = [
        _r("pago", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("100"), tiene_ml=False, filas_mp=(1,)),
        _r("envio", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("10"), neto_fin=D("0"), tiene_ml=False, filas_mp=(2,)),
        _r("mixto", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("80"), tiene_ml=False, filas_mp=(3, 4)),
        _r("cashback", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("5"), tiene_ml=False, filas_mp=(5,)),
    ]
    diag = diagnosticar_bloque_b(
        _rep(casos), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={i: date(2026, 7, 15) for i in tipos},
        tipos_movimiento_mp_por_fila=tipos,
        ids_orden_mp_por_fila={1: "orden-1"},
        montos_neto_mp_por_fila={1: D("100"), 2: D("10"), 3: D("100"), 4: D("-20"), 5: D("5")},
        tratamientos_mp_por_fila={
            1: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            2: TratamientoNetoComparable.COMPONENTE_YA_INCLUIDO,
            3: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            4: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            5: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
        },
    )
    assert diag.coherencia_operativa_dentro_periodo
    assert sum(r.cantidad_grupos for r in diag.resumen_operativo_dentro_periodo) == 4
    posibles = [m.id_grupo for m in diag.movimientos_mp_sin_venta if m.posible_venta_faltante]
    assert posibles == ["pago"]
    assert [m.id_grupo for m in filtrar_mp_sin_venta(diag.movimientos_mp_sin_venta, solo_prioritarios=True)] == ["pago"]
    assert [m.id_grupo for m in filtrar_mp_sin_venta(diag.movimientos_mp_sin_venta, filtro_combinacion="PAGO + DEVOLUCIÓN")] == ["mixto"]
    bloques = kpis_consolidados(_rep(casos), diag)
    bloque_d = {k.nombre: k for k in bloques["Bloque D — Calidad y pendientes"]}
    assert len(bloque_d) == 7
    assert "Duplicados o ambiguos" in bloque_d
    assert not ({"Pagos aprobados sin venta ML", "Grupos financieros mixtos", "Componentes de envío", "Otros movimientos no asociados a venta"} & bloque_d.keys())
    operativos = {k.nombre: k for k in bloques["Diagnóstico operativo MP sin venta dentro del período"]}
    assert {"Pagos aprobados puros detectados", "Candidatos válidos a venta faltante",
            "Pagos aprobados inconsistentes", "Grupos financieros mixtos",
            "Componentes de envío", "Otros movimientos no asociados a venta"} == operativos.keys()
    assert operativos["Pagos aprobados puros detectados"].valor == "1"
    assert operativos["Candidatos válidos a venta faltante"].valor == "1 · $ 100,00"
    wb = load_workbook(BytesIO(generar_bloque_b_mp_sin_venta_excel(diag)))
    headers = [c.value for c in wb["MP sin venta ML"][1]]
    for esperado in ("Prioridad operativa", "Combinación resumida", "Interpretación", "posible_venta_faltante"):
        assert esperado in headers
    assert "Composición de movimientos dentro del período ML sin venta encontrada" in {
        c.value for row in wb["Resumen MP sin ML"].iter_rows() for c in row
    }


def test_universo_comparable_611_611_0_permanece_sin_diferencia():
    """La nueva clasificación no altera la fórmula ni el universo del Bloque B."""
    casos = [_r(f"comparable-{i}", ml=D("100"), mp=D("100"), dif=D("0")) for i in range(611)]
    resumen = diagnosticar_bloque_b(_rep(casos)).resumen
    assert (resumen.comparables_totales, resumen.coincidencias, resumen.con_diferencia) == (611, 611, 0)
    assert resumen.diferencia_universo_comparable == D("0")
    from kiki_control.presentation.control_consolidado_view import formato_importe
    assert formato_importe(resumen.diferencia_universo_comparable) == "$ 0,00"


def test_composicion_productiva_sintetica_cierra_aun_con_inconsistencia_monetaria():
    resultados = []
    fechas, tipos, montos, tratamientos = {}, {}, {}, {}
    fila = 1

    def agregar(clave, movimientos, agregado):
        nonlocal fila
        filas = tuple(range(fila, fila + len(movimientos)))
        for numero, (tipo, monto, tratamiento) in zip(filas, movimientos):
            fechas[numero] = date(2026, 7, 10)
            tipos[numero] = tipo
            montos[numero] = monto
            tratamientos[numero] = tratamiento
        resultados.append(_r(clave, E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=agregado,
                              neto_fin=agregado, tiene_ml=False, filas_mp=filas))
        fila += len(movimientos)

    modifica = TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE
    incluido = TratamientoNetoComparable.COMPONENTE_YA_INCLUIDO
    for i in range(22):
        agregar(f"pago-{i}", [("PAGO_APROBADO", D("378536.64") if i == 0 else D("0"), modifica)],
                D("378536.64") if i == 0 else D("0"))
    for i in range(71):
        agregar(f"envio-{i}", [("PAGO_ENVIO", D("0"), incluido)], D("0"))
    for i in range(74):
        movimientos = [
            ("PAGO_APROBADO", D("943236.23") if i == 0 else D("0"), modifica),
            ("DEVOLUCION_DINERO", D("-942682.14") if i == 0 else D("0"), modifica),
        ]
        if i < 12:  # 74 grupos y 160 movimientos en MULTIPLES_TIPOS.
            movimientos.append(("DEVOLUCION_DINERO", D("0"), modifica))
        reconstruido = D("554.09") if i == 0 else D("0")
        agregar(f"mixto-{i}", movimientos, reconstruido + (D("1") if i == 0 else D("0")))
    agregar("otro", [("BONIFICACION", D("7892.77"), modifica)], D("7892.77"))

    diag = diagnosticar_bloque_b(
        _rep(resultados), date(2026, 7, 1), date(2026, 7, 31),
        fechas_origen_mp_por_fila=fechas,
        tipos_movimiento_mp_por_fila=tipos,
        montos_neto_mp_por_fila=montos,
        tratamientos_mp_por_fila=tratamientos,
    )
    assert len(diag.movimientos_mp_sin_venta) == 168
    assert sum(m.cantidad_movimientos for m in diag.movimientos_mp_sin_venta) == 254
    operativo = {r.subclasificacion_financiera.value: r for r in diag.resumen_operativo_dentro_periodo}
    assert operativo["PAGO_APROBADO"].cantidad_grupos == 22
    assert operativo["ENVIO"].cantidad_grupos == 71
    assert operativo["MULTIPLES_TIPOS"].cantidad_grupos == 74
    assert operativo["OTRO_MOVIMIENTO"].cantidad_grupos == 1
    assert operativo["MULTIPLES_TIPOS"].cantidad_movimientos == 160
    assert sum(r.neto_aprobado_bruto for r in diag.resumen_operativo_dentro_periodo) == D("1321772.87")
    assert sum(r.neto_financiero_total for r in diag.resumen_operativo_dentro_periodo) == D("386983.50")
    assert diag.coherencia_operativa_dentro_periodo
    assert diag.existen_grupos_monetarios_inconsistentes


def test_mp_sin_venta_reconstruye_pago_puro_y_excluye_fila_ajena():
    """El agregado amplio no puede contaminar las filas visibles del grupo."""
    r = _r("pago", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("100"),
           neto_fin=D("-900"), tiene_ml=False, filas_mp=(1,))
    diag = diagnosticar_bloque_b(
        _rep([r]), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={1: date(2026, 7, 2), 99: date(2020, 1, 1)},
        tipos_movimiento_mp_por_fila={1: "PAGO_APROBADO", 99: "DEVOLUCION_DINERO"},
        montos_neto_mp_por_fila={1: D("100"), 99: D("-1000")},
        tratamientos_mp_por_fila={
            1: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            99: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
        },
    )
    grupo = diag.movimientos_mp_sin_venta[0]
    assert grupo.cantidad_movimientos == 1
    assert grupo.neto_aprobado_mp == grupo.neto_financiero_total_mp == D("100")
    assert grupo.suma_reconstruida_movimientos_mp == D("100")
    assert grupo.diferencia_agregado_detalle_mp == D("-1000")
    assert not grupo.coherencia_grupo and not grupo.posible_venta_faltante
    assert not diag.coherencia_detalle_importes_mp_sin_venta
    assert diag.neto_financiero_total_mp_sin_venta == D("100")


def test_mp_sin_venta_varios_pagos_y_grupo_mixto_se_reconstruyen_del_detalle():
    casos = [
        _r("dividido", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("70"), tiene_ml=False, filas_mp=(1, 2)),
        _r("mixto", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("80"), tiene_ml=False, filas_mp=(3, 4)),
    ]
    diag = diagnosticar_bloque_b(
        _rep(casos), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={i: date(2026, 7, 2) for i in range(1, 5)},
        tipos_movimiento_mp_por_fila={1: "PAGO_APROBADO", 2: "PAGO_APROBADO", 3: "PAGO_APROBADO", 4: "DEVOLUCION_DINERO"},
        montos_neto_mp_por_fila={1: D("30"), 2: D("40"), 3: D("100"), 4: D("-20")},
        tratamientos_mp_por_fila={i: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE for i in range(1, 5)},
        ids_orden_mp_por_fila={1: None, 2: None, 3: "orden-3", 4: "orden-3"},
    )
    dividido, mixto = diag.movimientos_mp_sin_venta
    assert (dividido.cantidad_movimientos, dividido.neto_aprobado_mp, dividido.neto_financiero_total_mp) == (2, D("70"), D("70"))
    assert not dividido.tiene_id_orden_utilizable
    assert (mixto.neto_aprobado_mp, mixto.neto_financiero_total_mp) == (D("100"), D("80"))
    assert mixto.subclasificacion_financiera.value == "MULTIPLES_TIPOS"
    assert diag.coherencia_operativa_dentro_periodo
    assert sum(r.neto_financiero_total for r in diag.resumen_operativo_dentro_periodo) == D("150")


@pytest.mark.parametrize(("filas", "montos", "motivo"), [
    ((), {}, "No hay movimientos asociados"),
    ((1,), {1: None}, "Detalle monetario incompleto"),
])
def test_mp_sin_venta_sin_evidencia_monetaria_es_no_verificable(filas, montos, motivo):
    from kiki_control.presentation.control_consolidado_view import filas_mp_sin_venta

    r = _r("sin-evidencia", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("100"),
           tiene_ml=False, filas_mp=filas)
    diag = diagnosticar_bloque_b(
        _rep([r]), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={1: date(2026, 7, 2)},
        tipos_movimiento_mp_por_fila={1: "PAGO_APROBADO"},
        montos_neto_mp_por_fila=montos,
    )
    grupo = diag.movimientos_mp_sin_venta[0]
    assert grupo.estado_coherencia.value == "NO_VERIFICABLE"
    assert not grupo.coherencia_grupo and not grupo.posible_venta_faltante
    assert grupo.neto_aprobado_mp is None and grupo.neto_financiero_total_mp is None
    assert grupo.suma_reconstruida_movimientos_mp is None
    assert grupo.neto_financiero_agregado_original_mp == D("100")
    assert motivo in grupo.motivo_coherencia
    assert motivo in filas_mp_sin_venta((grupo,))[0]["Advertencia"]
    assert diag.neto_financiero_total_mp_sin_venta == D("0")

    wb = load_workbook(BytesIO(generar_bloque_b_mp_sin_venta_excel(diag)))
    ws = wb["MP sin venta ML"]
    headers = [cell.value for cell in ws[1]]
    assert ws.cell(2, headers.index("Estado de coherencia") + 1).value == "NO_VERIFICABLE"
    assert motivo in ws.cell(2, headers.index("Motivo de coherencia") + 1).value
    assert ws.cell(2, headers.index("Suma reconstruida desde movimientos") + 1).value is None


def test_mp_sin_venta_detalle_completo_sin_agregado_es_no_verificable():
    from kiki_control.presentation.control_consolidado_view import filas_mp_sin_venta

    motivo = ("El detalle monetario pudo reconstruirse, pero no existe agregado "
              "financiero original para verificar la coincidencia.")
    r = replace(
        _r("sin-agregado", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("100"),
           tiene_ml=False, filas_mp=(1,)),
        neto_financiero_total_mp=None,
    )
    diag = diagnosticar_bloque_b(
        _rep([r]), inicio_ml=date(2026, 7, 1), fin_ml=date(2026, 7, 31),
        fechas_origen_mp_por_fila={1: date(2026, 7, 2)},
        tipos_movimiento_mp_por_fila={1: "PAGO_APROBADO"},
        montos_neto_mp_por_fila={1: D("100")},
        tratamientos_mp_por_fila={1: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE},
    )
    grupo = diag.movimientos_mp_sin_venta[0]
    assert grupo.neto_aprobado_mp == D("100")
    assert grupo.neto_financiero_total_mp == D("100")
    assert grupo.suma_reconstruida_movimientos_mp == D("100")
    assert grupo.neto_financiero_agregado_original_mp is None
    assert grupo.diferencia_agregado_detalle_mp is None
    assert grupo.estado_coherencia.value == "NO_VERIFICABLE"
    assert not grupo.coherencia_grupo and not grupo.posible_venta_faltante
    assert grupo.motivo_coherencia == motivo

    fila_ui = filas_mp_sin_venta((grupo,))[0]
    assert fila_ui["Agregado financiero original"] == "No calculado"
    assert fila_ui["Motivo de coherencia"] == motivo
    wb = load_workbook(BytesIO(generar_bloque_b_mp_sin_venta_excel(diag)))
    ws = wb["MP sin venta ML"]
    headers = [cell.value for cell in ws[1]]
    assert ws.cell(2, headers.index("Agregado financiero original") + 1).value is None
    assert ws.cell(2, headers.index("Diferencia agregado − detalle") + 1).value is None
    assert ws.cell(2, headers.index("Suma reconstruida desde movimientos") + 1).value == 100
    assert ws.cell(2, headers.index("Estado de coherencia") + 1).value == "NO_VERIFICABLE"
    assert ws.cell(2, headers.index("Motivo de coherencia") + 1).value == motivo

    calidad = diag.calidad_monetaria_mp_sin_venta
    assert calidad.importe_reconstruido_excluido_kpi == D("100")
    assert calidad.agregado_original_referencia is None
    assert calidad.diferencia_agregado_detalle is None
    assert calidad.importe_no_verificable == D("0")


def test_calidad_incoherente_distingue_reconstruido_agregado_y_diferencia_en_ui_y_excel():
    from kiki_control.presentation.control_consolidado_view import filas_inconsistencias_mp_sin_venta

    r = _r("incoherente", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("100"),
           neto_fin=D("90"), tiene_ml=False, filas_mp=(1,))
    diag = diagnosticar_bloque_b(
        _rep([r]), date(2026, 7, 1), date(2026, 7, 31),
        fechas_origen_mp_por_fila={1: date(2026, 7, 2)},
        tipos_movimiento_mp_por_fila={1: "PAGO_APROBADO"},
        montos_neto_mp_por_fila={1: D("100")},
        tratamientos_mp_por_fila={1: TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE},
    )
    calidad = diag.calidad_monetaria_mp_sin_venta
    assert calidad.importe_reconstruido_excluido_kpi == D("100")
    assert calidad.agregado_original_referencia == D("90")
    assert calidad.diferencia_agregado_detalle == D("-10")
    assert calidad.importe_no_verificable == D("0")

    fila = filas_inconsistencias_mp_sin_venta(diag)[0]
    assert fila["Suma reconstruida"] == "$ 100,00"
    assert fila["Agregado original"] == "$ 90,00"
    assert fila["Diferencia"] == "$ -10,00"

    wb = load_workbook(BytesIO(generar_bloque_b_mp_sin_venta_excel(diag)))
    resumen = wb["Resumen MP sin ML"]
    valores = {row[0].value: row[1].value for row in resumen.iter_rows() if row[0].value}
    assert valores["Importe reconstruido excluido de KPI"] == D("100")
    assert valores["Agregado original de referencia"] == D("90")
    assert valores["Diferencia agregado − detalle"] == D("-10")


def test_no_verificable_no_reemplaza_reconstruido_ausente_con_agregado():
    r = _r("incompleto", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("90"),
           neto_fin=D("90"), tiene_ml=False, filas_mp=(1,))
    diag = diagnosticar_bloque_b(
        _rep([r]), date(2026, 7, 1), date(2026, 7, 31),
        fechas_origen_mp_por_fila={1: date(2026, 7, 2)},
        tipos_movimiento_mp_por_fila={1: "PAGO_APROBADO"},
        montos_neto_mp_por_fila={1: None},
    )
    calidad = diag.calidad_monetaria_mp_sin_venta
    assert calidad.importe_reconstruido_excluido_kpi is None
    assert calidad.agregado_original_referencia == D("90")
    assert calidad.diferencia_agregado_detalle is None
    assert calidad.importe_no_verificable is None
    assert calidad.cantidad_grupos_sin_reconstruccion == 1


def test_enriquecimiento_atomico_conserva_fila_excel_tipo_id_importe_y_envio_excluido():
    r = _r("orden-1", tiene_ml=False, ml=None, mp=D("90"), neto_fin=D("90"),
           filas_mp=(2, 3, 4), tipo=TipoMovimientoFinanciero.ORDEN_FINANCIERA)
    enriquecimientos = {
        2: _enriq(2, "pago", "PAGO_APROBADO", "100"),
        3: _enriq(3, "devolucion", "DEVOLUCION_DINERO", "-10"),
        4: _enriq(4, "envio", "PAGO_ENVIO", "7"),
    }
    diag = diagnosticar_bloque_b(_rep([r]), date(2026, 7, 1), date(2026, 7, 31),
                                 enriquecimientos_mp_por_fila=enriquecimientos)
    grupo = diag.movimientos_mp_sin_venta[0]
    assert [(d.fila_origen, d.id_movimiento_mp, d.tipo_movimiento, d.monto_neto_impactado)
            for d in grupo.movimientos_asociados] == [
                (2, "pago", "PAGO_APROBADO", D("100")),
                (3, "devolucion", "DEVOLUCION_DINERO", D("-10")),
                (4, "envio", "PAGO_ENVIO", D("7")),
            ]
    assert grupo.neto_aprobado_mp == D("100")
    assert grupo.neto_financiero_total_mp == D("90")


def test_clave_desplazada_y_pago_aprobado_negativo_son_inconsistentes_y_no_prioritarios():
    r = _r("orden-1", tiene_ml=False, ml=None, mp=D("-50"), neto_fin=D("-50"),
           filas_mp=(2,), tipo=TipoMovimientoFinanciero.ORDEN_FINANCIERA)
    diag = diagnosticar_bloque_b(_rep([r]), date(2026, 7, 1), date(2026, 7, 31),
                                 enriquecimientos_mp_por_fila={2: _enriq(3, "pago-3", "PAGO_APROBADO", "-50")})
    grupo = diag.movimientos_mp_sin_venta[0]
    detalle = grupo.movimientos_asociados[0]
    assert detalle.estado_correspondencia_fila.startswith("ESTADO_DATO_INCONSISTENTE")
    assert grupo.neto_aprobado_mp is None
    assert grupo.neto_financiero_total_mp is None
    assert not grupo.posible_venta_faltante
    assert not grupo.coherencia_grupo
    # La calidad monetaria no contamina la validación independiente de composición.
    assert diag.coherencia_operativa_dentro_periodo
    assert diag.composicion_cantidades_coherente
    assert diag.composicion_movimientos_coherente
    assert diag.composicion_neto_aprobado_coherente
    assert diag.composicion_neto_financiero_coherente
    assert diag.existen_grupos_monetarios_inconsistentes
    calidad = diag.calidad_monetaria_mp_sin_venta
    assert calidad is not None
    assert calidad.grupos_incoherentes == 1
    assert calidad.movimientos_correspondencia_inconsistente == 1
    assert calidad.pagos_aprobados_negativos == 1
    assert calidad.importe_reconstruido_confiable == D("0")
    assert calidad.importe_reconstruido_excluido_kpi is None
    assert calidad.agregado_original_referencia == D("-50")
    assert calidad.diferencia_agregado_detalle is None
    assert calidad.importe_no_verificable is None

    from kiki_control.presentation.control_consolidado_view import filas_inconsistencias_mp_sin_venta
    filas = filas_inconsistencias_mp_sin_venta(diag)
    assert len(filas) == 1
    assert filas[0]["ID de grupo"] == grupo.id_grupo
    assert filas[0]["Estado de coherencia"] == "INCOHERENTE"
    assert filas_inconsistencias_mp_sin_venta(diag, False) == filas


def test_regresion_integral_pagos_aprobados_desde_movimientos_tablas_y_excel():
    """Certifica 22/19/3 desde filas MP reales, sin precargar el diagnóstico."""
    from kiki_control.presentation.control_consolidado_view import (
        filas_candidatos_venta_faltante,
        filas_pagos_aprobados_inconsistentes,
    )

    resultados = []
    enriquecimientos = {}
    importes_validos = [D("378518.64"), *([D("1.00")] * 18)]
    for indice, importe in enumerate(importes_validos, start=1):
        id_movimiento = "=MOVIMIENTO_PELIGROSO" if indice == 1 else f"200000000000000{indice:02d}"
        id_orden = f"900000000000000{indice:02d}" if indice <= 10 else None
        resultados.append(_r(
            f"candidato-{indice}", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None,
            mp=importe, neto_fin=importe, tiene_ml=False, filas_mp=(indice,),
            tipo=TipoMovimientoFinanciero.ORDEN_FINANCIERA,
        ))
        enriquecimientos[indice] = EnriquecimientoMovimientoMpPorFila(
            indice, id_movimiento, id_orden, "PAGO_APROBADO", importe,
            TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            date(2026, 7, 10), date(2026, 7, 11), date(2026, 7, 12), str(importe),
        )

    casos_excluidos = (
        (20, "negativo", D("-10"), D("-10"), 20, "90000000000000020"),
        (21, "correspondencia", D("25"), D("25"), 999, None),
        (22, "sin-importe", D("50"), None, 22, None),
    )
    for fila, clave, agregado, detalle, fila_enriquecida, id_orden in casos_excluidos:
        resultados.append(_r(
            clave, E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=agregado,
            neto_fin=agregado, tiene_ml=False, filas_mp=(fila,),
            tipo=TipoMovimientoFinanciero.ORDEN_FINANCIERA,
        ))
        enriquecimientos[fila] = EnriquecimientoMovimientoMpPorFila(
            fila_enriquecida, f"mp-{clave}", id_orden, "PAGO_APROBADO", detalle,
            TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            date(2026, 7, 10), date(2026, 7, 11), date(2026, 7, 12),
            "" if detalle is None else str(detalle),
        )

    reporte = _rep(resultados)
    diag = diagnosticar_bloque_b(
        reporte, date(2026, 7, 1), date(2026, 7, 31),
        enriquecimientos_mp_por_fila=enriquecimientos,
    )
    pagos = diag.diagnostico_pagos_aprobados
    assert pagos is not None
    assert (len(pagos.detectados), len(pagos.candidatos_validos), len(pagos.inconsistentes)) == (22, 19, 3)
    assert pagos.importe_valido_candidatos == D("378536.64")
    assert pagos.no_candidatos_importe_no_positivo == ()
    assert (pagos.detectados_con_id, pagos.detectados_sin_id) == (11, 11)
    assert (pagos.candidatos_con_id, pagos.candidatos_sin_id) == (10, 9)
    assert pagos.detectados_con_id + pagos.detectados_sin_id == 22
    assert pagos.candidatos_con_id + pagos.candidatos_sin_id == 19
    assert all(
        m.categoria_principal == CategoriaPrincipalMpSinVenta.DENTRO_DEL_PERIODO_ML_SIN_VENTA
        and m.subclasificacion_financiera == SubclasificacionFinanciera.PAGO_APROBADO
        and m.estado_coherencia == EstadoCoherenciaGrupo.COHERENTE
        and m.suma_reconstruida_movimientos_mp is not None
        and m.suma_reconstruida_movimientos_mp > D("0")
        and m.posible_venta_faltante
        and all(d.estado_correspondencia_fila == "CORRESPONDENCIA_OK" for d in m.movimientos_asociados)
        for m in pagos.candidatos_validos
    )
    assert not any(m.posible_venta_faltante for m in pagos.inconsistentes)
    assert {m.id_grupo for m in pagos.inconsistentes} == {"negativo", "correspondencia", "sin-importe"}
    assert "22 grupos" in pagos.conclusion_ejecutiva
    assert "19 cumplen" in pagos.conclusion_ejecutiva
    assert "$ 378.536,64" in pagos.conclusion_ejecutiva
    assert "3 casos restantes" in pagos.conclusion_ejecutiva

    candidatos = filas_candidatos_venta_faltante(diag)
    inconsistentes = filas_pagos_aprobados_inconsistentes(diag)
    assert len(candidatos) == 19
    assert len(inconsistentes) == 3
    assert {f["ID de grupo"] for f in candidatos} == {f"candidato-{i}" for i in range(1, 20)}
    assert not ({"negativo", "correspondencia", "sin-importe"} & {f["ID de grupo"] for f in candidatos})
    assert set(candidatos[0]) == {
        "ID de grupo", "ID movimiento MP", "ID de orden", "Fila original MP",
        "Fecha de origen", "Fecha de aprobación", "Fecha de liquidación", "Importe crudo",
        "Importe normalizado", "Neto reconstruido", "Estado de correspondencia",
        "Estado monetario", "Motivo de posible venta faltante", "Acción recomendada",
    }
    motivos = {f["ID movimiento"]: f["Motivo de exclusión"] for f in inconsistentes}
    assert "filas" in motivos["mp-negativo"]
    assert "filas" in motivos["mp-correspondencia"]
    assert "falta monto_neto_impactado" in motivos["mp-sin-importe"]

    for contenido in (
        generar_bloque_b_mp_sin_venta_excel(diag),
        generar_reporte_consolidado_excel(reporte, diag_bloque_b=diag),
    ):
        wb = load_workbook(BytesIO(contenido), data_only=False)
        assert {"Candidatos venta faltante", "Pagos MP inconsistentes"} <= set(wb.sheetnames)
        ws_c = wb["Candidatos venta faltante"]
        ws_i = wb["Pagos MP inconsistentes"]
        assert [c.value for c in ws_c[1]] == [
            "ID de grupo", "ID movimiento MP", "ID de orden", "Fila original MP",
            "Fecha de origen", "Fecha de aprobación", "Fecha de liquidación", "Importe crudo",
            "Importe normalizado", "Neto reconstruido", "Estado de correspondencia",
            "Estado monetario", "Motivo de posible venta faltante", "Acción recomendada",
        ]
        assert [c.value for c in ws_i[1]] == [
            "Fila original", "ID movimiento", "ID orden", "Importe",
            "Motivo de exclusión", "Estado", "Acción recomendada",
        ]
        assert ws_c.max_row - 1 == 19
        assert ws_i.max_row - 1 == 3
        assert isinstance(ws_c.cell(2, 9).value, (int, float, Decimal))
        assert isinstance(ws_c.cell(2, 10).value, (int, float, Decimal))
        assert ws_c.cell(2, 2).data_type == "s"
        assert ws_c.cell(2, 2).value == "'=MOVIMIENTO_PELIGROSO"
        assert ws_c.cell(2, 3).data_type == "s"
        fila_sin_importe = next(row for row in ws_i.iter_rows(min_row=2) if row[1].value == "mp-sin-importe")
        assert fila_sin_importe[3].value is None


def test_pago_coherente_cero_es_no_candidato_pero_no_inconsistente():
    """Opción A: un cero coherente tiene categoría propia y no contamina inconsistencias."""
    r = _r("pago-cero", E.SOLO_MOVIMIENTO_FINANCIERO, ml=None, mp=D("0"),
           neto_fin=D("0"), tiene_ml=False, filas_mp=(1,))
    diag = diagnosticar_bloque_b(
        _rep([r]), date(2026, 7, 1), date(2026, 7, 31),
        enriquecimientos_mp_por_fila={1: EnriquecimientoMovimientoMpPorFila(
            1, "mp-cero", None, "PAGO_APROBADO", D("0"),
            TratamientoNetoComparable.MODIFICA_NETO_COMPARABLE,
            date(2026, 7, 2), date(2026, 7, 2), date(2026, 7, 3), "0",
        )},
    )
    pagos = diag.diagnostico_pagos_aprobados
    assert pagos is not None
    assert len(pagos.detectados) == 1
    assert pagos.candidatos_validos == ()
    assert pagos.inconsistentes == ()
    assert [m.id_grupo for m in pagos.no_candidatos_importe_no_positivo] == ["pago-cero"]
    assert pagos.importe_valido_candidatos == D("0")
